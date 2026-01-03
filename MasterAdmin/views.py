"""
MASTER ADMIN Views - Complete Dashboard and CRUD Operations
Provides unrestricted access to all system data for super administrator.
"""
from rest_framework import status, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.decorators import action
from django.db.models import Sum, Count, Avg, F, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.apps import apps
from datetime import timedelta
import logging
import os

from .authentication import (
    MasterAdminAuthentication,
    IsMasterAdmin,
    validate_master_admin_credentials,
    create_master_admin_session,
    invalidate_master_admin_session,
)

from accounts.models import CustomUser, LeadProfile, FFInvitation, FreightForwarderProfile
from SalesModule.models import (
    LeadCotizacion, Shipment, ShipmentTracking, PreLiquidation,
    FreightRate, InsuranceRate, CustomsDutyRate, 
    InlandTransportQuoteRate, CustomsBrokerageRate,
    QuoteScenario, QuoteLineItem,
    Port, Airport, AirportRegion, LogisticsProvider, ProviderRate,
    ShippingInstruction, ShipmentMilestone, TrackingTemplate
)
import csv
import io
from django.http import HttpResponse
from datetime import datetime


logger = logging.getLogger(__name__)


class MasterAdminLoginView(APIView):
    """
    Exclusive login endpoint for MASTER ADMIN.
    Uses credentials from Replit Secrets - NOT linked to regular user auth.
    """
    permission_classes = [AllowAny]
    authentication_classes = []
    
    def post(self, request):
        username = request.data.get('username', '')
        password = request.data.get('password', '')
        
        if not username or not password:
            return Response({
                'success': False,
                'error': 'Credenciales requeridas'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if validate_master_admin_credentials(username, password):
            token = create_master_admin_session()
            logger.info(f"MASTER ADMIN login successful from {request.META.get('REMOTE_ADDR')}")
            
            return Response({
                'success': True,
                'message': 'Acceso MASTER ADMIN concedido',
                'token': token,
                'expires_in_hours': 8
            })
        else:
            logger.warning(f"Failed MASTER ADMIN login attempt from {request.META.get('REMOTE_ADDR')}")
            return Response({
                'success': False,
                'error': 'Credenciales inválidas'
            }, status=status.HTTP_401_UNAUTHORIZED)


class MasterAdminLogoutView(APIView):
    """Logout endpoint for MASTER ADMIN."""
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def post(self, request):
        token = request.headers.get('X-Master-Admin-Token')
        if token:
            invalidate_master_admin_session(token)
        return Response({'success': True, 'message': 'Sesión MASTER ADMIN cerrada'})


class MasterAdminDashboardView(APIView):
    """
    Main Dashboard - High-level KPIs and system overview.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        total_leads = CustomUser.objects.filter(role='lead').count()
        total_cotizaciones = LeadCotizacion.objects.count()
        cotizaciones_aprobadas = LeadCotizacion.objects.filter(estado='aprobada').count()
        cotizaciones_rechazadas = LeadCotizacion.objects.filter(estado='rechazada').count()
        cotizaciones_pendientes = LeadCotizacion.objects.filter(estado='pendiente').count()
        
        ros_activos = LeadCotizacion.objects.filter(
            estado__in=['ro_generado', 'en_transito'],
            ro_number__isnull=False
        ).count()
        
        embarques_activos = Shipment.objects.exclude(
            current_status__in=['entregado', 'incidencia']
        ).count()
        
        valor_total_cotizado = LeadCotizacion.objects.aggregate(
            total=Sum('total_usd')
        )['total'] or 0
        
        tributos_totales = PreLiquidation.objects.filter(
            is_confirmed=True
        ).aggregate(
            total=Sum('total_tributos_usd')
        )['total'] or 0
        
        cotizaciones_mes = LeadCotizacion.objects.filter(
            fecha_creacion__gte=month_start
        ).count()
        
        cotizaciones_por_estado = LeadCotizacion.objects.values('estado').annotate(
            count=Count('id')
        ).order_by('estado')
        
        embarques_por_estado = Shipment.objects.values('current_status').annotate(
            count=Count('id')
        ).order_by('current_status')
        
        return Response({
            'kpis': {
                'total_leads': total_leads,
                'total_cotizaciones': total_cotizaciones,
                'cotizaciones_aprobadas': cotizaciones_aprobadas,
                'cotizaciones_rechazadas': cotizaciones_rechazadas,
                'cotizaciones_pendientes': cotizaciones_pendientes,
                'ros_activos': ros_activos,
                'embarques_activos': embarques_activos,
                'valor_total_cotizado_usd': float(valor_total_cotizado),
                'tributos_totales_usd': float(tributos_totales),
            },
            'metricas_mes': {
                'cotizaciones_nuevas': cotizaciones_mes,
            },
            'distribucion': {
                'cotizaciones_por_estado': list(cotizaciones_por_estado),
                'embarques_por_estado': list(embarques_por_estado),
            },
            'timestamp': now.isoformat()
        })


class MasterAdminUsersView(APIView):
    """
    Full CRUD access to all Users (LEADs).
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        from accounts.models import CustomerRUC
        
        user_id = request.query_params.get('id')
        
        if user_id:
            try:
                user = CustomUser.objects.get(id=user_id)
                profile = None
                try:
                    profile = LeadProfile.objects.get(user=user)
                    profile_data = {
                        'ruc': profile.ruc,
                        'razon_social': profile.razon_social,
                        'tipo_persona': profile.tipo_persona,
                        'codigo_senae': profile.codigo_senae,
                        'direccion_matriz': profile.direccion_matriz,
                        'ciudad': profile.ciudad,
                        'provincia': profile.provincia,
                        'telefono_empresa': profile.telefono_empresa,
                        'productos_frecuentes': profile.productos_frecuentes,
                        'paises_origen_frecuentes': profile.paises_origen_frecuentes,
                    }
                except LeadProfile.DoesNotExist:
                    profile_data = None
                
                return Response({
                    'id': user.id,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'phone': user.phone,
                    'role': user.role,
                    'is_active': user.is_active,
                    'date_joined': user.date_joined,
                    'last_login': user.last_login,
                    'profile': profile_data
                })
            except CustomUser.DoesNotExist:
                return Response({'error': 'Usuario no encontrado'}, status=404)
        
        users = CustomUser.objects.all().order_by('-date_joined')
        
        role_filter = request.query_params.get('role')
        if role_filter:
            users = users.filter(role=role_filter)
        
        search = request.query_params.get('search')
        if search:
            users = users.filter(
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        
        status_filter = request.query_params.get('status', 'all')
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)
        
        ruc_status_filter = request.query_params.get('ruc_status', 'all')
        if ruc_status_filter != 'all':
            user_ids_with_ruc_status = CustomerRUC.objects.filter(
                status=ruc_status_filter
            ).values_list('user_id', flat=True).distinct()
            users = users.filter(id__in=user_ids_with_ruc_status)
        
        date_from = request.query_params.get('date_from')
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d')
                users = users.filter(date_joined__date__gte=from_date.date())
            except ValueError:
                pass
        
        date_to = request.query_params.get('date_to')
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d')
                users = users.filter(date_joined__date__lte=to_date.date())
            except ValueError:
                pass
        
        users_list = users[:100]
        user_ids = [u.id for u in users_list]
        
        ruc_statuses = {}
        rucs = CustomerRUC.objects.filter(user_id__in=user_ids).values('user_id', 'status')
        for ruc in rucs:
            user_id = ruc['user_id']
            if user_id not in ruc_statuses:
                ruc_statuses[user_id] = ruc['status']
            elif ruc['status'] == 'approved':
                ruc_statuses[user_id] = 'approved'
        
        return Response({
            'total': users.count(),
            'users': [
                {
                    'id': u.id,
                    'email': u.email,
                    'first_name': u.first_name,
                    'last_name': u.last_name,
                    'phone': u.phone,
                    'role': u.role,
                    'is_active': u.is_active,
                    'date_joined': u.date_joined,
                    'last_login': u.last_login,
                    'ruc_status': ruc_statuses.get(u.id, None),
                }
                for u in users_list
            ]
        })
    
    def put(self, request):
        user_id = request.data.get('id')
        if not user_id:
            return Response({'error': 'ID de usuario requerido'}, status=400)
        
        try:
            user = CustomUser.objects.get(id=user_id)
            
            updatable_fields = ['first_name', 'last_name', 'phone', 'is_active', 'role']
            for field in updatable_fields:
                if field in request.data:
                    setattr(user, field, request.data[field])
            
            user.save()
            return Response({'success': True, 'message': 'Usuario actualizado'})
        except CustomUser.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=404)
    
    def delete(self, request):
        user_id = request.query_params.get('id')
        if not user_id:
            return Response({'error': 'ID de usuario requerido'}, status=400)
        
        try:
            user = CustomUser.objects.get(id=user_id)
            user.is_active = False
            user.save()
            return Response({'success': True, 'message': 'Usuario desactivado'})
        except CustomUser.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=404)


class MasterAdminUserDetailView(APIView):
    """
    Complete user detail view with all related data:
    - Profile info, RUCs
    - Cotizaciones (quotes)
    - Shipping Instructions (SI)
    - Routing Orders (RO)
    - Shipments and tracking
    - Documents and invoices
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request, user_id):
        from accounts.models import CustomerRUC
        
        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=404)
        
        profile_data = None
        try:
            profile = LeadProfile.objects.get(user=user)
            profile_data = {
                'ruc': profile.ruc,
                'legal_type': profile.legal_type,
                'is_active_importer': profile.is_active_importer,
                'senae_code': profile.senae_code,
                'business_address': profile.business_address,
                'postal_code': profile.postal_code,
                'preferred_trade_lane': profile.preferred_trade_lane,
                'preferred_transport': profile.preferred_transport,
                'main_products': profile.main_products,
                'average_monthly_volume': profile.average_monthly_volume,
                'customs_broker_name': profile.customs_broker_name,
                'customs_broker_code': profile.customs_broker_code,
                'notification_email': profile.notification_email,
                'notification_whatsapp': profile.notification_whatsapp,
                'notes': profile.notes,
                'is_profile_complete': profile.is_profile_complete,
            }
        except LeadProfile.DoesNotExist:
            pass
        
        rucs = CustomerRUC.objects.filter(user=user).order_by('-is_primary', '-created_at')
        rucs_data = [{
            'id': r.id,
            'ruc': r.ruc,
            'company_name': r.company_name,
            'is_primary': r.is_primary,
            'status': r.status,
            'justification': r.justification,
            'relationship_description': r.relationship_description,
            'is_oce_registered': r.is_oce_registered,
            'created_at': r.created_at,
        } for r in rucs]
        
        cotizaciones = LeadCotizacion.objects.filter(lead_user=user).order_by('-fecha_creacion')
        cotizaciones_data = [{
            'id': c.id,
            'numero_cotizacion': c.numero_cotizacion,
            'tipo_carga': c.tipo_carga,
            'origen_pais': c.origen_pais,
            'origen_ciudad': c.origen_ciudad,
            'destino_ciudad': c.destino_ciudad,
            'descripcion_mercancia': c.descripcion_mercancia,
            'valor_mercancia_usd': float(c.valor_mercancia_usd) if c.valor_mercancia_usd else 0,
            'total_usd': float(c.total_usd) if c.total_usd else 0,
            'estado': c.estado,
            'ro_number': c.ro_number,
            'fecha_creacion': c.fecha_creacion,
            'fecha_aprobacion': c.fecha_aprobacion,
        } for c in cotizaciones]
        
        cotizaciones_aprobadas = cotizaciones.filter(estado__in=['aprobada', 'ro_generado', 'en_transito', 'entregado'])
        
        shipping_instructions = ShippingInstruction.objects.filter(
            cotizacion__lead_user=user
        ).select_related('cotizacion').order_by('-created_at')
        si_data = [{
            'id': si.id,
            'ro_number': si.ro_number,
            'cotizacion_numero': si.cotizacion.numero_cotizacion if si.cotizacion else None,
            'shipper_name': si.shipper_name,
            'consignee_name': si.consignee_name,
            'notify_party': si.notify_party,
            'status': si.status,
            'created_at': si.created_at,
            'submitted_at': si.submitted_at,
        } for si in shipping_instructions]
        
        ros_data = []
        for cot in cotizaciones.filter(ro_number__isnull=False):
            try:
                si = ShippingInstruction.objects.get(cotizacion=cot)
                docs = ShippingInstructionDocument.objects.filter(shipping_instruction=si)
                milestones = ShipmentMilestone.objects.filter(shipping_instruction=si).order_by('milestone_order')
                
                ros_data.append({
                    'ro_number': cot.ro_number,
                    'cotizacion_id': cot.id,
                    'numero_cotizacion': cot.numero_cotizacion,
                    'origen': f"{cot.origen_ciudad}, {cot.origen_pais}",
                    'destino': cot.destino_ciudad,
                    'estado_cotizacion': cot.estado,
                    'si_status': si.status,
                    'fecha_ro': cot.fecha_aprobacion,
                    'documentos': [{
                        'id': d.id,
                        'document_type': d.document_type,
                        'file_name': d.file_name,
                        'file_url': d.file.url if d.file else None,
                        'uploaded_at': d.uploaded_at,
                    } for d in docs],
                    'milestones': [{
                        'id': m.id,
                        'milestone_type': m.milestone_type,
                        'title': m.title,
                        'is_completed': m.is_completed,
                        'completed_at': m.completed_at,
                        'notes': m.notes,
                    } for m in milestones],
                })
            except ShippingInstruction.DoesNotExist:
                ros_data.append({
                    'ro_number': cot.ro_number,
                    'cotizacion_id': cot.id,
                    'numero_cotizacion': cot.numero_cotizacion,
                    'origen': f"{cot.origen_ciudad}, {cot.origen_pais}",
                    'destino': cot.destino_ciudad,
                    'estado_cotizacion': cot.estado,
                    'si_status': 'pending',
                    'fecha_ro': cot.fecha_aprobacion,
                    'documentos': [],
                    'milestones': [],
                })
        
        shipments = Shipment.objects.filter(lead_user=user).order_by('-created_at')
        shipments_data = [{
            'id': s.id,
            'tracking_number': s.tracking_number,
            'ro_number': s.ro_number,
            'current_status': s.current_status,
            'origin_port': s.origin_port,
            'destination_port': s.destination_port,
            'carrier': s.carrier,
            'vessel_name': s.vessel_name,
            'eta': s.eta,
            'ata': s.ata,
            'created_at': s.created_at,
        } for s in shipments]
        
        shipments_arribados = shipments.filter(current_status='entregado')
        
        preliquidations = PreLiquidation.objects.filter(
            cotizacion__lead_user=user
        ).select_related('cotizacion').order_by('-created_at')
        preliq_data = [{
            'id': p.id,
            'cotizacion_numero': p.cotizacion.numero_cotizacion if p.cotizacion else None,
            'hs_code': p.hs_code,
            'cif_usd': float(p.cif_usd) if p.cif_usd else 0,
            'total_tributos_usd': float(p.total_tributos_usd) if p.total_tributos_usd else 0,
            'is_confirmed': p.is_confirmed,
            'created_at': p.created_at,
        } for p in preliquidations]
        
        return Response({
            'user': {
                'id': user.id,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'full_name': user.get_full_name(),
                'phone': user.phone,
                'whatsapp': user.whatsapp,
                'company_name': user.company_name,
                'city': user.city,
                'country': user.country,
                'role': user.role,
                'platform': user.platform,
                'is_active': user.is_active,
                'is_email_verified': user.is_email_verified,
                'date_joined': user.date_joined,
                'last_login': user.last_login,
            },
            'profile': profile_data,
            'rucs': rucs_data,
            'stats': {
                'total_cotizaciones': cotizaciones.count(),
                'cotizaciones_aprobadas': cotizaciones_aprobadas.count(),
                'total_ros': len(ros_data),
                'total_shipments': shipments.count(),
                'shipments_arribados': shipments_arribados.count(),
                'total_preliquidations': preliquidations.count(),
            },
            'cotizaciones': cotizaciones_data,
            'shipping_instructions': si_data,
            'routing_orders': ros_data,
            'shipments': shipments_data,
            'preliquidations': preliq_data,
        })
    
    def put(self, request, user_id):
        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=404)
        
        user_fields = ['first_name', 'last_name', 'phone', 'whatsapp', 'company_name', 
                       'city', 'country', 'is_active', 'is_email_verified']
        for field in user_fields:
            if field in request.data:
                setattr(user, field, request.data[field])
        user.save()
        
        profile_data = request.data.get('profile')
        if profile_data:
            profile, _ = LeadProfile.objects.get_or_create(user=user)
            profile_fields = ['legal_type', 'is_active_importer', 'senae_code', 'business_address',
                              'postal_code', 'preferred_trade_lane', 'preferred_transport',
                              'main_products', 'average_monthly_volume', 'customs_broker_name',
                              'customs_broker_code', 'notification_email', 'notification_whatsapp', 'notes']
            for field in profile_fields:
                if field in profile_data:
                    setattr(profile, field, profile_data[field])
            profile.save()
        
        return Response({'success': True, 'message': 'Usuario actualizado correctamente'})
    
    def delete(self, request, user_id):
        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=404)
        
        action = request.query_params.get('action', 'deactivate')
        
        if action == 'deactivate':
            user.is_active = False
            user.save()
            return Response({'success': True, 'message': 'Usuario desactivado'})
        elif action == 'delete':
            has_quotes = LeadCotizacion.objects.filter(lead_user=user).exists()
            has_shipments = Shipment.objects.filter(lead_user=user).exists()
            
            if has_quotes or has_shipments:
                return Response({
                    'error': 'No se puede eliminar usuario con cotizaciones o embarques. Use desactivar.',
                    'has_quotes': has_quotes,
                    'has_shipments': has_shipments,
                }, status=400)
            
            user.delete()
            return Response({'success': True, 'message': 'Usuario eliminado permanentemente'})
        
        return Response({'error': 'Acción no válida. Use deactivate o delete'}, status=400)


class MasterAdminCotizacionesView(APIView):
    """
    Full CRUD access to all Cotizaciones.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        cotizacion_id = request.query_params.get('id')
        
        if cotizacion_id:
            try:
                cot = LeadCotizacion.objects.get(id=cotizacion_id)
                scenarios = QuoteScenario.objects.filter(cotizacion=cot)
                
                return Response({
                    'cotizacion': {
                        'id': cot.id,
                        'numero_cotizacion': cot.numero_cotizacion,
                        'lead_user_id': cot.lead_user_id,
                        'lead_email': cot.lead_user.email if cot.lead_user else None,
                        'tipo_carga': cot.tipo_carga,
                        'origen_pais': cot.origen_pais,
                        'origen_ciudad': cot.origen_ciudad,
                        'destino_ciudad': cot.destino_ciudad,
                        'descripcion_mercancia': cot.descripcion_mercancia,
                        'peso_kg': float(cot.peso_kg) if cot.peso_kg else 0,
                        'volumen_cbm': float(cot.volumen_cbm) if cot.volumen_cbm else 0,
                        'valor_mercancia_usd': float(cot.valor_mercancia_usd) if cot.valor_mercancia_usd else 0,
                        'incoterm': cot.incoterm,
                        'flete_usd': float(cot.flete_usd) if cot.flete_usd else 0,
                        'seguro_usd': float(cot.seguro_usd) if cot.seguro_usd else 0,
                        'aduana_usd': float(cot.aduana_usd) if cot.aduana_usd else 0,
                        'transporte_interno_usd': float(cot.transporte_interno_usd) if cot.transporte_interno_usd else 0,
                        'otros_usd': float(cot.otros_usd) if cot.otros_usd else 0,
                        'total_usd': float(cot.total_usd) if cot.total_usd else 0,
                        'estado': cot.estado,
                        'ro_number': cot.ro_number,
                        'fecha_creacion': cot.fecha_creacion,
                        'fecha_aprobacion': cot.fecha_aprobacion,
                    },
                    'scenarios': [
                        {
                            'id': s.id,
                            'nombre': s.nombre,
                            'tipo_transporte': s.tipo_transporte,
                            'total_usd': float(s.total_usd) if s.total_usd else 0,
                            'is_selected': s.is_selected,
                        }
                        for s in scenarios
                    ]
                })
            except LeadCotizacion.DoesNotExist:
                return Response({'error': 'Cotización no encontrada'}, status=404)
        
        cotizaciones = LeadCotizacion.objects.all().order_by('-fecha_creacion')
        
        estado_filter = request.query_params.get('estado')
        if estado_filter:
            cotizaciones = cotizaciones.filter(estado=estado_filter)
        
        search = request.query_params.get('search')
        if search:
            cotizaciones = cotizaciones.filter(
                Q(numero_cotizacion__icontains=search) |
                Q(descripcion_mercancia__icontains=search) |
                Q(ro_number__icontains=search)
            )
        
        return Response({
            'total': cotizaciones.count(),
            'cotizaciones': [
                {
                    'id': c.id,
                    'numero_cotizacion': c.numero_cotizacion,
                    'lead_email': c.lead_user.email if c.lead_user else None,
                    'origen': f"{c.origen_ciudad}, {c.origen_pais}",
                    'destino': c.destino_ciudad,
                    'total_usd': float(c.total_usd) if c.total_usd else 0,
                    'estado': c.estado,
                    'ro_number': c.ro_number,
                    'fecha_creacion': c.fecha_creacion,
                }
                for c in cotizaciones[:100]
            ]
        })
    
    def put(self, request):
        cotizacion_id = request.data.get('id')
        if not cotizacion_id:
            return Response({'error': 'ID de cotización requerido'}, status=400)
        
        try:
            cot = LeadCotizacion.objects.get(id=cotizacion_id)
            
            updatable_fields = [
                'estado', 'flete_usd', 'seguro_usd', 'aduana_usd',
                'transporte_interno_usd', 'otros_usd', 'total_usd'
            ]
            for field in updatable_fields:
                if field in request.data:
                    setattr(cot, field, request.data[field])
            
            cot.save()
            return Response({'success': True, 'message': 'Cotización actualizada'})
        except LeadCotizacion.DoesNotExist:
            return Response({'error': 'Cotización no encontrada'}, status=404)
    
    def delete(self, request):
        cotizacion_id = request.query_params.get('id')
        if not cotizacion_id:
            return Response({'error': 'ID de cotización requerido'}, status=400)
        
        try:
            cot = LeadCotizacion.objects.get(id=cotizacion_id)
            cot.delete()
            return Response({'success': True, 'message': 'Cotización eliminada'})
        except LeadCotizacion.DoesNotExist:
            return Response({'error': 'Cotización no encontrada'}, status=404)


class MasterAdminCotizacionDetailView(APIView):
    """
    Detailed view of a single Cotización with all related data.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request, cotizacion_id):
        try:
            cot = LeadCotizacion.objects.select_related('lead_user').get(id=cotizacion_id)
        except LeadCotizacion.DoesNotExist:
            return Response({'error': 'Cotización no encontrada'}, status=404)
        
        user = cot.lead_user
        customer_info = {
            'id': user.id if user else None,
            'name': user.get_full_name() if user else '',
            'email': user.email if user else '',
            'company': user.company_name if user else '',
            'phone': user.phone if user else '',
            'city': user.city if user else '',
            'country': user.country if user else '',
            'ruc': '',
        }
        
        if user:
            from accounts.models import CustomerRUC
            primary_ruc = CustomerRUC.objects.filter(user=user, is_primary=True, status='approved').first()
            if primary_ruc:
                customer_info['ruc'] = primary_ruc.ruc
                customer_info['company'] = primary_ruc.company_name or customer_info['company']
        
        scenarios = QuoteScenario.objects.filter(cotizacion=cot)
        scenarios_data = []
        for s in scenarios:
            line_items = QuoteLineItem.objects.filter(escenario=s).order_by('categoria')
            lines_data = [{
                'id': li.id,
                'categoria': li.categoria,
                'categoria_display': li.get_categoria_display() if hasattr(li, 'get_categoria_display') else li.categoria,
                'descripcion': li.descripcion,
                'cantidad': float(li.cantidad) if li.cantidad else 1,
                'precio_unitario_usd': float(li.precio_unitario_usd) if li.precio_unitario_usd else 0,
                'subtotal_usd': float(li.subtotal_usd) if li.subtotal_usd else 0,
                'es_estimado': li.es_estimado,
                'notas': li.notas,
            } for li in line_items]
            
            scenarios_data.append({
                'id': s.id,
                'nombre': s.nombre,
                'tipo': s.tipo,
                'tipo_transporte': getattr(s, 'tipo_transporte', ''),
                'total_usd': float(s.total_usd) if hasattr(s, 'total_usd') and s.total_usd else 0,
                'is_selected': getattr(s, 'is_selected', False),
                'lineas': lines_data,
            })
        
        costs_breakdown = {
            'flete_usd': float(cot.flete_usd) if cot.flete_usd else 0,
            'seguro_usd': float(cot.seguro_usd) if cot.seguro_usd else 0,
            'aduana_usd': float(cot.aduana_usd) if cot.aduana_usd else 0,
            'transporte_interno_usd': float(cot.transporte_interno_usd) if cot.transporte_interno_usd else 0,
            'otros_usd': float(cot.otros_usd) if cot.otros_usd else 0,
            'total_usd': float(cot.total_usd) if cot.total_usd else 0,
        }
        
        documents = []
        try:
            si = ShippingInstruction.objects.filter(cotizacion=cot).first()
            if si:
                from SalesModule.models import ShippingInstructionDocument
                si_docs = ShippingInstructionDocument.objects.filter(shipping_instruction=si)
                for doc in si_docs:
                    documents.append({
                        'id': doc.id,
                        'type': doc.document_type,
                        'type_display': doc.get_document_type_display() if hasattr(doc, 'get_document_type_display') else doc.document_type,
                        'file_name': doc.file_name,
                        'file_url': doc.file.url if doc.file else None,
                        'uploaded_at': doc.uploaded_at,
                        'source': 'shipping_instruction',
                    })
        except Exception:
            pass
        
        try:
            preliq = PreLiquidation.objects.filter(cotizacion=cot).first()
            if preliq:
                from SalesModule.models import PreLiquidationDocument
                preliq_docs = PreLiquidationDocument.objects.filter(pre_liquidation=preliq)
                for doc in preliq_docs:
                    documents.append({
                        'id': doc.id,
                        'type': doc.document_type,
                        'type_display': doc.get_document_type_display() if hasattr(doc, 'get_document_type_display') else doc.document_type,
                        'file_name': doc.file_name,
                        'file_url': doc.file_path,
                        'uploaded_at': doc.created_at if hasattr(doc, 'created_at') else None,
                        'source': 'preliquidation',
                    })
        except Exception:
            pass
        
        return Response({
            'success': True,
            'cotizacion': {
                'id': cot.id,
                'numero_cotizacion': cot.numero_cotizacion,
                'estado': cot.estado,
                'tipo_carga': cot.tipo_carga,
                'origen_pais': cot.origen_pais,
                'origen_ciudad': cot.origen_ciudad,
                'destino_ciudad': cot.destino_ciudad,
                'incoterm': cot.incoterm,
                'descripcion_mercancia': cot.descripcion_mercancia,
                'peso_kg': float(cot.peso_kg) if cot.peso_kg else 0,
                'volumen_cbm': float(cot.volumen_cbm) if cot.volumen_cbm else 0,
                'valor_mercancia_usd': float(cot.valor_mercancia_usd) if cot.valor_mercancia_usd else 0,
                'requiere_seguro': cot.requiere_seguro,
                'requiere_transporte_interno': cot.requiere_transporte_interno,
                'notas_adicionales': cot.notas_adicionales,
                'ro_number': cot.ro_number,
                'shipper_name': cot.shipper_name,
                'shipper_address': cot.shipper_address,
                'consignee_name': cot.consignee_name,
                'consignee_address': cot.consignee_address,
                'notify_party': cot.notify_party,
                'fecha_embarque_estimada': cot.fecha_embarque_estimada,
                'fecha_creacion': cot.fecha_creacion,
                'fecha_actualizacion': cot.fecha_actualizacion,
                'fecha_aprobacion': cot.fecha_aprobacion,
            },
            'customer': customer_info,
            'costs': costs_breakdown,
            'scenarios': scenarios_data,
            'documents': documents,
        })


class MasterAdminShipmentsView(APIView):
    """
    Full CRUD access to all Shipments and Tracking.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        shipment_id = request.query_params.get('id')
        
        if shipment_id:
            try:
                ship = Shipment.objects.get(id=shipment_id)
                tracking = ShipmentTracking.objects.filter(shipment=ship).order_by('-timestamp')
                
                return Response({
                    'shipment': {
                        'id': ship.id,
                        'tracking_number': ship.tracking_number,
                        'cotizacion_id': ship.cotizacion_id,
                        'transport_type': ship.transport_type,
                        'carrier_name': ship.carrier_name,
                        'bl_awb_number': ship.bl_awb_number,
                        'container_number': ship.container_number,
                        'origin_country': ship.origin_country,
                        'origin_city': ship.origin_city,
                        'destination_city': ship.destination_city,
                        'current_status': ship.current_status,
                        'current_location': ship.current_location,
                        'estimated_departure': ship.estimated_departure,
                        'actual_departure': ship.actual_departure,
                        'estimated_arrival': ship.estimated_arrival,
                        'actual_arrival': ship.actual_arrival,
                        'estimated_delivery': ship.estimated_delivery,
                        'actual_delivery': ship.actual_delivery,
                    },
                    'tracking_history': [
                        {
                            'id': t.id,
                            'status': t.status,
                            'location': t.location,
                            'description': t.description,
                            'timestamp': t.timestamp,
                        }
                        for t in tracking
                    ]
                })
            except Shipment.DoesNotExist:
                return Response({'error': 'Embarque no encontrado'}, status=404)
        
        shipments = Shipment.objects.all().order_by('-created_at')
        
        status_filter = request.query_params.get('status')
        if status_filter:
            shipments = shipments.filter(current_status=status_filter)
        
        return Response({
            'total': shipments.count(),
            'shipments': [
                {
                    'id': s.id,
                    'tracking_number': s.tracking_number,
                    'origin': f"{s.origin_city}, {s.origin_country}",
                    'destination': s.destination_city,
                    'transport_type': s.transport_type,
                    'current_status': s.current_status,
                    'current_location': s.current_location,
                    'created_at': s.created_at,
                }
                for s in shipments[:100]
            ]
        })
    
    def put(self, request):
        shipment_id = request.data.get('id')
        if not shipment_id:
            return Response({'error': 'ID de embarque requerido'}, status=400)
        
        try:
            ship = Shipment.objects.get(id=shipment_id)
            
            updatable_fields = [
                'current_status', 'current_location', 'carrier_name',
                'estimated_arrival', 'actual_arrival', 'estimated_delivery', 'actual_delivery'
            ]
            for field in updatable_fields:
                if field in request.data:
                    setattr(ship, field, request.data[field])
            
            ship.save()
            return Response({'success': True, 'message': 'Embarque actualizado'})
        except Shipment.DoesNotExist:
            return Response({'error': 'Embarque no encontrado'}, status=404)


class MasterAdminRatesView(APIView):
    """
    Full CRUD access to all Cost/Rate tables.
    Supports: FreightRate, InsuranceRate, CustomsDutyRate, InlandTransportQuoteRate, CustomsBrokerageRate
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def _serialize_freight_rate(self, r):
        return {
            'id': r.id,
            'origin_country': r.origin_country,
            'origin_port': r.origin_port,
            'destination_country': r.destination_country,
            'destination_port': r.destination_port,
            'transport_type': r.transport_type,
            'transport_type_display': r.get_transport_type_display(),
            'rate_usd': float(r.rate_usd),
            'unit': r.unit,
            'unit_display': r.get_unit_display(),
            'min_rate_usd': float(r.min_rate_usd) if r.min_rate_usd else None,
            'carrier_name': r.carrier_name,
            'transit_days_min': r.transit_days_min,
            'transit_days_max': r.transit_days_max,
            'valid_from': r.valid_from.isoformat() if r.valid_from else None,
            'valid_until': r.valid_until.isoformat() if r.valid_until else None,
            'is_active': r.is_active,
            'notes': r.notes,
            'created_at': r.created_at.isoformat() if r.created_at else None,
        }
    
    def _serialize_insurance_rate(self, r):
        return {
            'id': r.id,
            'name': r.name,
            'coverage_type': r.coverage_type,
            'coverage_type_display': r.get_coverage_type_display(),
            'rate_percentage': float(r.rate_percentage),
            'min_premium_usd': float(r.min_premium_usd),
            'deductible_percentage': float(r.deductible_percentage),
            'insurance_company': r.insurance_company,
            'policy_number': r.policy_number,
            'valid_from': r.valid_from.isoformat() if r.valid_from else None,
            'valid_until': r.valid_until.isoformat() if r.valid_until else None,
            'is_active': r.is_active,
            'notes': r.notes,
            'created_at': r.created_at.isoformat() if r.created_at else None,
        }
    
    def _serialize_customs_rate(self, r):
        return {
            'id': r.id,
            'hs_code': r.hs_code,
            'description': r.description,
            'ad_valorem_percentage': float(r.ad_valorem_percentage),
            'iva_percentage': float(r.iva_percentage),
            'fodinfa_percentage': float(r.fodinfa_percentage),
            'ice_percentage': float(r.ice_percentage) if r.ice_percentage else 0,
            'salvaguardia_percentage': float(r.salvaguardia_percentage) if r.salvaguardia_percentage else 0,
            'specific_duty_usd': float(r.specific_duty_usd) if r.specific_duty_usd else 0,
            'specific_duty_unit': r.specific_duty_unit,
            'requires_import_license': r.requires_import_license,
            'requires_phytosanitary': r.requires_phytosanitary,
            'requires_inen_certification': r.requires_inen_certification,
            'valid_from': r.valid_from.isoformat() if r.valid_from else None,
            'valid_until': r.valid_until.isoformat() if r.valid_until else None,
            'is_active': r.is_active,
            'notes': r.notes,
            'created_at': r.created_at.isoformat() if r.created_at else None,
        }
    
    def _serialize_inland_rate(self, r):
        return {
            'id': r.id,
            'origin_city': r.origin_city,
            'destination_city': r.destination_city,
            'vehicle_type': r.vehicle_type,
            'vehicle_type_display': r.get_vehicle_type_display(),
            'rate_usd': float(r.rate_usd),
            'rate_per_kg_usd': float(r.rate_per_kg_usd) if r.rate_per_kg_usd else None,
            'estimated_hours': r.estimated_hours,
            'distance_km': r.distance_km,
            'includes_loading': r.includes_loading,
            'includes_unloading': r.includes_unloading,
            'carrier_name': r.carrier_name,
            'valid_from': r.valid_from.isoformat() if r.valid_from else None,
            'valid_until': r.valid_until.isoformat() if r.valid_until else None,
            'is_active': r.is_active,
            'notes': r.notes,
            'created_at': r.created_at.isoformat() if r.created_at else None,
        }
    
    def _serialize_brokerage_rate(self, r):
        return {
            'id': r.id,
            'name': r.name,
            'service_type': r.service_type,
            'service_type_display': r.get_service_type_display(),
            'fixed_rate_usd': float(r.fixed_rate_usd),
            'percentage_rate': float(r.percentage_rate),
            'min_rate_usd': float(r.min_rate_usd),
            'includes_aforo': r.includes_aforo,
            'includes_transmision': r.includes_transmision,
            'includes_almacenaje': r.includes_almacenaje,
            'valid_from': r.valid_from.isoformat() if r.valid_from else None,
            'valid_until': r.valid_until.isoformat() if r.valid_until else None,
            'is_active': r.is_active,
            'notes': r.notes,
            'created_at': r.created_at.isoformat() if r.created_at else None,
        }
    
    def _get_model_and_serializer(self, rate_type):
        mapping = {
            'freight': (FreightRate, self._serialize_freight_rate),
            'insurance': (InsuranceRate, self._serialize_insurance_rate),
            'customs': (CustomsDutyRate, self._serialize_customs_rate),
            'inland': (InlandTransportQuoteRate, self._serialize_inland_rate),
            'brokerage': (CustomsBrokerageRate, self._serialize_brokerage_rate),
        }
        return mapping.get(rate_type)
    
    def get(self, request):
        rate_type = request.query_params.get('type', 'freight')
        rate_id = request.query_params.get('id')
        
        result = self._get_model_and_serializer(rate_type)
        if not result:
            return Response({'error': 'Tipo de tarifa no válido'}, status=400)
        
        Model, serializer = result
        
        if rate_id:
            try:
                rate = Model.objects.get(id=rate_id)
                return Response({'rate': serializer(rate)})
            except Model.DoesNotExist:
                return Response({'error': 'Tarifa no encontrada'}, status=404)
        
        rates = Model.objects.all().order_by('-created_at')
        
        search = request.query_params.get('search', '').strip()
        if search:
            if rate_type == 'freight':
                rates = rates.filter(
                    Q(origin_port__icontains=search) | 
                    Q(destination_port__icontains=search) |
                    Q(origin_country__icontains=search) |
                    Q(carrier_name__icontains=search)
                )
            elif rate_type == 'insurance':
                rates = rates.filter(
                    Q(name__icontains=search) | 
                    Q(insurance_company__icontains=search)
                )
            elif rate_type == 'customs':
                rates = rates.filter(
                    Q(hs_code__icontains=search) | 
                    Q(description__icontains=search)
                )
            elif rate_type == 'inland':
                rates = rates.filter(
                    Q(origin_city__icontains=search) | 
                    Q(destination_city__icontains=search) |
                    Q(carrier_name__icontains=search)
                )
            elif rate_type == 'brokerage':
                rates = rates.filter(Q(name__icontains=search))
        
        is_active = request.query_params.get('is_active')
        if is_active is not None:
            rates = rates.filter(is_active=is_active.lower() == 'true')
        
        return Response({
            'type': rate_type,
            'total': rates.count(),
            'rates': [serializer(r) for r in rates[:200]]
        })
    
    def post(self, request):
        rate_type = request.data.get('type', 'freight')
        data = request.data.get('data', {})
        
        result = self._get_model_and_serializer(rate_type)
        if not result:
            return Response({'error': 'Tipo de tarifa no válido'}, status=400)
        
        Model, serializer = result
        
        try:
            if rate_type == 'freight':
                rate = FreightRate.objects.create(
                    origin_country=data.get('origin_country', ''),
                    origin_port=data.get('origin_port', ''),
                    destination_country=data.get('destination_country', 'Ecuador'),
                    destination_port=data.get('destination_port', ''),
                    transport_type=data.get('transport_type', 'aereo'),
                    rate_usd=data.get('rate_usd', 0),
                    unit=data.get('unit', 'kg'),
                    min_rate_usd=data.get('min_rate_usd'),
                    carrier_name=data.get('carrier_name', ''),
                    transit_days_min=data.get('transit_days_min'),
                    transit_days_max=data.get('transit_days_max'),
                    valid_from=data.get('valid_from'),
                    valid_until=data.get('valid_until'),
                    is_active=data.get('is_active', True),
                    notes=data.get('notes', ''),
                )
            elif rate_type == 'insurance':
                rate = InsuranceRate.objects.create(
                    name=data.get('name', ''),
                    coverage_type=data.get('coverage_type', 'basico'),
                    rate_percentage=data.get('rate_percentage', 0),
                    min_premium_usd=data.get('min_premium_usd', 25),
                    deductible_percentage=data.get('deductible_percentage', 0),
                    insurance_company=data.get('insurance_company', ''),
                    policy_number=data.get('policy_number', ''),
                    valid_from=data.get('valid_from'),
                    valid_until=data.get('valid_until'),
                    is_active=data.get('is_active', True),
                    notes=data.get('notes', ''),
                )
            elif rate_type == 'customs':
                rate = CustomsDutyRate.objects.create(
                    hs_code=data.get('hs_code', ''),
                    description=data.get('description', ''),
                    ad_valorem_percentage=data.get('ad_valorem_percentage', 0),
                    iva_percentage=data.get('iva_percentage', 15),
                    fodinfa_percentage=data.get('fodinfa_percentage', 0.5),
                    ice_percentage=data.get('ice_percentage', 0),
                    salvaguardia_percentage=data.get('salvaguardia_percentage', 0),
                    specific_duty_usd=data.get('specific_duty_usd', 0),
                    specific_duty_unit=data.get('specific_duty_unit', ''),
                    requires_import_license=data.get('requires_import_license', False),
                    requires_phytosanitary=data.get('requires_phytosanitary', False),
                    requires_inen_certification=data.get('requires_inen_certification', False),
                    valid_from=data.get('valid_from'),
                    valid_until=data.get('valid_until'),
                    is_active=data.get('is_active', True),
                    notes=data.get('notes', ''),
                )
            elif rate_type == 'inland':
                rate = InlandTransportQuoteRate.objects.create(
                    origin_city=data.get('origin_city', 'Guayaquil'),
                    destination_city=data.get('destination_city', ''),
                    vehicle_type=data.get('vehicle_type', 'contenedor_20'),
                    rate_usd=data.get('rate_usd', 0),
                    rate_per_kg_usd=data.get('rate_per_kg_usd'),
                    estimated_hours=data.get('estimated_hours'),
                    distance_km=data.get('distance_km'),
                    includes_loading=data.get('includes_loading', False),
                    includes_unloading=data.get('includes_unloading', False),
                    carrier_name=data.get('carrier_name', ''),
                    valid_from=data.get('valid_from'),
                    valid_until=data.get('valid_until'),
                    is_active=data.get('is_active', True),
                    notes=data.get('notes', ''),
                )
            elif rate_type == 'brokerage':
                rate = CustomsBrokerageRate.objects.create(
                    name=data.get('name', ''),
                    service_type=data.get('service_type', 'importacion_general'),
                    fixed_rate_usd=data.get('fixed_rate_usd', 150),
                    percentage_rate=data.get('percentage_rate', 0),
                    min_rate_usd=data.get('min_rate_usd', 150),
                    includes_aforo=data.get('includes_aforo', True),
                    includes_transmision=data.get('includes_transmision', True),
                    includes_almacenaje=data.get('includes_almacenaje', False),
                    valid_from=data.get('valid_from'),
                    valid_until=data.get('valid_until'),
                    is_active=data.get('is_active', True),
                    notes=data.get('notes', ''),
                )
            else:
                return Response({'error': 'Tipo de tarifa no válido'}, status=400)
            
            return Response({
                'success': True, 
                'message': 'Tarifa creada correctamente',
                'rate': serializer(rate)
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error creating rate: {e}")
            return Response({'error': f'Error al crear tarifa: {str(e)}'}, status=400)
    
    def put(self, request):
        rate_type = request.data.get('type', 'freight')
        rate_id = request.data.get('id')
        data = request.data.get('data', {})
        
        if not rate_id:
            return Response({'error': 'ID de tarifa requerido'}, status=400)
        
        result = self._get_model_and_serializer(rate_type)
        if not result:
            return Response({'error': 'Tipo de tarifa no válido'}, status=400)
        
        Model, serializer = result
        
        try:
            rate = Model.objects.get(id=rate_id)
            
            updatable_fields = []
            if rate_type == 'freight':
                updatable_fields = [
                    'origin_country', 'origin_port', 'destination_country', 'destination_port',
                    'transport_type', 'rate_usd', 'unit', 'min_rate_usd', 'carrier_name',
                    'transit_days_min', 'transit_days_max', 'valid_from', 'valid_until',
                    'is_active', 'notes'
                ]
            elif rate_type == 'insurance':
                updatable_fields = [
                    'name', 'coverage_type', 'rate_percentage', 'min_premium_usd',
                    'deductible_percentage', 'insurance_company', 'policy_number',
                    'valid_from', 'valid_until', 'is_active', 'notes'
                ]
            elif rate_type == 'customs':
                updatable_fields = [
                    'hs_code', 'description', 'ad_valorem_percentage', 'iva_percentage',
                    'fodinfa_percentage', 'ice_percentage', 'salvaguardia_percentage',
                    'specific_duty_usd', 'specific_duty_unit', 'requires_import_license',
                    'requires_phytosanitary', 'requires_inen_certification',
                    'valid_from', 'valid_until', 'is_active', 'notes'
                ]
            elif rate_type == 'inland':
                updatable_fields = [
                    'origin_city', 'destination_city', 'vehicle_type', 'rate_usd',
                    'rate_per_kg_usd', 'estimated_hours', 'distance_km',
                    'includes_loading', 'includes_unloading', 'carrier_name',
                    'valid_from', 'valid_until', 'is_active', 'notes'
                ]
            elif rate_type == 'brokerage':
                updatable_fields = [
                    'name', 'service_type', 'fixed_rate_usd', 'percentage_rate',
                    'min_rate_usd', 'includes_aforo', 'includes_transmision',
                    'includes_almacenaje', 'valid_from', 'valid_until', 'is_active', 'notes'
                ]
            
            for field in updatable_fields:
                if field in data:
                    setattr(rate, field, data[field])
            
            rate.save()
            
            return Response({
                'success': True, 
                'message': 'Tarifa actualizada correctamente',
                'rate': serializer(rate)
            })
            
        except Model.DoesNotExist:
            return Response({'error': 'Tarifa no encontrada'}, status=404)
        except Exception as e:
            logger.error(f"Error updating rate: {e}")
            return Response({'error': f'Error al actualizar tarifa: {str(e)}'}, status=400)
    
    def delete(self, request):
        rate_type = request.query_params.get('type', 'freight')
        rate_id = request.query_params.get('id')
        
        if not rate_id:
            return Response({'error': 'ID de tarifa requerido'}, status=400)
        
        result = self._get_model_and_serializer(rate_type)
        if not result:
            return Response({'error': 'Tipo de tarifa no válido'}, status=400)
        
        Model, _ = result
        
        try:
            rate = Model.objects.get(id=rate_id)
            rate.delete()
            return Response({'success': True, 'message': 'Tarifa eliminada correctamente'})
        except Model.DoesNotExist:
            return Response({'error': 'Tarifa no encontrada'}, status=404)
        except Exception as e:
            logger.error(f"Error deleting rate: {e}")
            return Response({'error': f'Error al eliminar tarifa: {str(e)}'}, status=400)


class MasterAdminProfitReviewView(APIView):
    """
    Financial Reporting - Profit Review by RO and Freight Forwarder.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        now = timezone.now()
        ros = LeadCotizacion.objects.filter(
            ro_number__isnull=False
        ).order_by('-fecha_creacion')
        
        profit_data = []
        total_revenue = 0
        total_costs = 0
        
        monthly_profits = {}
        transport_profits = {'FCL': 0, 'LCL': 0, 'AIR': 0}
        
        for ro in ros:
            flete = float(ro.flete_usd) if ro.flete_usd else 0
            seguro = float(ro.seguro_usd) if ro.seguro_usd else 0
            aduana = float(ro.aduana_usd) if ro.aduana_usd else 0
            transporte = float(ro.transporte_interno_usd) if ro.transporte_interno_usd else 0
            otros = float(ro.otros_usd) if ro.otros_usd else 0
            total = float(ro.total_usd) if ro.total_usd else 0
            
            estimated_cost = flete * 0.7 + seguro * 0.8 + aduana * 0.9 + transporte * 0.75
            margin = total - estimated_cost
            margin_pct = (margin / total * 100) if total > 0 else 0
            
            total_revenue += total
            total_costs += estimated_cost
            
            if ro.fecha_creacion:
                month_key = ro.fecha_creacion.strftime('%Y-%m')
                if month_key not in monthly_profits:
                    monthly_profits[month_key] = {'profit': 0, 'count': 0}
                monthly_profits[month_key]['profit'] += margin
                monthly_profits[month_key]['count'] += 1
            
            tipo_carga = (ro.tipo_carga or '').upper()
            if 'FCL' in tipo_carga:
                transport_profits['FCL'] += margin
            elif 'LCL' in tipo_carga:
                transport_profits['LCL'] += margin
            elif 'AIR' in tipo_carga or 'AERE' in tipo_carga or 'AÉREO' in tipo_carga:
                transport_profits['AIR'] += margin
            
            profit_data.append({
                'ro_number': ro.ro_number,
                'cotizacion_numero': ro.numero_cotizacion,
                'cliente_email': ro.lead_user.email if ro.lead_user else 'N/A',
                'origen': f"{ro.origen_ciudad}, {ro.origen_pais}",
                'destino': ro.destino_ciudad,
                'tipo_carga': ro.tipo_carga,
                'estado': ro.estado,
                'desglose': {
                    'flete_usd': flete,
                    'seguro_usd': seguro,
                    'aduana_usd': aduana,
                    'transporte_interno_usd': transporte,
                    'otros_usd': otros,
                },
                'total_facturado_usd': total,
                'costo_estimado_usd': round(estimated_cost, 2),
                'margen_usd': round(margin, 2),
                'margen_porcentaje': round(margin_pct, 2),
                'fecha_creacion': ro.fecha_creacion,
            })
        
        total_margin = total_revenue - total_costs
        avg_margin_pct = (total_margin / total_revenue * 100) if total_revenue > 0 else 0
        avg_profit_per_quote = (total_margin / len(profit_data)) if profit_data else 0
        
        monthly_chart_data = []
        for i in range(11, -1, -1):
            month_date = now - timedelta(days=i*30)
            month_key = month_date.strftime('%Y-%m')
            month_name = month_date.strftime('%b %Y')
            month_data = monthly_profits.get(month_key, {'profit': 0, 'count': 0})
            monthly_chart_data.append({
                'month': month_name,
                'profit': round(month_data['profit'], 2),
                'count': month_data['count']
            })
        
        transport_chart_data = [
            {'name': 'FCL', 'value': round(transport_profits['FCL'], 2)},
            {'name': 'LCL', 'value': round(transport_profits['LCL'], 2)},
            {'name': 'AIR', 'value': round(transport_profits['AIR'], 2)},
        ]
        
        return Response({
            'resumen': {
                'total_ros': len(profit_data),
                'ingresos_totales_usd': round(total_revenue, 2),
                'costos_totales_usd': round(total_costs, 2),
                'margen_total_usd': round(total_margin, 2),
                'margen_promedio_porcentaje': round(avg_margin_pct, 2),
                'promedio_profit_por_cotizacion': round(avg_profit_per_quote, 2),
            },
            'charts': {
                'monthly_profits': monthly_chart_data,
                'transport_breakdown': transport_chart_data,
            },
            'ros': profit_data,
            'export_available': True
        })


class MasterAdminLogsView(APIView):
    """
    Activity Logs Viewer with filtering support.
    Supports filtering by: search, action_type, date_from, date_to, user_id, level
    Also returns available action types for filter dropdown.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        from .models import ActivityLog
        from accounts.models import CustomUser
        
        search = request.query_params.get('search', '').strip()
        action_type = request.query_params.get('action_type', '').strip()
        date_from = request.query_params.get('date_from', '').strip()
        date_to = request.query_params.get('date_to', '').strip()
        user_id = request.query_params.get('user_id', '').strip()
        level = request.query_params.get('level', '').strip()
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 50))
        
        logs_qs = ActivityLog.objects.all()
        
        if search:
            logs_qs = logs_qs.filter(Q(message__icontains=search) | Q(user_email__icontains=search))
        
        if action_type:
            logs_qs = logs_qs.filter(action_type=action_type)
        
        if level:
            logs_qs = logs_qs.filter(level=level)
        
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d')
                logs_qs = logs_qs.filter(created_at__date__gte=from_date.date())
            except ValueError:
                pass
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d')
                logs_qs = logs_qs.filter(created_at__date__lte=to_date.date())
            except ValueError:
                pass
        
        if user_id:
            try:
                logs_qs = logs_qs.filter(user_id=int(user_id))
            except ValueError:
                pass
        
        total_count = logs_qs.count()
        total_pages = (total_count + page_size - 1) // page_size
        
        offset = (page - 1) * page_size
        logs_page = logs_qs[offset:offset + page_size]
        
        logs_list = []
        for log in logs_page:
            logs_list.append({
                'id': log.id,
                'action_type': log.action_type,
                'action_type_display': dict(ActivityLog.ACTION_TYPE_CHOICES).get(log.action_type, log.action_type),
                'level': log.level,
                'message': log.message,
                'user_id': log.user_id,
                'user_email': log.user_email,
                'ip_address': log.ip_address,
                'related_object_type': log.related_object_type,
                'related_object_id': log.related_object_id,
                'created_at': log.created_at.isoformat() if log.created_at else None,
            })
        
        action_types = ActivityLog.get_action_types()
        
        level_choices = [{'value': k, 'label': v} for k, v in ActivityLog.LEVEL_CHOICES]
        
        api_status = {
            'database': 'connected',
            'celery': 'unknown',
            'email_service': 'mock',
            'senae_api': 'not_configured',
            'tracking_api': 'mock',
        }
        
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
            api_status['database'] = 'connected'
        except Exception:
            api_status['database'] = 'disconnected'
        
        return Response({
            'logs': logs_list,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_count': total_count,
                'total_pages': total_pages,
            },
            'filters': {
                'action_types': action_types,
                'levels': level_choices,
            },
            'api_status': api_status,
            'timestamp': timezone.now().isoformat()
        })


class MasterAdminExportView(APIView):
    """
    Export data to CSV/Excel format.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        export_type = request.query_params.get('type', 'profit')
        format_type = request.query_params.get('format', 'csv')
        
        if export_type == 'profit':
            ros = LeadCotizacion.objects.filter(ro_number__isnull=False)
            
            csv_data = "RO Number,Cotizacion,Cliente,Origen,Destino,Total USD,Margen Estimado USD,Fecha\n"
            for ro in ros:
                total = float(ro.total_usd) if ro.total_usd else 0
                margin = total * 0.25
                csv_data += f"{ro.ro_number},{ro.numero_cotizacion},{ro.lead_user.email if ro.lead_user else 'N/A'},\"{ro.origen_ciudad}, {ro.origen_pais}\",{ro.destino_ciudad},{total},{round(margin, 2)},{ro.fecha_creacion.strftime('%Y-%m-%d')}\n"
            
            return Response({
                'success': True,
                'format': 'csv',
                'filename': f'profit_report_{timezone.now().strftime("%Y%m%d")}.csv',
                'data': csv_data
            })
        
        return Response({'error': 'Tipo de exportación no válido'}, status=400)


class MasterAdminPortsView(APIView):
    """
    Full CRUD access to World Ports database.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        port_id = request.query_params.get('id')
        
        if port_id:
            try:
                port = Port.objects.get(id=port_id)
                return Response({
                    'id': port.id,
                    'un_locode': port.un_locode,
                    'name': port.name,
                    'country': port.country,
                    'region': port.region,
                    'is_active': port.is_active,
                    'created_at': port.created_at,
                    'updated_at': port.updated_at,
                })
            except Port.DoesNotExist:
                return Response({'error': 'Puerto no encontrado'}, status=404)
        
        ports = Port.objects.all().order_by('region', 'country', 'name')
        
        region_filter = request.query_params.get('region')
        if region_filter:
            ports = ports.filter(region=region_filter)
        
        country_filter = request.query_params.get('country')
        if country_filter:
            ports = ports.filter(country__icontains=country_filter)
        
        search = request.query_params.get('search')
        if search:
            ports = ports.filter(
                Q(name__icontains=search) |
                Q(un_locode__icontains=search) |
                Q(country__icontains=search)
            )
        
        regions_summary = Port.objects.values('region').annotate(count=Count('id')).order_by('region')
        
        return Response({
            'total': ports.count(),
            'regions_summary': list(regions_summary),
            'ports': [
                {
                    'id': p.id,
                    'un_locode': p.un_locode,
                    'name': p.name,
                    'country': p.country,
                    'region': p.region,
                    'is_active': p.is_active,
                }
                for p in ports[:200]
            ]
        })
    
    def post(self, request):
        un_locode = request.data.get('un_locode', '').upper()
        name = request.data.get('name', '')
        country = request.data.get('country', '')
        region = request.data.get('region', '')
        
        if not un_locode or not name or not country or not region:
            return Response({'error': 'Todos los campos son requeridos'}, status=400)
        
        if Port.objects.filter(un_locode=un_locode).exists():
            return Response({'error': f'El código UN/LOCODE {un_locode} ya existe'}, status=400)
        
        port = Port.objects.create(
            un_locode=un_locode,
            name=name,
            country=country,
            region=region,
            is_active=request.data.get('is_active', True)
        )
        
        return Response({
            'success': True,
            'message': f'Puerto {name} ({un_locode}) creado exitosamente',
            'id': port.id
        })
    
    def put(self, request):
        port_id = request.data.get('id')
        if not port_id:
            return Response({'error': 'ID de puerto requerido'}, status=400)
        
        try:
            port = Port.objects.get(id=port_id)
            
            updatable_fields = ['un_locode', 'name', 'country', 'region', 'is_active']
            for field in updatable_fields:
                if field in request.data:
                    value = request.data[field]
                    if field == 'un_locode':
                        value = value.upper()
                    setattr(port, field, value)
            
            port.save()
            return Response({'success': True, 'message': 'Puerto actualizado'})
        except Port.DoesNotExist:
            return Response({'error': 'Puerto no encontrado'}, status=404)
    
    def delete(self, request):
        port_id = request.query_params.get('id')
        if not port_id:
            return Response({'error': 'ID de puerto requerido'}, status=400)
        
        try:
            port = Port.objects.get(id=port_id)
            port_name = port.name
            port.delete()
            return Response({'success': True, 'message': f'Puerto {port_name} eliminado'})
        except Port.DoesNotExist:
            return Response({'error': 'Puerto no encontrado'}, status=404)


class MasterAdminAirportsView(APIView):
    """
    Full CRUD access to World Airports database.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        airport_id = request.query_params.get('id')
        
        if airport_id:
            try:
                airport = Airport.objects.get(id=airport_id)
                return Response({
                    'id': airport.id,
                    'iata_code': airport.iata_code,
                    'icao_code': airport.icao_code,
                    'name': airport.name,
                    'ciudad_exacta': airport.ciudad_exacta,
                    'country': airport.country,
                    'region_name': airport.region_name,
                    'latitude': float(airport.latitude) if airport.latitude else None,
                    'longitude': float(airport.longitude) if airport.longitude else None,
                    'timezone': airport.timezone,
                    'is_active': airport.is_active,
                })
            except Airport.DoesNotExist:
                return Response({'error': 'Aeropuerto no encontrado'}, status=404)
        
        airports = Airport.objects.all().order_by('region_name', 'country', 'ciudad_exacta')
        
        region_filter = request.query_params.get('region')
        if region_filter:
            airports = airports.filter(region_name__icontains=region_filter)
        
        country_filter = request.query_params.get('country')
        if country_filter:
            airports = airports.filter(country__icontains=country_filter)
        
        search = request.query_params.get('search')
        if search:
            airports = airports.filter(
                Q(name__icontains=search) |
                Q(iata_code__icontains=search) |
                Q(ciudad_exacta__icontains=search) |
                Q(country__icontains=search)
            )
        
        regions_summary = Airport.objects.values('region_name').annotate(count=Count('id')).order_by('region_name')
        
        return Response({
            'total': airports.count(),
            'regions_summary': list(regions_summary),
            'airports': [
                {
                    'id': a.id,
                    'iata_code': a.iata_code,
                    'icao_code': a.icao_code,
                    'name': a.name,
                    'ciudad_exacta': a.ciudad_exacta,
                    'country': a.country,
                    'region_name': a.region_name,
                    'is_active': a.is_active,
                }
                for a in airports[:200]
            ]
        })
    
    def post(self, request):
        iata_code = request.data.get('iata_code', '').upper()
        name = request.data.get('name', '')
        ciudad_exacta = request.data.get('ciudad_exacta', '')
        country = request.data.get('country', '')
        region_name = request.data.get('region_name', '')
        
        if not iata_code or not name or not ciudad_exacta or not country or not region_name:
            return Response({'error': 'Los campos iata_code, name, ciudad_exacta, country y region_name son requeridos'}, status=400)
        
        if Airport.objects.filter(iata_code=iata_code).exists():
            return Response({'error': f'El código IATA {iata_code} ya existe'}, status=400)
        
        airport = Airport.objects.create(
            iata_code=iata_code,
            icao_code=request.data.get('icao_code', '').upper(),
            name=name,
            ciudad_exacta=ciudad_exacta,
            country=country,
            region_name=region_name,
            latitude=request.data.get('latitude'),
            longitude=request.data.get('longitude'),
            timezone=request.data.get('timezone', ''),
            is_active=request.data.get('is_active', True)
        )
        
        return Response({
            'success': True,
            'message': f'Aeropuerto {name} ({iata_code}) creado exitosamente',
            'id': airport.id
        })
    
    def put(self, request):
        airport_id = request.data.get('id')
        if not airport_id:
            return Response({'error': 'ID de aeropuerto requerido'}, status=400)
        
        try:
            airport = Airport.objects.get(id=airport_id)
            
            updatable_fields = ['iata_code', 'icao_code', 'name', 'ciudad_exacta', 'country', 
                               'region_name', 'latitude', 'longitude', 'timezone', 'is_active']
            for field in updatable_fields:
                if field in request.data:
                    value = request.data[field]
                    if field in ['iata_code', 'icao_code']:
                        value = value.upper() if value else ''
                    setattr(airport, field, value)
            
            airport.save()
            return Response({'success': True, 'message': 'Aeropuerto actualizado'})
        except Airport.DoesNotExist:
            return Response({'error': 'Aeropuerto no encontrado'}, status=404)
    
    def delete(self, request):
        airport_id = request.query_params.get('id')
        if not airport_id:
            return Response({'error': 'ID de aeropuerto requerido'}, status=400)
        
        try:
            airport = Airport.objects.get(id=airport_id)
            airport_name = airport.name
            airport.delete()
            return Response({'success': True, 'message': f'Aeropuerto {airport_name} eliminado'})
        except Airport.DoesNotExist:
            return Response({'error': 'Aeropuerto no encontrado'}, status=404)


class MasterAdminProvidersView(APIView):
    """
    Full CRUD access to Logistics Providers database.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        provider_id = request.query_params.get('id')
        
        if provider_id:
            try:
                provider = LogisticsProvider.objects.get(id=provider_id)
                rates = ProviderRate.objects.filter(provider=provider)
                return Response({
                    'id': provider.id,
                    'name': provider.name,
                    'code': provider.code,
                    'transport_type': provider.transport_type,
                    'contact_name': provider.contact_name,
                    'contact_email': provider.contact_email,
                    'contact_phone': provider.contact_phone,
                    'website': provider.website,
                    'notes': provider.notes,
                    'priority': provider.priority,
                    'is_active': provider.is_active,
                    'rates_count': rates.count(),
                    'rates': [
                        {
                            'id': r.id,
                            'origin_port': r.origin_port,
                            'destination': r.destination,
                            'container_type': r.container_type,
                            'rate_usd': float(r.rate_usd) if r.rate_usd else 0,
                            'unit': r.unit,
                            'transit_days': r.transit_days,
                            'is_active': r.is_active,
                        }
                        for r in rates[:50]
                    ]
                })
            except LogisticsProvider.DoesNotExist:
                return Response({'error': 'Proveedor no encontrado'}, status=404)
        
        providers = LogisticsProvider.objects.all().order_by('transport_type', 'priority', 'name')
        
        transport_filter = request.query_params.get('transport_type')
        if transport_filter:
            providers = providers.filter(transport_type=transport_filter)
        
        search = request.query_params.get('search')
        if search:
            providers = providers.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search)
            )
        
        type_summary = LogisticsProvider.objects.values('transport_type').annotate(count=Count('id')).order_by('transport_type')
        
        return Response({
            'total': providers.count(),
            'type_summary': list(type_summary),
            'providers': [
                {
                    'id': p.id,
                    'name': p.name,
                    'code': p.code,
                    'transport_type': p.transport_type,
                    'contact_email': p.contact_email,
                    'priority': p.priority,
                    'is_active': p.is_active,
                    'rates_count': ProviderRate.objects.filter(provider=p).count(),
                }
                for p in providers[:100]
            ]
        })
    
    def post(self, request):
        name = request.data.get('name', '')
        code = request.data.get('code', '').upper()
        transport_type = request.data.get('transport_type', '')
        
        if not name or not code or not transport_type:
            return Response({'error': 'name, code y transport_type son requeridos'}, status=400)
        
        if LogisticsProvider.objects.filter(code=code).exists():
            return Response({'error': f'El código {code} ya existe'}, status=400)
        
        provider = LogisticsProvider.objects.create(
            name=name,
            code=code,
            transport_type=transport_type,
            contact_name=request.data.get('contact_name', ''),
            contact_email=request.data.get('contact_email', ''),
            contact_phone=request.data.get('contact_phone', ''),
            website=request.data.get('website', ''),
            notes=request.data.get('notes', ''),
            priority=request.data.get('priority', 5),
            is_active=request.data.get('is_active', True)
        )
        
        return Response({
            'success': True,
            'message': f'Proveedor {name} ({code}) creado exitosamente',
            'id': provider.id
        })
    
    def put(self, request):
        provider_id = request.data.get('id')
        if not provider_id:
            return Response({'error': 'ID de proveedor requerido'}, status=400)
        
        try:
            provider = LogisticsProvider.objects.get(id=provider_id)
            
            updatable_fields = ['name', 'code', 'transport_type', 'contact_name', 'contact_email',
                               'contact_phone', 'website', 'notes', 'priority', 'is_active']
            for field in updatable_fields:
                if field in request.data:
                    value = request.data[field]
                    if field == 'code':
                        value = value.upper()
                    setattr(provider, field, value)
            
            provider.save()
            return Response({'success': True, 'message': 'Proveedor actualizado'})
        except LogisticsProvider.DoesNotExist:
            return Response({'error': 'Proveedor no encontrado'}, status=404)
    
    def delete(self, request):
        provider_id = request.query_params.get('id')
        if not provider_id:
            return Response({'error': 'ID de proveedor requerido'}, status=400)
        
        try:
            provider = LogisticsProvider.objects.get(id=provider_id)
            provider_name = provider.name
            provider.delete()
            return Response({'success': True, 'message': f'Proveedor {provider_name} eliminado'})
        except LogisticsProvider.DoesNotExist:
            return Response({'error': 'Proveedor no encontrado'}, status=404)


class MasterAdminProviderRatesView(APIView):
    """
    Full CRUD access to Provider Rates database.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        rate_id = request.query_params.get('id')
        provider_id = request.query_params.get('provider_id')
        
        if rate_id:
            try:
                rate = ProviderRate.objects.get(id=rate_id)
                return Response({
                    'id': rate.id,
                    'provider_id': rate.provider_id,
                    'provider_name': rate.provider.name if rate.provider else None,
                    'origin_port': rate.origin_port,
                    'destination': rate.destination,
                    'container_type': rate.container_type,
                    'rate_usd': float(rate.rate_usd) if rate.rate_usd else 0,
                    'unit': rate.unit,
                    'transit_days': rate.transit_days,
                    'valid_from': rate.valid_from,
                    'valid_until': rate.valid_until,
                    'notes': rate.notes,
                    'is_active': rate.is_active,
                })
            except ProviderRate.DoesNotExist:
                return Response({'error': 'Tarifa no encontrada'}, status=404)
        
        rates = ProviderRate.objects.all().select_related('provider').order_by('provider__name', 'origin_port')
        
        if provider_id:
            rates = rates.filter(provider_id=provider_id)
        
        origin_filter = request.query_params.get('origin_port')
        if origin_filter:
            rates = rates.filter(origin_port__icontains=origin_filter)
        
        container_filter = request.query_params.get('container_type')
        if container_filter:
            rates = rates.filter(container_type=container_filter)
        
        return Response({
            'total': rates.count(),
            'rates': [
                {
                    'id': r.id,
                    'provider_id': r.provider_id,
                    'provider_name': r.provider.name if r.provider else None,
                    'provider_code': r.provider.code if r.provider else None,
                    'origin_port': r.origin_port,
                    'destination': r.destination,
                    'container_type': r.container_type,
                    'rate_usd': float(r.rate_usd) if r.rate_usd else 0,
                    'unit': r.unit,
                    'transit_days': r.transit_days,
                    'is_active': r.is_active,
                }
                for r in rates[:200]
            ]
        })
    
    def post(self, request):
        provider_id = request.data.get('provider_id')
        origin_port = request.data.get('origin_port', '')
        destination = request.data.get('destination', '')
        rate_usd = request.data.get('rate_usd')
        
        if not provider_id or not origin_port or not destination or rate_usd is None:
            return Response({'error': 'provider_id, origin_port, destination y rate_usd son requeridos'}, status=400)
        
        try:
            provider = LogisticsProvider.objects.get(id=provider_id)
        except LogisticsProvider.DoesNotExist:
            return Response({'error': 'Proveedor no encontrado'}, status=404)
        
        rate = ProviderRate.objects.create(
            provider=provider,
            origin_port=origin_port,
            destination=destination,
            container_type=request.data.get('container_type', ''),
            rate_usd=rate_usd,
            unit=request.data.get('unit', 'CONTAINER'),
            transit_days=request.data.get('transit_days'),
            valid_from=request.data.get('valid_from'),
            valid_until=request.data.get('valid_until'),
            notes=request.data.get('notes', ''),
            is_active=request.data.get('is_active', True)
        )
        
        return Response({
            'success': True,
            'message': f'Tarifa para {provider.name} creada exitosamente',
            'id': rate.id
        })
    
    def put(self, request):
        rate_id = request.data.get('id')
        if not rate_id:
            return Response({'error': 'ID de tarifa requerido'}, status=400)
        
        try:
            rate = ProviderRate.objects.get(id=rate_id)
            
            updatable_fields = ['origin_port', 'destination', 'container_type', 'rate_usd',
                               'unit', 'transit_days', 'valid_from', 'valid_until', 'notes', 'is_active']
            for field in updatable_fields:
                if field in request.data:
                    setattr(rate, field, request.data[field])
            
            rate.save()
            return Response({'success': True, 'message': 'Tarifa actualizada'})
        except ProviderRate.DoesNotExist:
            return Response({'error': 'Tarifa no encontrada'}, status=404)
    
    def delete(self, request):
        rate_id = request.query_params.get('id')
        if not rate_id:
            return Response({'error': 'ID de tarifa requerido'}, status=400)
        
        try:
            rate = ProviderRate.objects.get(id=rate_id)
            rate.delete()
            return Response({'success': True, 'message': 'Tarifa eliminada'})
        except ProviderRate.DoesNotExist:
            return Response({'error': 'Tarifa no encontrada'}, status=404)


class MasterAdminFreightRateFCLView(APIView):
    """
    Full CRUD access to unified freight rates (FCL, LCL, AEREO).
    Includes 1,775 rates imported from carrier contracts.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        rate_id = request.query_params.get('id')
        
        if rate_id:
            try:
                rate = FreightRateFCL.objects.get(id=rate_id)
                serializer = FreightRateFCLSerializer(rate)
                return Response(serializer.data)
            except FreightRateFCL.DoesNotExist:
                return Response({'error': 'Tarifa no encontrada'}, status=404)
        
        rates = FreightRateFCL.objects.all()
        
        transport_type = request.query_params.get('transport_type')
        if transport_type:
            rates = rates.filter(transport_type__icontains=transport_type)
        
        pol = request.query_params.get('pol')
        if pol:
            rates = rates.filter(pol_name__icontains=pol)
        
        pod = request.query_params.get('pod')
        if pod:
            rates = rates.filter(pod_name__icontains=pod)
        
        carrier = request.query_params.get('carrier')
        if carrier:
            rates = rates.filter(carrier_name__icontains=carrier)
        
        is_active = request.query_params.get('is_active')
        if is_active is not None:
            rates = rates.filter(is_active=is_active.lower() == 'true')
        
        rates = rates.order_by('transport_type', 'pol_name', 'pod_name', '-validity_date')
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 50))
        start = (page - 1) * page_size
        end = start + page_size
        
        total = rates.count()
        rates_page = rates[start:end]
        
        serializer = FreightRateFCLListSerializer(rates_page, many=True)
        
        summary = FreightRateFCL.objects.values('transport_type').annotate(
            count=Count('id')
        ).order_by('transport_type')
        
        return Response({
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'summary_by_type': list(summary),
            'rates': serializer.data
        })
    
    def post(self, request):
        serializer = FreightRateFCLSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'message': 'Tarifa creada exitosamente',
                'id': serializer.data['id']
            })
        return Response({'error': serializer.errors}, status=400)
    
    def put(self, request):
        rate_id = request.data.get('id')
        if not rate_id:
            return Response({'error': 'ID de tarifa requerido'}, status=400)
        
        try:
            rate = FreightRateFCL.objects.get(id=rate_id)
            serializer = FreightRateFCLSerializer(rate, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'success': True, 'message': 'Tarifa actualizada'})
            return Response({'error': serializer.errors}, status=400)
        except FreightRateFCL.DoesNotExist:
            return Response({'error': 'Tarifa no encontrada'}, status=404)
    
    def delete(self, request):
        rate_id = request.query_params.get('id')
        if not rate_id:
            return Response({'error': 'ID de tarifa requerido'}, status=400)
        
        try:
            rate = FreightRateFCL.objects.get(id=rate_id)
            rate.delete()
            return Response({'success': True, 'message': 'Tarifa eliminada'})
        except FreightRateFCL.DoesNotExist:
            return Response({'error': 'Tarifa no encontrada'}, status=404)


class MasterAdminProfitMarginView(APIView):
    """
    Full CRUD access to profit margin configuration.
    Manages margins by transport type and item type.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        margin_id = request.query_params.get('id')
        
        if margin_id:
            try:
                margin = ProfitMarginConfig.objects.get(id=margin_id)
                serializer = ProfitMarginConfigSerializer(margin)
                return Response(serializer.data)
            except ProfitMarginConfig.DoesNotExist:
                return Response({'error': 'Configuración de margen no encontrada'}, status=404)
        
        margins = ProfitMarginConfig.objects.all()
        
        transport_type = request.query_params.get('transport_type')
        if transport_type:
            margins = margins.filter(transport_type=transport_type)
        
        item_type = request.query_params.get('item_type')
        if item_type:
            margins = margins.filter(item_type=item_type)
        
        is_active = request.query_params.get('is_active')
        if is_active is not None:
            margins = margins.filter(is_active=is_active.lower() == 'true')
        
        margins = margins.order_by('priority', 'transport_type', 'item_type')
        
        serializer = ProfitMarginConfigSerializer(margins, many=True)
        
        return Response({
            'total': margins.count(),
            'transport_type_choices': [
                {'value': k, 'label': v} 
                for k, v in ProfitMarginConfig.TRANSPORT_TYPE_CHOICES
            ],
            'item_type_choices': [
                {'value': k, 'label': v} 
                for k, v in ProfitMarginConfig.ITEM_TYPE_CHOICES
            ],
            'margin_type_choices': [
                {'value': k, 'label': v} 
                for k, v in ProfitMarginConfig.MARGIN_TYPE_CHOICES
            ],
            'margins': serializer.data
        })
    
    def post(self, request):
        serializer = ProfitMarginConfigSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'message': 'Configuración de margen creada exitosamente',
                'id': serializer.data['id']
            })
        return Response({'error': serializer.errors}, status=400)
    
    def put(self, request):
        margin_id = request.data.get('id')
        if not margin_id:
            return Response({'error': 'ID de configuración requerido'}, status=400)
        
        try:
            margin = ProfitMarginConfig.objects.get(id=margin_id)
            serializer = ProfitMarginConfigSerializer(margin, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'success': True, 'message': 'Configuración actualizada'})
            return Response({'error': serializer.errors}, status=400)
        except ProfitMarginConfig.DoesNotExist:
            return Response({'error': 'Configuración no encontrada'}, status=404)
    
    def delete(self, request):
        margin_id = request.query_params.get('id')
        if not margin_id:
            return Response({'error': 'ID de configuración requerido'}, status=400)
        
        try:
            margin = ProfitMarginConfig.objects.get(id=margin_id)
            margin.delete()
            return Response({'success': True, 'message': 'Configuración eliminada'})
        except ProfitMarginConfig.DoesNotExist:
            return Response({'error': 'Configuración no encontrada'}, status=404)


class MasterAdminLocalCostView(APIView):
    """
    Full CRUD access to local destination costs (THC, handling, fees, etc).
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        cost_id = request.query_params.get('id')
        
        if cost_id:
            try:
                cost = LocalDestinationCost.objects.get(id=cost_id)
                serializer = LocalDestinationCostSerializer(cost)
                return Response(serializer.data)
            except LocalDestinationCost.DoesNotExist:
                return Response({'error': 'Gasto local no encontrado'}, status=404)
        
        costs = LocalDestinationCost.objects.all()
        
        transport_type = request.query_params.get('transport_type')
        if transport_type:
            costs = costs.filter(transport_type=transport_type)
        
        cost_type = request.query_params.get('cost_type')
        if cost_type:
            costs = costs.filter(cost_type=cost_type)
        
        port = request.query_params.get('port')
        if port:
            costs = costs.filter(port=port)
        
        carrier_code = request.query_params.get('carrier_code')
        if carrier_code:
            costs = costs.filter(carrier_code=carrier_code)
        
        is_active = request.query_params.get('is_active')
        if is_active is not None:
            costs = costs.filter(is_active=is_active.lower() == 'true')
        
        costs = costs.order_by('transport_type', 'carrier_code', 'cost_type', 'port')
        
        serializer = LocalDestinationCostSerializer(costs, many=True)
        
        return Response({
            'total': costs.count(),
            'transport_type_choices': [
                {'value': k, 'label': v} 
                for k, v in LocalDestinationCost.TRANSPORT_TYPE_CHOICES
            ],
            'cost_type_choices': [
                {'value': k, 'label': v} 
                for k, v in LocalDestinationCost.COST_TYPE_CHOICES
            ],
            'port_choices': [
                {'value': k, 'label': v} 
                for k, v in LocalDestinationCost.PORT_CHOICES
            ],
            'container_type_choices': [
                {'value': k, 'label': v} 
                for k, v in LocalDestinationCost.CONTAINER_TYPE_CHOICES
            ],
            'costs': serializer.data
        })
    
    def post(self, request):
        serializer = LocalDestinationCostSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'message': 'Gasto local creado exitosamente',
                'id': serializer.data['id']
            })
        return Response({'error': serializer.errors}, status=400)
    
    def put(self, request):
        cost_id = request.data.get('id')
        if not cost_id:
            return Response({'error': 'ID de gasto local requerido'}, status=400)
        
        try:
            cost = LocalDestinationCost.objects.get(id=cost_id)
            serializer = LocalDestinationCostSerializer(cost, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'success': True, 'message': 'Gasto local actualizado'})
            return Response({'error': serializer.errors}, status=400)
        except LocalDestinationCost.DoesNotExist:
            return Response({'error': 'Gasto local no encontrado'}, status=404)
    
    def delete(self, request):
        cost_id = request.query_params.get('id')
        if not cost_id:
            return Response({'error': 'ID de gasto local requerido'}, status=400)
        
        try:
            cost = LocalDestinationCost.objects.get(id=cost_id)
            cost.delete()
            return Response({'success': True, 'message': 'Gasto local eliminado'})
        except LocalDestinationCost.DoesNotExist:
            return Response({'error': 'Gasto local no encontrado'}, status=404)


class MasterAdminTrackingView(APIView):
    """
    Cargo Tracking management with CSV bulk import/export.
    BLOQUE 4: Sistema de Ingesta de Datos para el Master Admin.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    MILESTONE_COLUMNS = [
        ('SI_ENVIADA', 'SI Enviada'),
        ('SI_RECIBIDA_FORWARDER', 'SI Recibida Forwarder'),
        ('SHIPPER_CONTACTADO', 'Shipper Contactado'),
        ('BOOKING_ORDER_RECIBIDA', 'Booking Order Recibida'),
        ('ETD_SOLICITADO', 'ETD Solicitado'),
        ('BOOKING_CONFIRMADO', 'Booking Confirmado'),
        ('SO_LIBERADO', 'S/O Liberado'),
        ('RETIRO_CARGUE_VACIOS', 'Retiro Cargue Vacios'),
        ('CONTENEDORES_CARGADOS', 'Contenedores Cargados'),
        ('GATE_IN', 'Gate In'),
        ('ETD', 'ETD'),
        ('ATD', 'ATD'),
        ('TRANSHIPMENT', 'Transhipment'),
        ('ETA_DESTINO', 'ETA Destino'),
    ]
    
    def get(self, request):
        action = request.query_params.get('action', 'list')
        
        if action == 'template':
            return self._generate_csv_template()
        elif action == 'list':
            return self._list_tracking_ros()
        else:
            return Response({'error': 'Acción no válida'}, status=400)
    
    def _generate_csv_template(self):
        """GET /api/admin/tracking/?action=template - Genera plantilla CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="tracking_milestones_template.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        headers = ['RO_NUMBER', 'COMPANY_NAME']
        for key, label in self.MILESTONE_COLUMNS:
            headers.append(f'ACTUAL_{key}')
            headers.append(f'PLANNED_{key}')
        
        headers.extend([
            'BOOKING_CUTOFF', 'DOC_CUTOFF', 'VGM_CUTOFF', 'PORT_CUTOFF',
            'NOTES'
        ])
        
        writer.writerow(headers)
        
        active_ros = ShippingInstruction.objects.filter(
            ro_number__isnull=False,
            status__in=['ro_generated', 'sent_to_forwarder', 'confirmed', 'in_transit']
        ).select_related('quote_submission')
        
        for si in active_ros:
            row = [
                si.ro_number,
                si.quote_submission.company_name if si.quote_submission else ''
            ]
            
            milestones = {m.milestone_key: m for m in si.milestones.all()}
            
            for key, label in self.MILESTONE_COLUMNS:
                milestone = milestones.get(key)
                if milestone:
                    row.append(milestone.actual_date.strftime('%Y-%m-%d %H:%M') if milestone.actual_date else '')
                    row.append(milestone.planned_date.strftime('%Y-%m-%d %H:%M') if milestone.planned_date else '')
                else:
                    row.extend(['', ''])
            
            booking_milestone = milestones.get('BOOKING_CONFIRMADO')
            if booking_milestone and booking_milestone.meta_data:
                row.append(booking_milestone.meta_data.get('booking_cutoff', ''))
                row.append(booking_milestone.meta_data.get('doc_cutoff', ''))
                row.append(booking_milestone.meta_data.get('vgm_cutoff', ''))
                row.append(booking_milestone.meta_data.get('port_cutoff', ''))
            else:
                row.extend(['', '', '', ''])
            
            row.append('')
            writer.writerow(row)
        
        return response
    
    def _list_tracking_ros(self):
        """Lista todos los ROs con tracking"""
        ros = ShippingInstruction.objects.filter(
            ro_number__isnull=False
        ).select_related('quote_submission').prefetch_related('milestones').order_by('-created_at')
        
        data = []
        for si in ros[:100]:
            milestones = si.milestones.all()
            completed_count = milestones.filter(status='COMPLETED').count()
            current = ShipmentMilestone.get_current_milestone(si)
            
            data.append({
                'id': si.id,
                'ro_number': si.ro_number,
                'company_name': si.quote_submission.company_name if si.quote_submission else '',
                'origin': si.quote_submission.origin if si.quote_submission else '',
                'destination': si.quote_submission.destination if si.quote_submission else '',
                'status': si.status,
                'current_milestone': current.get_milestone_key_display() if current else 'Sin iniciar',
                'progress': f"{completed_count}/14",
                'progress_pct': round((completed_count / 14) * 100),
                'created_at': si.created_at,
            })
        
        return Response({
            'total': ros.count(),
            'ros': data
        })
    
    def post(self, request):
        """POST /api/admin/tracking/ - Importa CSV con datos de milestones"""
        csv_file = request.FILES.get('file')
        if not csv_file:
            return Response({'error': 'Archivo CSV requerido'}, status=400)
        
        try:
            decoded = csv_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            
            processed = 0
            errors = []
            
            for row_num, row in enumerate(reader, start=2):
                ro_number = row.get('RO_NUMBER', '').strip()
                if not ro_number:
                    continue
                
                try:
                    si = ShippingInstruction.objects.get(ro_number=ro_number)
                except ShippingInstruction.DoesNotExist:
                    errors.append(f"Fila {row_num}: RO '{ro_number}' no encontrado")
                    continue
                
                if not si.milestones.exists():
                    ShipmentMilestone.create_initial_milestones(si)
                
                for key, label in self.MILESTONE_COLUMNS:
                    actual_col = f'ACTUAL_{key}'
                    planned_col = f'PLANNED_{key}'
                    
                    actual_date_str = row.get(actual_col, '').strip()
                    planned_date_str = row.get(planned_col, '').strip()
                    
                    if actual_date_str or planned_date_str:
                        try:
                            milestone = si.milestones.get(milestone_key=key)
                            
                            if actual_date_str:
                                try:
                                    milestone.actual_date = datetime.strptime(actual_date_str, '%Y-%m-%d %H:%M')
                                    milestone.status = 'COMPLETED'
                                except ValueError:
                                    try:
                                        milestone.actual_date = datetime.strptime(actual_date_str, '%Y-%m-%d')
                                        milestone.status = 'COMPLETED'
                                    except ValueError:
                                        pass
                            
                            if planned_date_str:
                                try:
                                    milestone.planned_date = datetime.strptime(planned_date_str, '%Y-%m-%d %H:%M')
                                except ValueError:
                                    try:
                                        milestone.planned_date = datetime.strptime(planned_date_str, '%Y-%m-%d')
                                    except ValueError:
                                        pass
                            
                            milestone.save()
                        except ShipmentMilestone.DoesNotExist:
                            pass
                
                booking_milestone = si.milestones.filter(milestone_key='BOOKING_CONFIRMADO').first()
                if booking_milestone:
                    meta = booking_milestone.meta_data or {}
                    if row.get('BOOKING_CUTOFF'):
                        meta['booking_cutoff'] = row['BOOKING_CUTOFF']
                    if row.get('DOC_CUTOFF'):
                        meta['doc_cutoff'] = row['DOC_CUTOFF']
                    if row.get('VGM_CUTOFF'):
                        meta['vgm_cutoff'] = row['VGM_CUTOFF']
                    if row.get('PORT_CUTOFF'):
                        meta['port_cutoff'] = row['PORT_CUTOFF']
                    booking_milestone.meta_data = meta
                    booking_milestone.save()
                
                ShipmentMilestone.get_current_milestone(si)
                processed += 1
            
            return Response({
                'success': True,
                'message': f'Importación completada: {processed} ROs procesados',
                'processed': processed,
                'errors': errors[:20] if errors else []
            })
            
        except Exception as e:
            logger.error(f"Error importing tracking CSV: {str(e)}")
            return Response({'error': f'Error procesando CSV: {str(e)}'}, status=400)


class FFInvitationManagementView(APIView):
    """
    Master Admin management of Freight Forwarder invitations.
    Create, list, and send invitations for FF portal access.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        """List all FF invitations"""
        invitations = FFInvitation.objects.all().order_by('-created_at')
        
        return Response({
            'success': True,
            'total': invitations.count(),
            'invitations': [
                {
                    'id': inv.id,
                    'email': inv.email,
                    'company_name': inv.company_name,
                    'status': inv.status,
                    'expires_at': inv.expires_at.isoformat(),
                    'is_expired': inv.is_expired,
                    'sent_at': inv.sent_at.isoformat() if inv.sent_at else None,
                    'accepted_at': inv.accepted_at.isoformat() if inv.accepted_at else None,
                    'created_at': inv.created_at.isoformat(),
                }
                for inv in invitations
            ]
        })
    
    def post(self, request):
        """Create and send new FF invitation"""
        email = request.data.get('email')
        company_name = request.data.get('company_name')
        days_valid = request.data.get('days_valid', 7)
        send_email = request.data.get('send_email', True)
        
        if not email or not company_name:
            return Response({
                'success': False,
                'error': 'Email y nombre de empresa son requeridos.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        existing_pending = FFInvitation.objects.filter(
            email=email,
            status__in=['pending', 'sent']
        ).first()
        
        if existing_pending and not existing_pending.is_expired:
            return Response({
                'success': False,
                'error': 'Ya existe una invitación pendiente para este email.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        invitation = FFInvitation.create_invitation(
            email=email,
            company_name=company_name,
            days_valid=days_valid
        )
        
        registration_url = f"/ff-portal/register?token={invitation.token}"
        
        if send_email:
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                subject = 'ImportaYa.ia - Invitación al Portal de Freight Forwarders'
                message = f"""
Estimado/a,

Ha sido invitado/a a acceder al Portal de Freight Forwarders de ImportaYa.ia.

Empresa: {company_name}

Por favor, complete su registro usando el siguiente enlace:
{os.environ.get('REPLIT_DOMAINS', 'http://localhost:5000').split(',')[0]}{registration_url}

Este enlace es válido por {days_valid} días.

Saludos cordiales,
Equipo ImportaYa.ia
                """
                
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=True,
                )
                
                invitation.status = 'sent'
                invitation.sent_at = timezone.now()
                invitation.save()
                
            except Exception as e:
                logger.warning(f"Failed to send FF invitation email: {e}")
        
        return Response({
            'success': True,
            'message': f'Invitación {"enviada" if send_email else "creada"} exitosamente.',
            'invitation': {
                'id': invitation.id,
                'email': invitation.email,
                'company_name': invitation.company_name,
                'token': invitation.token,
                'registration_url': registration_url,
                'expires_at': invitation.expires_at.isoformat(),
                'status': invitation.status,
            }
        }, status=status.HTTP_201_CREATED)
    
    def delete(self, request):
        """Revoke/delete an FF invitation by ID. Only pending or expired invitations can be deleted."""
        invitation_id = request.query_params.get('id')
        
        if not invitation_id:
            return Response({
                'success': False,
                'error': 'ID de invitación es requerido.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            invitation = FFInvitation.objects.get(id=invitation_id)
        except FFInvitation.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Invitación no encontrada.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if invitation.status not in ['pending', 'sent', 'expired'] and not invitation.is_expired:
            return Response({
                'success': False,
                'error': 'Solo se pueden revocar invitaciones pendientes o expiradas.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if invitation.status == 'accepted':
            return Response({
                'success': False,
                'error': 'No se puede revocar una invitación ya aceptada.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        email = invitation.email
        company_name = invitation.company_name
        invitation.delete()
        
        logger.info(f"FF invitation revoked: {email} ({company_name})")
        
        return Response({
            'success': True,
            'message': f'Invitación para {email} revocada exitosamente.'
        })


class FFAssignmentView(APIView):
    """
    Assign Freight Forwarder users to Shipping Instructions.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        """List FF users, unassigned ROs, and assigned ROs"""
        ff_users = CustomUser.objects.filter(role='freight_forwarder', is_active=True)
        unassigned_ros = ShippingInstruction.objects.filter(
            ro_number__isnull=False,
            assigned_ff_user__isnull=True,
            status__in=['ro_generated', 'sent_to_forwarder']
        )
        
        assigned_ros = ShippingInstruction.objects.filter(
            ro_number__isnull=False,
            assigned_ff_user__isnull=False
        ).select_related('assigned_ff_user').order_by('-ff_assignment_date')[:100]
        
        return Response({
            'success': True,
            'ff_users': [
                {
                    'id': user.id,
                    'email': user.email,
                    'company_name': getattr(user, 'ff_profile', None).company_name if hasattr(user, 'ff_profile') and user.ff_profile else '',
                    'assigned_count': ShippingInstruction.objects.filter(assigned_ff_user=user).count(),
                }
                for user in ff_users
            ],
            'unassigned_ros': [
                {
                    'id': si.id,
                    'ro_number': si.ro_number,
                    'consignee_name': si.consignee_name,
                    'status': si.status,
                    'created_at': si.created_at.isoformat(),
                }
                for si in unassigned_ros[:50]
            ],
            'assigned_ros': [
                {
                    'id': si.id,
                    'ro_number': si.ro_number,
                    'consignee_name': si.consignee_name,
                    'status': si.status,
                    'created_at': si.created_at.isoformat() if si.created_at else None,
                    'assigned_ff_id': si.assigned_ff_user.id,
                    'assigned_ff_email': si.assigned_ff_user.email,
                    'assigned_ff_company': getattr(si.assigned_ff_user, 'ff_profile', None).company_name if hasattr(si.assigned_ff_user, 'ff_profile') and si.assigned_ff_user.ff_profile else '',
                    'ff_assignment_date': si.ff_assignment_date.isoformat() if si.ff_assignment_date else None,
                }
                for si in assigned_ros
            ]
        })
    
    def post(self, request):
        """Assign FF user to a Shipping Instruction"""
        ro_id = request.data.get('ro_id')
        ff_user_id = request.data.get('ff_user_id')
        notify = request.data.get('notify', True)
        
        if not ro_id or not ff_user_id:
            return Response({
                'success': False,
                'error': 'RO ID y FF User ID son requeridos.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            si = ShippingInstruction.objects.get(id=ro_id)
            ff_user = CustomUser.objects.get(id=ff_user_id, role='freight_forwarder')
        except ShippingInstruction.DoesNotExist:
            return Response({
                'success': False,
                'error': 'RO no encontrado.'
            }, status=status.HTTP_404_NOT_FOUND)
        except CustomUser.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Usuario FF no encontrado.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        si.assigned_ff_user = ff_user
        si.ff_assignment_date = timezone.now()
        si.save(update_fields=['assigned_ff_user', 'ff_assignment_date'])
        
        if notify:
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                send_mail(
                    f'ImportaYa.ia - Nuevo embarque asignado: {si.ro_number}',
                    f'''
Estimado/a,

Se le ha asignado un nuevo embarque para actualización de tracking:

RO: {si.ro_number}
Consignatario: {si.consignee_name}

Por favor ingrese al Portal de Freight Forwarders para actualizar el estado.

Saludos,
Equipo ImportaYa.ia
                    ''',
                    settings.DEFAULT_FROM_EMAIL,
                    [ff_user.email],
                    fail_silently=True,
                )
            except Exception as e:
                logger.warning(f"Failed to send assignment notification: {e}")
        
        return Response({
            'success': True,
            'message': f'RO {si.ro_number} asignado a {ff_user.email}',
            'assignment': {
                'ro_number': si.ro_number,
                'ff_email': ff_user.email,
                'assigned_at': si.ff_assignment_date.isoformat(),
            }
        })
    
    def put(self, request):
        """Reassign an RO from one FF to another"""
        ro_id = request.data.get('ro_id')
        new_ff_user_id = request.data.get('new_ff_user_id')
        notify = request.data.get('notify', True)
        
        if not ro_id or not new_ff_user_id:
            return Response({
                'success': False,
                'error': 'ro_id y new_ff_user_id son requeridos.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            si = ShippingInstruction.objects.get(id=ro_id)
        except ShippingInstruction.DoesNotExist:
            return Response({
                'success': False,
                'error': 'RO no encontrado.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if not si.assigned_ff_user:
            return Response({
                'success': False,
                'error': 'Este RO no tiene un FF asignado actualmente. Use POST para asignar.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        previous_ff = si.assigned_ff_user
        
        try:
            new_ff_user = CustomUser.objects.get(id=new_ff_user_id, role='freight_forwarder')
        except CustomUser.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Usuario FF no encontrado.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if previous_ff.id == new_ff_user.id:
            return Response({
                'success': False,
                'error': 'El nuevo FF es el mismo que el actual.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        si.assigned_ff_user = new_ff_user
        si.ff_assignment_date = timezone.now()
        si.save(update_fields=['assigned_ff_user', 'ff_assignment_date'])
        
        if notify:
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                send_mail(
                    f'ImportaYa.ia - Embarque reasignado: {si.ro_number}',
                    f'''
Estimado/a,

Se le ha reasignado un embarque para actualización de tracking:

RO: {si.ro_number}
Consignatario: {si.consignee_name}

Este embarque fue previamente gestionado por otro Freight Forwarder.

Por favor ingrese al Portal de Freight Forwarders para revisar y actualizar el estado.

Saludos,
Equipo ImportaYa.ia
                    ''',
                    settings.DEFAULT_FROM_EMAIL,
                    [new_ff_user.email],
                    fail_silently=True,
                )
            except Exception as e:
                logger.warning(f"Failed to send reassignment notification: {e}")
        
        logger.info(f"RO {si.ro_number} reassigned from {previous_ff.email} to {new_ff_user.email}")
        
        return Response({
            'success': True,
            'message': f'RO {si.ro_number} reasignado de {previous_ff.email} a {new_ff_user.email}',
            'reassignment': {
                'ro_number': si.ro_number,
                'previous_ff_email': previous_ff.email,
                'new_ff_email': new_ff_user.email,
                'reassigned_at': si.ff_assignment_date.isoformat(),
            }
        })


class FFConfigView(APIView):
    """
    Freight Forwarder Configuration Management.
    Configure global FF assignment mode (single/multi/manual) and route assignments.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        """Get FF configuration and route assignments"""
        from accounts.models import FFGlobalConfig, FFRouteAssignment
        
        config = FFGlobalConfig.get_config()
        
        ff_profiles = FreightForwarderProfile.objects.filter(
            is_verified=True
        ).select_related('user')
        
        assignments = FFRouteAssignment.objects.filter(
            is_active=True
        ).select_related('freight_forwarder', 'freight_forwarder__user').order_by('priority')
        
        return Response({
            'success': True,
            'config': {
                'assignment_mode': config.assignment_mode,
                'assignment_mode_display': config.get_assignment_mode_display(),
                'default_ff_id': config.default_ff.id if config.default_ff else None,
                'default_ff_name': config.default_ff.company_name if config.default_ff else None,
                'auto_assign_on_ro': config.auto_assign_on_ro,
                'notes': config.notes,
                'updated_at': config.updated_at,
            },
            'available_ffs': [
                {
                    'id': ff.id,
                    'company_name': ff.company_name,
                    'contact_name': ff.contact_name,
                    'email': ff.user.email,
                    'is_verified': ff.is_verified,
                }
                for ff in ff_profiles
            ],
            'route_assignments': [
                {
                    'id': a.id,
                    'ff_id': a.freight_forwarder.id,
                    'ff_name': a.freight_forwarder.company_name,
                    'transport_type': a.transport_type,
                    'transport_type_display': a.get_transport_type_display(),
                    'origin_country': a.origin_country,
                    'origin_port': a.origin_port,
                    'destination_city': a.destination_city,
                    'carrier_name': a.carrier_name,
                    'priority': a.priority,
                    'is_active': a.is_active,
                    'notes': a.notes,
                }
                for a in assignments
            ],
            'mode_options': [
                {'value': 'single', 'label': 'FF Único Global', 'description': 'Un solo FF maneja todos los embarques'},
                {'value': 'multi', 'label': 'Asignación por Criterios', 'description': 'Diferentes FFs según ruta, transporte, naviera, etc.'},
                {'value': 'manual', 'label': 'Asignación Manual', 'description': 'El Master Admin asigna manualmente cada embarque'},
            ],
            'transport_options': [
                {'value': 'all', 'label': 'Todos'},
                {'value': 'FCL', 'label': 'Marítimo FCL'},
                {'value': 'LCL', 'label': 'Marítimo LCL'},
                {'value': 'AEREO', 'label': 'Aéreo'},
            ]
        })
    
    def put(self, request):
        """Update FF global configuration"""
        from accounts.models import FFGlobalConfig
        
        config = FFGlobalConfig.get_config()
        
        if 'assignment_mode' in request.data:
            if request.data['assignment_mode'] in ['single', 'multi', 'manual']:
                config.assignment_mode = request.data['assignment_mode']
        
        if 'default_ff_id' in request.data:
            ff_id = request.data['default_ff_id']
            if ff_id:
                try:
                    config.default_ff = FreightForwarderProfile.objects.get(id=ff_id)
                except FreightForwarderProfile.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'FF no encontrado'
                    }, status=404)
            else:
                config.default_ff = None
        
        if 'auto_assign_on_ro' in request.data:
            config.auto_assign_on_ro = bool(request.data['auto_assign_on_ro'])
        
        if 'notes' in request.data:
            config.notes = request.data['notes']
        
        config.save()
        
        return Response({
            'success': True,
            'message': 'Configuración actualizada correctamente',
            'config': {
                'assignment_mode': config.assignment_mode,
                'default_ff_id': config.default_ff.id if config.default_ff else None,
                'auto_assign_on_ro': config.auto_assign_on_ro,
            }
        })
    
    def post(self, request):
        """Create new route assignment"""
        from accounts.models import FFRouteAssignment
        
        ff_id = request.data.get('ff_id')
        if not ff_id:
            return Response({
                'success': False,
                'error': 'FF ID es requerido'
            }, status=400)
        
        try:
            ff = FreightForwarderProfile.objects.get(id=ff_id)
        except FreightForwarderProfile.DoesNotExist:
            return Response({
                'success': False,
                'error': 'FF no encontrado'
            }, status=404)
        
        assignment = FFRouteAssignment.objects.create(
            freight_forwarder=ff,
            transport_type=request.data.get('transport_type', 'all'),
            origin_country=request.data.get('origin_country', ''),
            origin_port=request.data.get('origin_port', ''),
            destination_city=request.data.get('destination_city', ''),
            carrier_name=request.data.get('carrier_name', ''),
            priority=request.data.get('priority', 1),
            notes=request.data.get('notes', ''),
            is_active=True,
        )
        
        return Response({
            'success': True,
            'message': 'Asignación de ruta creada',
            'assignment': {
                'id': assignment.id,
                'ff_name': ff.company_name,
                'transport_type': assignment.transport_type,
            }
        }, status=201)
    
    def delete(self, request):
        """Delete route assignment"""
        from accounts.models import FFRouteAssignment
        
        assignment_id = request.query_params.get('id')
        if not assignment_id:
            return Response({
                'success': False,
                'error': 'ID de asignación requerido'
            }, status=400)
        
        try:
            assignment = FFRouteAssignment.objects.get(id=assignment_id)
            assignment.delete()
            return Response({
                'success': True,
                'message': 'Asignación eliminada'
            })
        except FFRouteAssignment.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Asignación no encontrada'
            }, status=404)


class HSCodeManagementView(APIView):
    """
    CRUD for HS Code entries (partidas arancelarias)
    GET: List all with search/filter
    POST: Create new entry
    PUT: Update entry
    DELETE: Delete entry
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        from SalesModule.models import HSCodeEntry
        
        search = request.query_params.get('search', '')
        category = request.query_params.get('category', '')
        institution = request.query_params.get('institution', '')
        include_inactive = request.query_params.get('include_inactive', 'false').lower() == 'true'
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        
        if include_inactive:
            queryset = HSCodeEntry.objects.all()
        else:
            queryset = HSCodeEntry.objects.filter(is_active=True)
        
        if search:
            from django.db.models import Q
            search_lower = search.lower()
            queryset = queryset.filter(
                Q(hs_code__icontains=search_lower) |
                Q(description__icontains=search_lower) |
                Q(description_en__icontains=search_lower) |
                Q(keywords__icontains=search_lower) |
                Q(category__icontains=search_lower)
            )
        
        if category:
            queryset = queryset.filter(category__icontains=category)
        
        if institution:
            queryset = queryset.filter(permit_institution=institution)
        
        total = queryset.count()
        start = (page - 1) * per_page
        end = start + per_page
        entries = queryset[start:end]
        
        categories = HSCodeEntry.objects.values_list('category', flat=True).distinct()
        categories = [c for c in categories if c]
        
        return Response({
            'entries': [{
                'id': e.id,
                'hs_code': e.hs_code,
                'description': e.description,
                'description_en': e.description_en,
                'category': e.category,
                'chapter': e.chapter,
                'ad_valorem_rate': float(e.ad_valorem_rate) if e.ad_valorem_rate else 0,
                'ice_rate': float(e.ice_rate) if e.ice_rate else 0,
                'unit': e.unit,
                'requires_permit': e.requires_permit,
                'permit_institution': e.permit_institution,
                'permit_name': e.permit_name,
                'permit_processing_days': e.permit_processing_days,
                'keywords': e.keywords,
                'notes': e.notes,
                'is_active': e.is_active,
                'created_at': e.created_at.isoformat() if e.created_at else None,
                'updated_at': e.updated_at.isoformat() if e.updated_at else None,
            } for e in entries],
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': max(1, (total + per_page - 1) // per_page) if per_page > 0 else 1,
            'categories': list(categories),
        })
    
    def post(self, request):
        from SalesModule.models import HSCodeEntry
        
        data = request.data
        
        if HSCodeEntry.objects.filter(hs_code=data.get('hs_code')).exists():
            return Response({
                'error': f'Ya existe una entrada con el código HS {data.get("hs_code")}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        entry = HSCodeEntry.objects.create(
            hs_code=data.get('hs_code', ''),
            description=data.get('description', ''),
            description_en=data.get('description_en', ''),
            category=data.get('category', ''),
            chapter=data.get('chapter', ''),
            ad_valorem_rate=data.get('ad_valorem_rate', 0),
            ice_rate=data.get('ice_rate', 0),
            unit=data.get('unit', 'kg'),
            requires_permit=data.get('requires_permit', False),
            permit_institution=data.get('permit_institution', ''),
            permit_name=data.get('permit_name', ''),
            permit_processing_days=data.get('permit_processing_days', ''),
            keywords=data.get('keywords', ''),
            notes=data.get('notes', ''),
            is_active=data.get('is_active', True),
        )
        
        return Response({
            'success': True,
            'message': 'Partida arancelaria creada',
            'id': entry.id
        }, status=status.HTTP_201_CREATED)
    
    def put(self, request):
        from SalesModule.models import HSCodeEntry
        
        entry_id = request.data.get('id')
        if not entry_id:
            return Response({'error': 'ID requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            entry = HSCodeEntry.objects.get(id=entry_id)
        except HSCodeEntry.DoesNotExist:
            return Response({'error': 'Entrada no encontrada'}, status=status.HTTP_404_NOT_FOUND)
        
        fields = ['hs_code', 'description', 'description_en', 'category', 'chapter',
                  'ad_valorem_rate', 'ice_rate', 'unit', 'requires_permit', 
                  'permit_institution', 'permit_name', 'permit_processing_days',
                  'keywords', 'notes', 'is_active']
        
        for field in fields:
            if field in request.data:
                setattr(entry, field, request.data[field])
        
        entry.save()
        
        return Response({
            'success': True,
            'message': 'Partida arancelaria actualizada'
        })
    
    def delete(self, request):
        from SalesModule.models import HSCodeEntry
        
        entry_id = request.query_params.get('id')
        if not entry_id:
            return Response({'error': 'ID requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            entry = HSCodeEntry.objects.get(id=entry_id)
            entry.delete()
            return Response({
                'success': True,
                'message': 'Partida arancelaria eliminada'
            })
        except HSCodeEntry.DoesNotExist:
            return Response({'error': 'Entrada no encontrada'}, status=status.HTTP_404_NOT_FOUND)


class HSCodeImportView(APIView):
    """
    POST: Import HS codes from CSV/Excel
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def post(self, request):
        from SalesModule.models import HSCodeEntry
        from decimal import Decimal
        
        if 'file' not in request.FILES:
            return Response({'error': 'No se proporcionó archivo'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        created_count = 0
        updated_count = 0
        errors = []
        
        try:
            if file_extension == 'csv':
                content = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif file_extension in ['xlsx', 'xls']:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response({'error': 'Formato no soportado. Use CSV o Excel.'}, status=status.HTTP_400_BAD_REQUEST)
            
            for i, row in enumerate(rows, start=2):
                try:
                    hs_code = str(row.get('hs_code', row.get('codigo_hs', ''))).strip()
                    if not hs_code:
                        continue
                    
                    description = str(row.get('description', row.get('descripcion', ''))).strip()
                    ad_valorem = row.get('ad_valorem_rate', row.get('ad_valorem', 0))
                    if ad_valorem is None:
                        ad_valorem = 0
                    
                    entry, created = HSCodeEntry.objects.update_or_create(
                        hs_code=hs_code,
                        defaults={
                            'description': description,
                            'description_en': str(row.get('description_en', row.get('descripcion_ingles', ''))).strip(),
                            'category': str(row.get('category', row.get('categoria', ''))).strip(),
                            'chapter': str(row.get('chapter', row.get('capitulo', ''))).strip(),
                            'ad_valorem_rate': Decimal(str(ad_valorem)),
                            'ice_rate': Decimal(str(row.get('ice_rate', row.get('ice', 0)) or 0)),
                            'unit': str(row.get('unit', row.get('unidad', 'kg'))).strip() or 'kg',
                            'requires_permit': str(row.get('requires_permit', row.get('requiere_permiso', 'false'))).lower() in ['true', '1', 'si', 'yes'],
                            'permit_institution': str(row.get('permit_institution', row.get('institucion', ''))).strip(),
                            'permit_name': str(row.get('permit_name', row.get('permiso', ''))).strip(),
                            'keywords': str(row.get('keywords', row.get('palabras_clave', ''))).strip(),
                            'is_active': True,
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                        
                except Exception as e:
                    errors.append(f"Fila {i}: {str(e)}")
            
            return Response({
                'success': True,
                'message': f'Importación completada: {created_count} creados, {updated_count} actualizados',
                'created': created_count,
                'updated': updated_count,
                'errors': errors[:10] if errors else []
            })
            
        except Exception as e:
            return Response({
                'error': f'Error procesando archivo: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


class HSCodeExportView(APIView):
    """
    GET: Export HS codes to CSV or Excel
    Query params:
      - format: 'csv' or 'excel' (default: csv)
      - search: search term
      - category: filter by category
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        from SalesModule.models import HSCodeEntry
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        export_format = request.query_params.get('format', 'csv').lower()
        search = request.query_params.get('search', '')
        category = request.query_params.get('category', '')
        
        queryset = HSCodeEntry.objects.filter(is_active=True)
        
        if search:
            queryset = queryset.filter(
                Q(hs_code__icontains=search) |
                Q(description__icontains=search) |
                Q(description_en__icontains=search) |
                Q(keywords__icontains=search) |
                Q(category__icontains=search)
            )
        
        if category:
            queryset = queryset.filter(category__icontains=category)
        
        queryset = queryset.order_by('hs_code')
        
        headers = [
            'hs_code', 'description', 'description_en', 'category', 'chapter',
            'ad_valorem_rate', 'ice_rate', 'unit', 'requires_permit',
            'permit_institution', 'permit_name', 'permit_processing_days',
            'keywords', 'notes', 'is_active'
        ]
        
        rows = []
        for entry in queryset:
            rows.append([
                entry.hs_code,
                entry.description,
                entry.description_en or '',
                entry.category or '',
                entry.chapter or '',
                float(entry.ad_valorem_rate) if entry.ad_valorem_rate else 0,
                float(entry.ice_rate) if entry.ice_rate else 0,
                entry.unit or 'kg',
                'true' if entry.requires_permit else 'false',
                entry.permit_institution or '',
                entry.permit_name or '',
                entry.permit_processing_days or '',
                entry.keywords or '',
                entry.notes or '',
                'true' if entry.is_active else 'false',
            ])
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if export_format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'HS Codes'
            
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="hs_codes_{timestamp}.xlsx"'
            return response
        
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(rows)
            
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="hs_codes_{timestamp}.csv"'
            return response


class TrackingTemplatesView(APIView):
    """
    CRUD for Tracking Templates - milestone configurations by transport type.
    GET: List all templates grouped by transport_type
    POST: Create new milestone
    PUT: Update milestone (requires id in body)
    DELETE: Delete milestone by id in query params
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        templates = TrackingTemplate.objects.all().order_by('transport_type', 'milestone_order')
        
        grouped = {
            'FCL': [],
            'LCL': [],
            'AIR': []
        }
        
        for t in templates:
            grouped[t.transport_type].append({
                'id': t.id,
                'transport_type': t.transport_type,
                'milestone_name': t.milestone_name,
                'milestone_order': t.milestone_order,
                'is_active': t.is_active,
                'description': t.description,
                'created_at': t.created_at.isoformat() if t.created_at else None,
                'updated_at': t.updated_at.isoformat() if t.updated_at else None,
            })
        
        return Response({
            'success': True,
            'templates': grouped,
            'counts': {
                'FCL': len(grouped['FCL']),
                'LCL': len(grouped['LCL']),
                'AIR': len(grouped['AIR']),
                'total': sum(len(v) for v in grouped.values())
            }
        })
    
    def post(self, request):
        transport_type = request.data.get('transport_type')
        milestone_name = request.data.get('milestone_name')
        milestone_order = request.data.get('milestone_order')
        description = request.data.get('description', '')
        is_active = request.data.get('is_active', True)
        
        if not transport_type or not milestone_name:
            return Response({
                'error': 'transport_type y milestone_name son requeridos'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if transport_type not in ['FCL', 'LCL', 'AIR']:
            return Response({
                'error': 'transport_type debe ser FCL, LCL o AIR'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if milestone_order is None:
            max_order = TrackingTemplate.objects.filter(
                transport_type=transport_type
            ).aggregate(max_order=Count('id'))['max_order'] or 0
            milestone_order = max_order + 1
        
        if TrackingTemplate.objects.filter(
            transport_type=transport_type,
            milestone_order=milestone_order
        ).exists():
            TrackingTemplate.objects.filter(
                transport_type=transport_type,
                milestone_order__gte=milestone_order
            ).update(milestone_order=F('milestone_order') + 1)
        
        template = TrackingTemplate.objects.create(
            transport_type=transport_type,
            milestone_name=milestone_name,
            milestone_order=milestone_order,
            description=description,
            is_active=is_active
        )
        
        return Response({
            'success': True,
            'message': 'Hito creado exitosamente',
            'template': {
                'id': template.id,
                'transport_type': template.transport_type,
                'milestone_name': template.milestone_name,
                'milestone_order': template.milestone_order,
                'is_active': template.is_active,
                'description': template.description,
            }
        }, status=status.HTTP_201_CREATED)
    
    def put(self, request):
        template_id = request.data.get('id')
        if not template_id:
            return Response({'error': 'ID requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            template = TrackingTemplate.objects.get(id=template_id)
        except TrackingTemplate.DoesNotExist:
            return Response({'error': 'Plantilla no encontrada'}, status=status.HTTP_404_NOT_FOUND)
        
        if 'milestone_name' in request.data:
            template.milestone_name = request.data['milestone_name']
        if 'description' in request.data:
            template.description = request.data['description']
        if 'is_active' in request.data:
            template.is_active = request.data['is_active']
        if 'milestone_order' in request.data:
            new_order = request.data['milestone_order']
            old_order = template.milestone_order
            
            if new_order != old_order:
                if new_order < old_order:
                    TrackingTemplate.objects.filter(
                        transport_type=template.transport_type,
                        milestone_order__gte=new_order,
                        milestone_order__lt=old_order
                    ).exclude(id=template_id).update(milestone_order=F('milestone_order') + 1)
                else:
                    TrackingTemplate.objects.filter(
                        transport_type=template.transport_type,
                        milestone_order__gt=old_order,
                        milestone_order__lte=new_order
                    ).exclude(id=template_id).update(milestone_order=F('milestone_order') - 1)
                
                template.milestone_order = new_order
        
        template.save()
        
        return Response({
            'success': True,
            'message': 'Plantilla actualizada exitosamente',
            'template': {
                'id': template.id,
                'transport_type': template.transport_type,
                'milestone_name': template.milestone_name,
                'milestone_order': template.milestone_order,
                'is_active': template.is_active,
                'description': template.description,
            }
        })
    
    def delete(self, request):
        template_id = request.query_params.get('id')
        if not template_id:
            return Response({'error': 'ID requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            template = TrackingTemplate.objects.get(id=template_id)
            transport_type = template.transport_type
            deleted_order = template.milestone_order
            template.delete()
            
            TrackingTemplate.objects.filter(
                transport_type=transport_type,
                milestone_order__gt=deleted_order
            ).update(milestone_order=F('milestone_order') - 1)
            
            return Response({
                'success': True,
                'message': 'Plantilla eliminada exitosamente'
            })
        except TrackingTemplate.DoesNotExist:
            return Response({'error': 'Plantilla no encontrada'}, status=status.HTTP_404_NOT_FOUND)


class RUCApprovalHistoryView(APIView):
    """
    View to retrieve RUC approval/rejection history for Master Admin.
    """
    authentication_classes = [MasterAdminAuthentication]
    permission_classes = [IsMasterAdmin]
    
    def get(self, request):
        from accounts.models import RUCApprovalHistory
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        ruc_filter = request.query_params.get('ruc', None)
        action_filter = request.query_params.get('action', None)
        
        queryset = RUCApprovalHistory.objects.all().order_by('-performed_at')
        
        if ruc_filter:
            queryset = queryset.filter(ruc_number__icontains=ruc_filter)
        
        if action_filter and action_filter in ['approved', 'rejected']:
            queryset = queryset.filter(action=action_filter)
        
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        
        history_items = queryset[start:end]
        
        history_data = []
        for item in history_items:
            history_data.append({
                'id': item.id,
                'ruc_number': item.ruc_number,
                'company_name': item.company_name,
                'user_email': item.user_email,
                'user_name': item.user_name,
                'action': item.action,
                'action_display': item.get_action_display(),
                'admin_notes': item.admin_notes,
                'performed_at': item.performed_at.isoformat(),
            })
        
        return Response({
            'success': True,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'history': history_data
        })
