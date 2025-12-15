from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Q
from datetime import timedelta, datetime
import csv, io, secrets
import logging
import traceback

from accounts.mixins import OwnerFilterMixin
from .permissions import IsLeadUser

logger = logging.getLogger(__name__)
from .models import (
    Lead, Opportunity, Quote, TaskReminder, Meeting, APIKey, BulkLeadImport,
    QuoteSubmission, QuoteSubmissionDocument, CostRate, LeadCotizacion, QuoteScenario, QuoteLineItem,
    FreightRate, InsuranceRate, CustomsDutyRate, InlandTransportQuoteRate, CustomsBrokerageRate,
    Shipment, ShipmentTracking, PreLiquidation, PreLiquidationDocument, Port,
    ShippingInstruction, ShippingInstructionDocument, ShipmentMilestone
)
from rest_framework.parsers import MultiPartParser, FormParser
from .serializers import (
    LeadSerializer, OpportunitySerializer, QuoteSerializer, 
    QuoteGenerateSerializer, TaskReminderSerializer, MeetingSerializer,
    APIKeySerializer, BulkLeadImportSerializer, QuoteSubmissionSerializer, QuoteSubmissionDocumentSerializer,
    QuoteSubmissionDetailSerializer, CostRateSerializer, LeadCotizacionSerializer,
    LeadCotizacionInstruccionSerializer, LeadCotizacionDetailSerializer,
    QuoteScenarioSerializer, QuoteLineItemSerializer,
    QuoteScenarioSelectSerializer, GenerateScenariosSerializer,
    FreightRateSerializer, InsuranceRateSerializer, CustomsDutyRateSerializer,
    InlandTransportQuoteRateSerializer, CustomsBrokerageRateSerializer,
    CustomsDutyCalculationSerializer, BrokerageFeeCalculationSerializer,
    ShipmentSerializer, ShipmentDetailSerializer, ShipmentCreateSerializer,
    ShipmentTrackingSerializer, AddTrackingEventSerializer,
    PreLiquidationSerializer, PreLiquidationDocumentSerializer, HSCodeSuggestionSerializer,
    PreLiquidationUpdateHSCodeSerializer, RequestAssistanceSerializer,
    PortSerializer, PortSearchSerializer,
    ShippingInstructionSerializer, ShippingInstructionDetailSerializer,
    ShippingInstructionCreateSerializer, ShippingInstructionFormSerializer,
    ShippingInstructionDocumentSerializer,
    ShipmentMilestoneSerializer, CargoTrackingListSerializer, CargoTrackingDetailSerializer
)


class LeadViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = Lead.objects.all()
    serializer_class = LeadSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'source', 'country']
    search_fields = ['company_name', 'contact_name', 'email']
    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        
        if response.status_code == status.HTTP_201_CREATED:
            lead_data = response.data
            is_active_importer = lead_data.get('is_active_importer', False)
            
            if not is_active_importer:
                self._send_customs_email(lead_data)
        
        return response
    
    def _send_customs_email(self, lead_data):
        try:
            subject = f"Nuevo Lead: {lead_data['company_name']} - Oferta de Servicio de RUC"
            message = f"""
Estimado Departamento de Aduanas,

Se ha registrado un nuevo lead que podría ser cliente potencial para nuestro servicio de registro de RUC ante la SENAE.

Información del Lead:
- Empresa: {lead_data['company_name']}
- Contacto: {lead_data['contact_name']}
- Correo: {lead_data['email']}
- Teléfono: {lead_data.get('phone', 'N/A')}
- WhatsApp: {lead_data.get('whatsapp', 'N/A')}
- Ciudad: {lead_data.get('city', 'N/A')}
- Notas: {lead_data.get('notes', 'N/A')}

Por favor, contacte al lead para ofrecerle nuestro servicio de registro de RUC ante la SENAE - Aduana del Ecuador.

Sistema IntegralCargoSolutions ICS
"""
            
            customs_email = getattr(settings, 'CUSTOMS_DEPARTMENT_EMAIL', 'aduanas@integralcargosolutions.ec')
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [customs_email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending customs email: {str(e)}")


class OpportunityViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = Opportunity.objects.all()
    serializer_class = OpportunitySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['stage', 'lead']
    search_fields = ['opportunity_name']


class QuoteViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = Quote.objects.all()
    serializer_class = QuoteSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'incoterm', 'cargo_type']
    search_fields = ['quote_number']
    
    @action(detail=False, methods=['post'], url_path='generate')
    def generate_quote(self, request):
        serializer = QuoteGenerateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        try:
            opportunity = Opportunity.objects.get(id=data['opportunity_id'])
        except Opportunity.DoesNotExist:
            return Response({'error': 'Oportunidad no encontrada'}, status=status.HTTP_404_NOT_FOUND)
        
        final_price = data['base_rate'] + data['profit_margin']
        
        quote = Quote.objects.create(
            opportunity=opportunity,
            origin=data['origin'],
            destination=data['destination'],
            incoterm=data['incoterm'],
            cargo_type=data['cargo_type'],
            cargo_description=data.get('cargo_description', ''),
            base_rate=data['base_rate'],
            profit_margin=data['profit_margin'],
            final_price=final_price,
            valid_until=data.get('valid_until'),
            notes=data.get('notes', ''),
            status='borrador'
        )
        
        return Response(QuoteSerializer(quote).data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'], url_path='send')
    def send_quote(self, request, pk=None):
        quote = self.get_object()
        
        if quote.status == 'enviado':
            return Response({'error': 'La cotización ya fue enviada'}, status=status.HTTP_400_BAD_REQUEST)
        
        quote.status = 'enviado'
        quote.sent_at = timezone.now()
        quote.save()
        
        follow_up_time = timezone.now() + timedelta(hours=1)
        TaskReminder.objects.create(
            quote=quote,
            lead=quote.opportunity.lead,
            task_type='seguimiento_cotizacion',
            title=f'Seguimiento de Cotización {quote.quote_number}',
            description=f'Hacer seguimiento de la cotización {quote.quote_number} enviada a {quote.opportunity.lead.company_name}',
            priority='alta',
            due_date=follow_up_time
        )
        
        return Response({
            'message': 'Cotización enviada exitosamente',
            'quote': QuoteSerializer(quote).data,
            'follow_up_created': True,
            'follow_up_time': follow_up_time
        }, status=status.HTTP_200_OK)


class TaskReminderViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = TaskReminder.objects.all()
    serializer_class = TaskReminderSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'priority', 'task_type', 'lead']
    search_fields = ['title', 'description']
    
    @action(detail=True, methods=['post'], url_path='complete')
    def complete_task(self, request, pk=None):
        task = self.get_object()
        task.status = 'completada'
        task.completed_at = timezone.now()
        task.save()
        return Response(TaskReminderSerializer(task).data)


class MeetingViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'meeting_type', 'lead']
    search_fields = ['title']
    
    @action(detail=True, methods=['post'], url_path='sync-calendar')
    def sync_calendar(self, request, pk=None):
        meeting = self.get_object()
        
        calendar_type = request.data.get('calendar_type')
        
        if calendar_type == 'google':
            meeting.google_calendar_synced = True
            meeting.save()
            return Response({
                'message': 'Mock: Reunión sincronizada con Google Calendar',
                'meeting_id': meeting.id,
                'reminder_15min': True
            })
        elif calendar_type == 'outlook':
            meeting.outlook_synced = True
            meeting.save()
            return Response({
                'message': 'Mock: Reunión sincronizada con Outlook',
                'meeting_id': meeting.id,
                'reminder_15min': True
            })
        else:
            return Response({'error': 'Tipo de calendario no válido (google/outlook)'}, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.decorators import api_view
from rest_framework.views import APIView
from django.http import HttpResponse
from datetime import datetime
from .reports.generators import (
    generate_sales_metrics_report,
    generate_lead_conversion_report,
    generate_communication_report,
    generate_quote_analytics_report,
    generate_quote_submissions_report,
    generate_dashboard_summary,
    generate_cost_analytics_report,
    generate_operational_kpis,
    generate_import_trends_report,
    generate_financial_summary,
    export_report_to_excel,
    export_report_to_pdf
)


class DashboardAPIView(APIView):
    """Consolidated dashboard summary for LEAD portal"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        summary = generate_dashboard_summary(user=request.user)
        return Response(summary, status=status.HTTP_200_OK)


class ReportsAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        report_type = request.query_params.get('type', 'sales_metrics')
        format_type = request.query_params.get('format', 'json')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        start_date = None
        end_date = None
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        user = request.user
        
        if report_type == 'sales_metrics':
            report_data = generate_sales_metrics_report(start_date, end_date)
        elif report_type == 'lead_conversion':
            report_data = generate_lead_conversion_report(start_date, end_date)
        elif report_type == 'communication':
            report_data = generate_communication_report(start_date, end_date)
        elif report_type == 'quote_analytics':
            report_data = generate_quote_analytics_report(start_date, end_date)
        elif report_type == 'quote_submissions':
            report_data = generate_quote_submissions_report(start_date, end_date)
        elif report_type == 'cost_analytics':
            report_data = generate_cost_analytics_report(start_date, end_date, user=user)
        elif report_type == 'operational_kpis':
            report_data = generate_operational_kpis(start_date, end_date, user=user)
        elif report_type == 'import_trends':
            report_data = generate_import_trends_report(start_date, end_date, user=user)
        elif report_type == 'financial_summary':
            report_data = generate_financial_summary(start_date, end_date, user=user)
        else:
            return Response({'error': 'Tipo de reporte no válido'}, status=status.HTTP_400_BAD_REQUEST)
        
        if format_type == 'json':
            return Response(report_data, status=status.HTTP_200_OK)
        elif format_type == 'excel':
            excel_file = export_report_to_excel(report_data, report_type)
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            return response
        elif format_type == 'pdf':
            pdf_file = export_report_to_pdf(report_data, report_type)
            response = HttpResponse(pdf_file.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.now().strftime("%Y%m%d")}.pdf"'
            return response
        else:
            return Response({'error': 'Formato no válido'}, status=status.HTTP_400_BAD_REQUEST)


class APIKeyViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = APIKey.objects.all()
    serializer_class = APIKeySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_active', 'service_type']
    search_fields = ['name', 'service_type']
    
    def perform_create(self, serializer):
        key = 'ic_' + secrets.token_urlsafe(32)
        serializer.save(key=key, owner=self.request.user)


class QuoteSubmissionDocumentViewSet(viewsets.ModelViewSet):
    queryset = QuoteSubmissionDocument.objects.all()
    serializer_class = QuoteSubmissionDocumentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        submission_id = self.request.query_params.get('quote_submission')
        if submission_id:
            queryset = queryset.filter(quote_submission_id=submission_id)
        return queryset


class QuoteSubmissionViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = QuoteSubmission.objects.all()
    serializer_class = QuoteSubmissionSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'transport_type', 'city']
    search_fields = ['company_name', 'contact_email', 'origin', 'destination']
    
    def _is_test_mode(self, quote_submission, user=None):
        """Detecta si la solicitud es de un usuario de prueba basado en credenciales de login"""
        test_domains = ['@test.importaya.ia', '@test.importaya.ec', '@demo.importaya.ia']
        
        user_email = ''
        if user and hasattr(user, 'email'):
            user_email = user.email or ''
        elif quote_submission.owner and hasattr(quote_submission.owner, 'email'):
            user_email = quote_submission.owner.email or ''
        
        contact_email = quote_submission.contact_email or ''
        
        is_user_test = any(domain in user_email.lower() for domain in test_domains)
        is_contact_test = any(domain in contact_email.lower() for domain in test_domains)
        
        return is_user_test or is_contact_test
    
    def _is_complete_submission(self, quote_submission):
        """Verifica si la solicitud tiene la información mínima completa para generar cotización"""
        required_fields = [
            quote_submission.origin,
            quote_submission.destination,
            quote_submission.transport_type,
            quote_submission.cargo_weight_kg or quote_submission.cargo_volume_cbm,
        ]
        return all(required_fields)
    
    def perform_create(self, serializer):
        quote_submission = serializer.save(owner=self.request.user)
        
        is_test = self._is_test_mode(quote_submission, user=self.request.user)
        
        if is_test:
            logger.info(f"[MODO PRUEBA] Solicitud {quote_submission.submission_number} - Usuario de test detectado")
            quote_submission.is_oce_registered = True
            quote_submission.is_first_quote = False
            quote_submission.vendor_validation_status = 'test_mode'
            quote_submission.notes = (quote_submission.notes or '') + '\n[MODO PRUEBA] OCE asumido como registrado para testing.'
            quote_submission.save(update_fields=['is_oce_registered', 'is_first_quote', 'vendor_validation_status', 'notes'])
        
        if not is_test:
            if not quote_submission.is_oce_registered and not quote_submission.customs_alert_sent:
                self._send_customs_alert_email(quote_submission)
                quote_submission.customs_alert_sent = True
                quote_submission.save(update_fields=['customs_alert_sent'])
            
            self._check_first_quote_and_send_vendor_validation(quote_submission)
        
        if self._is_complete_submission(quote_submission):
            self._generate_intelligent_quote_with_ai(quote_submission)
        else:
            logger.warning(f"Solicitud {quote_submission.id} incompleta - No se genera cotización AI")
            quote_submission.status = 'validacion_pendiente'
            quote_submission.notes = (quote_submission.notes or '') + '\n[INFO] Solicitud requiere información adicional para generar cotización.'
            quote_submission.save(update_fields=['status', 'notes'])
        
        if quote_submission.status == 'validacion_pendiente':
            self._find_and_apply_cost_rates(quote_submission)
    
    def _check_first_quote_and_send_vendor_validation(self, quote_submission):
        """BLOQUE 2: Verifica si es primera cotización del RUC y envía email al Freight Forwarder"""
        try:
            if not quote_submission.company_ruc:
                return
            
            existing_quotes = QuoteSubmission.objects.filter(
                company_ruc=quote_submission.company_ruc
            ).exclude(id=quote_submission.id).exists()
            
            if existing_quotes:
                quote_submission.is_first_quote = False
                quote_submission.vendor_validation_status = 'not_required'
                quote_submission.save(update_fields=['is_first_quote', 'vendor_validation_status'])
            else:
                quote_submission.is_first_quote = True
                quote_submission.vendor_validation_status = 'pending'
                self._send_vendor_validation_email(quote_submission)
                quote_submission.save(update_fields=['is_first_quote', 'vendor_validation_status'])
                
        except Exception as e:
            logger.error(f"Error checking first quote: {e}")
    
    def _send_vendor_validation_email(self, quote_submission):
        """Envía email al Freight Forwarder para validar si el RUC es cliente existente"""
        try:
            subject = f"Validación de Cliente: RUC {quote_submission.company_ruc}"
            message = f"""
Estimado Freight Forwarder,

Se ha recibido una nueva solicitud de cotización de un cliente NUEVO en el sistema ImportaYa.ia.
Por favor confirme si este RUC corresponde a un cliente existente en su base de datos.

DATOS DEL CLIENTE:
- Empresa: {quote_submission.company_name}
- RUC: {quote_submission.company_ruc}
- Contacto: {quote_submission.contact_name}
- Email: {quote_submission.contact_email}
- Teléfono: {quote_submission.contact_phone}
- Ciudad: {quote_submission.city}

SOLICITUD DE COTIZACIÓN #{quote_submission.submission_number}:
- Origen: {quote_submission.origin}
- Destino: {quote_submission.destination}
- Tipo Transporte: {quote_submission.transport_type}
- Descripción: {quote_submission.cargo_description or quote_submission.product_description or 'No especificada'}
- Peso: {quote_submission.cargo_weight_kg} kg
- Volumen: {quote_submission.cargo_volume_cbm} CBM

ACCIÓN REQUERIDA:
Por favor acceda al panel de administración para confirmar:
- ¿Es cliente existente? → Se aplicará margen reducido ($500 mínimo)
- ¿Es cliente nuevo? → Se aplicará margen estándar

Sistema ImportaYa.ia
La logística de carga integral, ahora es Inteligente!
"""
            forwarder_email = getattr(settings, 'FREIGHT_FORWARDER_EMAIL', 'forwarder@importaya.ia')
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [forwarder_email],
                fail_silently=True,
            )
            logger.info(f"Vendor validation email sent for submission {quote_submission.id} - RUC: {quote_submission.company_ruc}")
        except Exception as e:
            logger.error(f"Error sending vendor validation email: {e}")
    
    def _send_customs_alert_email(self, quote_submission):
        """Envía alerta al Ejecutivo de Aduanas para leads sin OCE"""
        try:
            subject = f"Nuevo Lead Sin OCE: {quote_submission.company_name} - Registro de RUC/OCE"
            message = f"""
Estimado Ejecutivo de Aduanas,

Se ha registrado un nuevo lead que NO es Operador de Comercio Exterior (OCE) registrado ante SENAE.
Este cliente necesita asistencia para registrar su RUC y OCE antes de operar.

INFORMACIÓN DEL LEAD:
- Empresa: {quote_submission.company_name}
- Contacto: {quote_submission.contact_name}
- Email: {quote_submission.contact_email}
- Teléfono: {quote_submission.contact_phone}
- WhatsApp: {quote_submission.contact_whatsapp or quote_submission.contact_phone}
- RUC: {quote_submission.company_ruc or 'No proporcionado'}
- Ciudad: {quote_submission.city}

SOLICITUD DE COTIZACIÓN:
- Origen: {quote_submission.origin}
- Destino: {quote_submission.destination}
- Tipo Transporte: {quote_submission.transport_type}
- Descripción: {quote_submission.cargo_description or quote_submission.product_description or 'No especificada'}

Por favor, contacte al lead para ofrecerle nuestro servicio de registro de RUC/OCE ante la SENAE.

Sistema ImportaYa.ia
"""
            customs_email = getattr(settings, 'CUSTOMS_DEPARTMENT_EMAIL', 'aduanas@importaya.ia')
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [customs_email],
                fail_silently=True,
            )
            logger.info(f"Customs alert email sent for submission {quote_submission.id}")
        except Exception as e:
            logger.error(f"Error sending customs alert email: {e}")
    
    def _generate_intelligent_quote_with_ai(self, quote_submission):
        """Genera cotización inteligente usando Gemini AI"""
        try:
            from .gemini_service import generate_intelligent_quote
            import json
            
            product_desc = quote_submission.product_description or quote_submission.cargo_description or ''
            if quote_submission.product_origin_country:
                product_desc = f"{product_desc} (Origen: {quote_submission.product_origin_country})"
            
            fob_value = None
            if quote_submission.fob_value_usd:
                fob_value = float(quote_submission.fob_value_usd)
            
            ai_result = generate_intelligent_quote(
                cargo_description=product_desc,
                origin=quote_submission.origin,
                destination=quote_submission.destination,
                transport_type=quote_submission.transport_type,
                weight_kg=float(quote_submission.cargo_weight_kg) if quote_submission.cargo_weight_kg else None,
                volume_cbm=float(quote_submission.cargo_volume_cbm) if quote_submission.cargo_volume_cbm else None,
                incoterm=quote_submission.incoterm or "FOB",
                fob_value_usd=fob_value,
                container_type=quote_submission.container_type or None,
                hs_code_known=quote_submission.hs_code_known or None
            )
            
            clasificacion = ai_result.get('clasificacion', {})
            quote_submission.ai_hs_code = clasificacion.get('hs_code', '')
            quote_submission.ai_hs_confidence = clasificacion.get('confianza')
            quote_submission.ai_category = clasificacion.get('categoria', '')
            
            tributos = ai_result.get('tributos', {})
            quote_submission.ai_ad_valorem_pct = tributos.get('ad_valorem_pct')
            
            permisos = ai_result.get('permisos', [])
            quote_submission.ai_requires_permit = len(permisos) > 0
            quote_submission.ai_permit_institutions = json.dumps(permisos, ensure_ascii=False) if permisos else ''
            
            quote_submission.ai_response = json.dumps(ai_result, ensure_ascii=False)
            quote_submission.ai_status = ai_result.get('ai_status', 'unknown')
            
            if ai_result.get('ai_status') == 'success':
                self._create_scenarios_from_ai(quote_submission, ai_result)
                quote_submission.status = 'cotizacion_generada'
            
            quote_submission.save()
            logger.info(f"AI quote generated for submission {quote_submission.id}: {quote_submission.ai_status}")
            
        except Exception as e:
            logger.error(f"Error generating AI quote: {e}")
            quote_submission.ai_status = 'error'
            quote_submission.notes = f"Error IA: {str(e)}"
            quote_submission.save()
    
    def _create_scenarios_from_ai(self, quote_submission, ai_result):
        """Procesa escenarios de cotización desde respuesta de IA
        
        Nota: Los escenarios se almacenan en el campo ai_response como JSON.
        QuoteScenario está vinculado a LeadCotizacion, no a QuoteSubmission.
        Los escenarios generados por IA pueden usarse para crear una LeadCotizacion posteriormente.
        """
        from decimal import Decimal
        try:
            escenarios = ai_result.get('escenarios', [])
            if escenarios:
                logger.info(f"AI generated {len(escenarios)} scenarios for submission {quote_submission.id}")
                
                scenario_summary = []
                min_price = None
                
                for esc in escenarios:
                    subtotal = esc.get('subtotal_logistica_usd', 0) or 0
                    scenario_summary.append(f"{esc.get('tipo', 'N/A')}: ${subtotal:.2f} ({esc.get('tiempo_transito_dias', 'N/A')} días)")
                    
                    if min_price is None or subtotal < min_price:
                        min_price = subtotal
                
                if min_price and min_price > 0:
                    quote_submission.final_price = Decimal(str(min_price))
                    logger.info(f"Set final_price to ${min_price:.2f} for submission {quote_submission.id}")
                
                if quote_submission.notes:
                    quote_submission.notes += f"\n\nEscenarios IA: {', '.join(scenario_summary)}"
                else:
                    quote_submission.notes = f"Escenarios IA: {', '.join(scenario_summary)}"
            
        except Exception as e:
            logger.error(f"Error processing scenarios from AI: {e}")
    
    def _find_and_apply_cost_rates(self, quote_submission):
        """Busca COST_RATES según origin/destination/transport_type"""
        try:
            cost_rate = CostRate.objects.filter(
                origin__icontains=quote_submission.origin,
                destination__icontains=quote_submission.destination,
                transport_type=quote_submission.transport_type,
                is_active=True
            ).order_by('-created_at').first()
            
            if cost_rate:
                quote_submission.cost_rate = cost_rate.rate
                quote_submission.cost_rate_source = cost_rate.source
                quote_submission.calculate_final_price()
                quote_submission.status = 'cotizacion_generada'
                quote_submission.processed_at = timezone.now()
            else:
                quote_submission.status = 'procesando_costos'
            
            quote_submission.save()
        except Exception as e:
            quote_submission.status = 'error_costos'
            quote_submission.notes = str(e)
            quote_submission.save()
    
    @action(detail=False, methods=['get'], url_path='search-cost-rates')
    def search_cost_rates(self, request):
        """Busca COST_RATES disponibles para cotización"""
        origin = request.query_params.get('origin')
        destination = request.query_params.get('destination')
        transport_type = request.query_params.get('transport_type')
        
        if not all([origin, destination, transport_type]):
            return Response({'error': 'origin, destination y transport_type son requeridos'}, status=status.HTTP_400_BAD_REQUEST)
        
        rates = CostRate.objects.filter(
            origin__icontains=origin,
            destination__icontains=destination,
            transport_type=transport_type,
            is_active=True
        ).order_by('-created_at')
        
        return Response(CostRateSerializer(rates, many=True).data)
    
    @action(detail=True, methods=['post'], url_path='validate-vendor')
    def validate_vendor(self, request, pk=None):
        """BLOQUE 2: Valida si el cliente es existente del forwarder y aplica margen de ganancia
        
        Body:
        - is_existing_client: bool - True si el RUC ya es cliente del forwarder
        - notes: str (opcional) - Notas adicionales del forwarder
        """
        quote_submission = self.get_object()
        
        if quote_submission.vendor_validation_status not in ['pending', 'not_required']:
            return Response({
                'error': 'Esta cotización ya fue validada',
                'current_status': quote_submission.vendor_validation_status
            }, status=status.HTTP_400_BAD_REQUEST)
        
        is_existing_client = request.data.get('is_existing_client', False)
        forwarder_notes = request.data.get('notes', '')
        
        quote_submission.vendor_validated_at = timezone.now()
        
        if is_existing_client:
            quote_submission.vendor_validation_status = 'existing_client'
            min_profit = 500.00
            if quote_submission.profit_markup and quote_submission.profit_markup < min_profit:
                quote_submission.profit_markup = min_profit
            elif not quote_submission.profit_markup:
                quote_submission.profit_markup = min_profit
            
            message = f'Cliente existente confirmado. Margen mínimo aplicado: $500 USD.'
        else:
            quote_submission.vendor_validation_status = 'new_client'
            if not quote_submission.profit_markup:
                quote_submission.profit_markup = 100.00
            message = 'Cliente nuevo confirmado. Margen estándar aplicado.'
        
        if forwarder_notes:
            if quote_submission.notes:
                quote_submission.notes += f"\n\nNota del Forwarder: {forwarder_notes}"
            else:
                quote_submission.notes = f"Nota del Forwarder: {forwarder_notes}"
        
        quote_submission.calculate_final_price()
        quote_submission.save()
        
        return Response({
            'message': message,
            'vendor_validation_status': quote_submission.vendor_validation_status,
            'profit_markup': float(quote_submission.profit_markup) if quote_submission.profit_markup else None,
            'final_price': float(quote_submission.final_price) if quote_submission.final_price else None,
            'validated_at': quote_submission.vendor_validated_at.isoformat()
        })
    
    @action(detail=False, methods=['get'], url_path='pending-vendor-validations')
    def pending_vendor_validations(self, request):
        """Lista solicitudes pendientes de validación de vendor (para panel admin)"""
        pending = QuoteSubmission.objects.filter(
            vendor_validation_status='pending',
            is_first_quote=True
        ).order_by('-created_at')
        
        return Response(QuoteSubmissionSerializer(pending, many=True).data)
    
    @action(detail=True, methods=['get'], url_path='download-pdf')
    def download_pdf(self, request, pk=None):
        """Descarga PDF de cotización con formato profesional"""
        try:
            quote_submission = self.get_object()
            
            from .reports.quote_pdf_generator import generate_quote_pdf, generate_multi_scenario_pdf
            import json
            
            scenario_index = request.query_params.get('scenario')
            
            scenario_data = None
            if quote_submission.ai_response:
                try:
                    ai_response = json.loads(quote_submission.ai_response)
                    escenarios = ai_response.get('escenarios', [])
                    
                    if scenario_index is not None:
                        idx = int(scenario_index)
                        if 0 <= idx < len(escenarios):
                            raw_scenario = escenarios[idx]
                        else:
                            raw_scenario = escenarios[0] if escenarios else {}
                    else:
                        raw_scenario = escenarios[0] if escenarios else {}
                    
                    from .quotation_engine import obtener_gastos_locales_db
                    
                    transport_type = quote_submission.transport_type
                    container_type = getattr(quote_submission, 'container_type', None) or '40HC'
                    container_type_normalized = container_type.replace('1x', '').upper()
                    quantity = quote_submission.quantity or 1
                    
                    carrier_code = raw_scenario.get('naviera_codigo') or raw_scenario.get('carrier_code') or raw_scenario.get('naviera', '')
                    if carrier_code:
                        carrier_code = carrier_code.upper().replace(' ', '_')
                        carrier_mappings = {
                            'MAERSK': 'MSK_GYE', 'HAPAG': 'HPG', 'HAPAG-LLOYD': 'HPG',
                            'YANG MING': 'YML', 'SEABOARD': 'SBM', 'SEABOARD MARINE': 'SBM',
                            'EVERGREEN': 'EMC', 'WAN HAI': 'WHL'
                        }
                        carrier_code = carrier_mappings.get(carrier_code, carrier_code)
                    
                    gastos_db = obtener_gastos_locales_db(
                        transport_type=transport_type,
                        port='GYE',
                        container_type=container_type_normalized,
                        quantity=quantity,
                        carrier_code=carrier_code if transport_type == 'FCL' else None
                    )
                    
                    costos_locales_db = {}
                    for item in gastos_db.get('items', []):
                        codigo = item.get('codigo', '').upper()
                        monto = item.get('monto', 0)
                        is_exempt = item.get('is_iva_exempt', False)
                        if 'THC' in codigo or 'DTHC' in codigo:
                            costos_locales_db['thc'] = monto
                            costos_locales_db['thc_is_exempt'] = is_exempt
                        elif 'HANDLING' in codigo:
                            costos_locales_db['handling'] = monto
                        elif 'LOCALES_MBL' in codigo or 'DOC' in codigo:
                            costos_locales_db['locales_mbl'] = monto
                        elif 'LOCALES_CNTR' in codigo:
                            costos_locales_db['locales_cntr'] = monto
                    
                    scenario_data = {
                        'flete_base': raw_scenario.get('flete_maritimo_usd') or raw_scenario.get('flete_aereo_usd') or raw_scenario.get('flete_usd', 1600),
                        'tarifa_cbm': raw_scenario.get('tarifa_cbm', 85),
                        'tarifa_ton': raw_scenario.get('tarifa_ton', 85),
                        'tarifa_kg': raw_scenario.get('tarifa_kg', 4.50),
                        'dias_transito': raw_scenario.get('tiempo_transito_dias', 'N/A'),
                        'dias_libres': raw_scenario.get('dias_libres_demora', 21),
                        'carrier_name': carrier_code,
                        'costos_locales': {
                            'locales_mbl': costos_locales_db.get('locales_mbl', 100),
                            'handling': costos_locales_db.get('handling', 50),
                            'locales_cntr': costos_locales_db.get('locales_cntr', 400),
                            'thc': costos_locales_db.get('thc', 200),
                        }
                    }
                except (json.JSONDecodeError, ValueError):
                    pass
            
            pdf_buffer = generate_quote_pdf(quote_submission, scenario_data)
            
            filename = f"Cotizacion_{quote_submission.submission_number or quote_submission.id}_{quote_submission.transport_type}.pdf"
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"Error generating quote PDF: {e}")
            return Response({'error': f'Error generando PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='multi-port-quote')
    def multi_port_quote(self, request):
        """
        Genera cotización multi-puerto (tabla tarifario sin totalizar).
        
        Cuando hay múltiples POL y/o POD, genera una tabla comparativa de tarifas
        de flete sin totalizar. Los gastos locales se muestran por separado.
        """
        from .quotation_engine import generar_cotizacion_multipuerto
        from decimal import Decimal
        import json
        
        origin_ports = request.data.get('origin_ports', [])
        destination_ports = request.data.get('destination_ports', [])
        transport_type = request.data.get('transport_type', 'FCL')
        container_type = request.data.get('container_type', '40HC')
        quantity = request.data.get('quantity', 1)
        weight_kg = request.data.get('weight_kg')
        volume_cbm = request.data.get('volume_cbm')
        
        # Validaciones
        if not origin_ports or not isinstance(origin_ports, list):
            return Response(
                {'error': 'origin_ports es requerido y debe ser una lista'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not destination_ports or not isinstance(destination_ports, list):
            return Response(
                {'error': 'destination_ports es requerido y debe ser una lista'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Convertir a Decimal si se proporcionan
        weight_decimal = Decimal(str(weight_kg)) if weight_kg else None
        volume_decimal = Decimal(str(volume_cbm)) if volume_cbm else None
        
        try:
            result = generar_cotizacion_multipuerto(
                origin_ports=origin_ports,
                destination_ports=destination_ports,
                transport_type=transport_type.upper(),
                container_type=container_type,
                quantity=quantity,
                weight_kg=weight_decimal,
                volume_cbm=volume_decimal
            )
            
            return Response(result, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error generating multi-port quote: {e}")
            return Response(
                {'error': f'Error generando cotización multi-puerto: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='validate-inland-address', permission_classes=[IsAuthenticated])
    def validate_inland_address(self, request, pk=None):
        """
        Valida dirección de transporte terrestre usando Gemini AI.
        
        1. Valida la dirección con IA para obtener coordenadas exactas
        2. Genera link de Google Maps
        3. Guarda la información en la base de datos
        4. Envía correo automático al freight forwarder
        
        POST data:
        - address: Dirección de entrega (opcional, usa la guardada si no se proporciona)
        - city: Ciudad de destino (opcional)
        
        Requiere autenticación.
        """
        from .address_validation_service import (
            validate_address_with_gemini,
            send_forwarder_notification_email,
            process_inland_transport_address
        )
        from django.utils import timezone
        
        try:
            quote_submission = self.get_object()
            
            address = request.data.get('address', quote_submission.inland_transport_address)
            city = request.data.get('city', quote_submission.inland_transport_city or quote_submission.city)
            
            if address and address != quote_submission.inland_transport_address:
                quote_submission.inland_transport_address = address
            if city and city != quote_submission.inland_transport_city:
                quote_submission.inland_transport_city = city
            
            quote_submission.needs_inland_transport = True
            quote_submission.save()
            
            result = process_inland_transport_address(quote_submission)
            
            if result.get('success'):
                return Response({
                    'success': True,
                    'message': 'Dirección validada y forwarder notificado exitosamente',
                    'validation': result.get('validation_result'),
                    'email': result.get('email_result'),
                    'quote_submission_id': quote_submission.id,
                    'google_maps_link': quote_submission.inland_transport_google_maps_link
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'success': False,
                    'error': result.get('error', 'Error al procesar dirección'),
                    'validation': result.get('validation_result'),
                    'email': result.get('email_result')
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            logger.error(f"Error validating inland address: {e}")
            return Response(
                {'error': f'Error al validar dirección: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='validate-address-preview', permission_classes=[IsAuthenticated])
    def validate_address_preview(self, request):
        """
        Valida una dirección sin guardar - solo para preview.
        Útil para mostrar al usuario la validación antes de enviar.
        
        POST data:
        - address: Dirección a validar (requerido)
        - city: Ciudad de destino (requerido)
        
        Requiere autenticación.
        """
        from .address_validation_service import validate_address_with_gemini
        
        address = request.data.get('address')
        city = request.data.get('city')
        
        if not address:
            return Response(
                {'error': 'La dirección es requerida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not city:
            return Response(
                {'error': 'La ciudad es requerida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            result = validate_address_with_gemini(
                address=address,
                city=city,
                country='Ecuador'
            )
            
            return Response({
                'success': result.get('success', False),
                'validated_address': result.get('validated_address'),
                'latitude': str(result.get('latitude')) if result.get('latitude') else None,
                'longitude': str(result.get('longitude')) if result.get('longitude') else None,
                'google_maps_link': result.get('google_maps_link'),
                'confidence': result.get('confidence'),
                'is_valid': result.get('is_valid', True),
                'notes': result.get('notes', ''),
                'street': result.get('street', ''),
                'neighborhood': result.get('neighborhood', ''),
                'province': result.get('province', '')
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error in address preview validation: {e}")
            return Response(
                {'error': f'Error al validar dirección: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='rate-table')
    def rate_table(self, request):
        """
        Obtiene tabla de tarifas para una ruta específica.
        Retorna múltiples carriers/navieras para comparación.
        """
        from .quotation_engine import obtener_tarifas_tabla
        
        pol = request.query_params.get('pol')
        pod = request.query_params.get('pod')
        transport_type = request.query_params.get('transport_type', 'FCL')
        container_type = request.query_params.get('container_type', '40HC')
        limit = int(request.query_params.get('limit', 10))
        
        if not pol or not pod:
            return Response(
                {'error': 'Parámetros requeridos: pol, pod'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            tarifas = obtener_tarifas_tabla(
                pol=pol,
                pod=pod,
                transport_type=transport_type.upper(),
                container_type=container_type,
                limit=limit
            )
            
            return Response({
                'pol': pol,
                'pod': pod,
                'transport_type': transport_type,
                'container_type': container_type,
                'total_tarifas': len(tarifas),
                'tarifas': tarifas
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error getting rate table: {e}")
            return Response(
                {'error': f'Error obteniendo tabla de tarifas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='download-multi-scenario-pdf')
    def download_multi_scenario_pdf(self, request, pk=None):
        """Descarga PDF con múltiples escenarios de cotización"""
        try:
            quote_submission = self.get_object()
            
            from .reports.quote_pdf_generator import generate_multi_scenario_pdf
            import json
            
            scenarios = []
            if quote_submission.ai_response:
                try:
                    ai_response = json.loads(quote_submission.ai_response)
                    escenarios = ai_response.get('escenarios', [])
                    
                    for esc in escenarios:
                        scenarios.append({
                            'nombre': esc.get('tipo', 'N/A'),
                            'dias_transito': esc.get('tiempo_transito_dias', 'N/A'),
                            'total_usd': esc.get('subtotal_logistica_usd', 0)
                        })
                except json.JSONDecodeError:
                    pass
            
            if not scenarios:
                return Response({'error': 'No hay escenarios disponibles para esta cotización'}, status=status.HTTP_400_BAD_REQUEST)
            
            pdf_buffer = generate_multi_scenario_pdf(quote_submission, scenarios)
            
            filename = f"Cotizacion_Escenarios_{quote_submission.submission_number or quote_submission.id}.pdf"
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"Error generating multi-scenario PDF: {e}")
            return Response({'error': f'Error generando PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='my-submissions', permission_classes=[IsAuthenticated])
    def my_submissions(self, request):
        """
        Obtiene las solicitudes de cotización del usuario autenticado (Lead).
        Filtra por el email del usuario o por el lead asociado.
        """
        try:
            user = request.user
            
            queryset = QuoteSubmission.objects.filter(
                Q(contact_email__iexact=user.email) |
                Q(lead__email__iexact=user.email) |
                Q(owner=user)
            ).order_by('-created_at')
            
            status_filter = request.query_params.get('status')
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            transport_filter = request.query_params.get('transport_type')
            if transport_filter:
                queryset = queryset.filter(transport_type=transport_filter)
            
            date_from = request.query_params.get('date_from')
            if date_from:
                queryset = queryset.filter(created_at__date__gte=date_from)
            
            date_to = request.query_params.get('date_to')
            if date_to:
                queryset = queryset.filter(created_at__date__lte=date_to)
            
            serializer = QuoteSubmissionSerializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching my submissions: {e}")
            return Response(
                {'error': f'Error obteniendo solicitudes: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='approve', permission_classes=[IsAuthenticated])
    def approve(self, request, pk=None):
        """
        Aprueba una cotización y la prepara para generar instrucción de embarque.
        Cambia el estado de 'enviada' o 'cotizacion_generada' a 'aprobada'.
        """
        try:
            quote_submission = self.get_object()
            
            if quote_submission.status not in ['enviada', 'cotizacion_generada', 'cotizado', 'procesando_costos']:
                return Response(
                    {'error': f'No se puede aprobar una cotización con estado: {quote_submission.status}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            quote_submission.status = 'aprobada'
            quote_submission.save(update_fields=['status', 'updated_at'])
            
            try:
                subject = f"Cotización Aprobada: {quote_submission.submission_number}"
                message = f"""
Estimado Equipo de Operaciones,

La cotización {quote_submission.submission_number} ha sido APROBADA por el cliente.

DATOS DEL CLIENTE:
- Empresa: {quote_submission.company_name}
- Contacto: {quote_submission.contact_name}
- Email: {quote_submission.contact_email}
- Teléfono: {quote_submission.contact_phone}

DETALLES DE LA COTIZACIÓN:
- Origen: {quote_submission.origin}
- Destino: {quote_submission.destination}
- Tipo Transporte: {quote_submission.transport_type}
- Descripción: {quote_submission.cargo_description or quote_submission.product_description or 'N/A'}

ACCIÓN REQUERIDA:
El cliente está listo para enviar la Instrucción de Embarque.
Por favor, prepare la documentación necesaria.

Sistema ImportaYa.ia
"""
                operations_email = getattr(settings, 'OPERATIONS_EMAIL', 'operaciones@importaya.ia')
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [operations_email],
                    fail_silently=True,
                )
            except Exception as email_error:
                logger.warning(f"Error sending approval notification: {email_error}")
            
            serializer = QuoteSubmissionSerializer(quote_submission)
            return Response({
                'success': True,
                'message': 'Cotización aprobada exitosamente',
                'quote_submission': serializer.data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error approving quote submission: {e}")
            return Response(
                {'error': f'Error al aprobar cotización: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='reject', permission_classes=[IsAuthenticated])
    def reject(self, request, pk=None):
        """
        Rechaza una cotización.
        Cambia el estado a 'rechazada' y guarda el motivo.
        """
        try:
            quote_submission = self.get_object()
            
            if quote_submission.status in ['completada', 'en_transito', 'ro_generado']:
                return Response(
                    {'error': f'No se puede rechazar una cotización con estado: {quote_submission.status}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            reason = request.data.get('reason', '')
            
            quote_submission.status = 'cancelada'
            quote_submission.rejection_reason = reason
            quote_submission.rejected_at = timezone.now()
            
            if reason:
                current_notes = quote_submission.notes or ''
                quote_submission.notes = f"{current_notes}\n[RECHAZADA] {timezone.now().strftime('%Y-%m-%d %H:%M')}: {reason}".strip()
            
            quote_submission.save(update_fields=['status', 'notes', 'updated_at'])
            
            try:
                subject = f"Cotización Rechazada: {quote_submission.submission_number}"
                message = f"""
Estimado Equipo Comercial,

La cotización {quote_submission.submission_number} ha sido RECHAZADA por el cliente.

DATOS DEL CLIENTE:
- Empresa: {quote_submission.company_name}
- Contacto: {quote_submission.contact_name}
- Email: {quote_submission.contact_email}

MOTIVO DEL RECHAZO:
{reason or 'No especificado'}

ACCIÓN RECOMENDADA:
Considere hacer seguimiento con el cliente para entender sus necesidades.

Sistema ImportaYa.ia
"""
                sales_email = getattr(settings, 'SALES_EMAIL', 'ventas@importaya.ia')
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [sales_email],
                    fail_silently=True,
                )
            except Exception as email_error:
                logger.warning(f"Error sending rejection notification: {email_error}")
            
            serializer = QuoteSubmissionSerializer(quote_submission)
            return Response({
                'success': True,
                'message': 'Cotización rechazada',
                'quote_submission': serializer.data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error rejecting quote submission: {e}")
            return Response(
                {'error': f'Error al rechazar cotización: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='generate-ro', permission_classes=[IsAuthenticated])
    def generate_ro(self, request, pk=None):
        """
        Genera el Routing Order (RO) para una cotización aprobada.
        Almacena los datos de la instrucción de embarque y genera el número de RO único.
        """
        try:
            quote_submission = self.get_object()
            
            if quote_submission.status != 'aprobada':
                return Response(
                    {'error': f'Solo se puede generar RO para cotizaciones aprobadas. Estado actual: {quote_submission.status}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            year = timezone.now().year
            month = timezone.now().month
            
            existing_ro_count = QuoteSubmission.objects.filter(
                ro_number__isnull=False,
                created_at__year=year,
                created_at__month=month
            ).count()
            
            ro_number = f"RO-{year}{str(month).zfill(2)}-{str(existing_ro_count + 1).zfill(5)}"
            
            quote_submission.ro_number = ro_number
            quote_submission.status = 'ro_generado'
            quote_submission.ro_generated_at = timezone.now()
            
            quote_submission.shipper_name = request.data.get('shipper_name', '')
            quote_submission.shipper_address = request.data.get('shipper_address', '')
            quote_submission.consignee_name = request.data.get('consignee_name', quote_submission.company_name)
            quote_submission.consignee_address = request.data.get('consignee_address', '')
            quote_submission.notify_party = request.data.get('notify_party', '')
            
            embarque_date = request.data.get('fecha_embarque_estimada')
            if embarque_date:
                quote_submission.estimated_departure_date = embarque_date
            
            notas = request.data.get('notas', '')
            if notas:
                current_notes = quote_submission.notes or ''
                quote_submission.notes = f"{current_notes}\n[INSTRUCCIÓN EMBARQUE] {notas}".strip()
            
            quote_submission.save()
            
            try:
                subject = f"Nuevo RO Generado: {ro_number}"
                message = f"""
Estimado Equipo de Operaciones,

Se ha generado un nuevo Routing Order (RO).

NÚMERO DE RO: {ro_number}

DATOS DEL EMBARQUE:
- Solicitud: {quote_submission.submission_number}
- Origen: {quote_submission.origin}
- Destino: {quote_submission.destination}
- Tipo Transporte: {quote_submission.transport_type}

SHIPPER:
- Nombre: {quote_submission.shipper_name}
- Dirección: {quote_submission.shipper_address}

CONSIGNATARIO:
- Nombre: {quote_submission.consignee_name}
- Dirección: {quote_submission.consignee_address}

NOTIFY PARTY: {quote_submission.notify_party or 'N/A'}

NOTAS: {notas or 'N/A'}

Por favor, coordine el embarque correspondiente.

Sistema ImportaYa.ia
"""
                operations_email = getattr(settings, 'OPERATIONS_EMAIL', 'operaciones@importaya.ia')
                forwarder_email = getattr(settings, 'FREIGHT_FORWARDER_EMAIL', 'forwarder@importaya.ia')
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [operations_email, forwarder_email],
                    fail_silently=True,
                )
            except Exception as email_error:
                logger.warning(f"Error sending RO notification: {email_error}")
            
            serializer = QuoteSubmissionSerializer(quote_submission)
            return Response({
                'success': True,
                'message': f'Routing Order generado exitosamente: {ro_number}',
                'ro_number': ro_number,
                'quote_submission': serializer.data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error generating RO: {e}")
            return Response(
                {'error': f'Error al generar RO: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CostRateViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = CostRate.objects.filter(is_active=True)
    serializer_class = CostRateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['origin', 'destination', 'transport_type', 'source', 'is_active']
    search_fields = ['origin', 'destination', 'provider_name']
    
    @action(detail=False, methods=['get'], url_path='by-route')
    def get_by_route(self, request):
        """Obtiene tarifa para una ruta específica"""
        origin = request.query_params.get('origin')
        destination = request.query_params.get('destination')
        transport_type = request.query_params.get('transport_type')
        
        if not all([origin, destination, transport_type]):
            return Response({'error': 'Parámetros requeridos: origin, destination, transport_type'}, status=status.HTTP_400_BAD_REQUEST)
        
        rate = CostRate.objects.filter(
            origin__icontains=origin,
            destination__icontains=destination,
            transport_type=transport_type,
            is_active=True
        ).order_by('-created_at').first()
        
        if rate:
            return Response(CostRateSerializer(rate).data)
        return Response({'error': 'No hay tarifa disponible para esta ruta'}, status=status.HTTP_404_NOT_FOUND)


class BulkLeadImportViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    queryset = BulkLeadImport.objects.all()
    serializer_class = BulkLeadImportSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'file_type']
    
    @action(detail=False, methods=['get'], url_path='template')
    def generate_template(self, request):
        file_type = request.query_params.get('file_type', 'csv')
        # EXACTAMENTE los nombres de columna que espera el parser de upload
        columns = ['Empresa', 'Nombre Contacto', 'Correo', 'Teléfono', 'WhatsApp', 'País', 'Ciudad', 'Notas', '¿Es Importador Activo?', 'RUC']
        
        if file_type == 'csv':
            output = io.StringIO()
            # Usar punto y coma como delimitador (estándar en Ecuador)
            writer = csv.writer(output, delimiter=';', lineterminator='\r\n')
            writer.writerow(columns)
            
            # Agregar UTF-8 BOM para que Excel reconozca caracteres acentuados
            csv_content = output.getvalue()
            csv_bytes = '\ufeff' + csv_content  # UTF-8 BOM
            
            response = HttpResponse(csv_bytes.encode('utf-8'), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="plantilla_leads.csv"'
            return response
        
        elif file_type == 'xlsx':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Leads'
            ws.append(columns)
            
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="plantilla_leads.xlsx"'
            return response
        
        elif file_type == 'xls':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Leads'
            ws.append(columns)
            
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="plantilla_leads.xls"'
            return response
        
        elif file_type == 'txt':
            output = '\t'.join(columns)
            response = HttpResponse(output, content_type='text/plain')
            response['Content-Disposition'] = 'attachment; filename="plantilla_leads.txt"'
            return response
        
        return Response({'error': 'Tipo de archivo no válido'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'], url_path='upload')
    def upload_file(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Detectar tipo de archivo automáticamente si no se proporciona
        file_type = request.data.get('file_type', '').lower()
        if not file_type:
            filename_lower = file_obj.name.lower()
            if filename_lower.endswith('.xlsx'):
                file_type = 'xlsx'
            elif filename_lower.endswith('.xls'):
                file_type = 'xls'
            elif filename_lower.endswith('.csv'):
                file_type = 'csv'
            elif filename_lower.endswith('.txt'):
                file_type = 'txt'
            else:
                return Response({'error': f'Tipo de archivo no soportado: {file_obj.name}'}, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"Archivo: {file_obj.name}, Tipo detectado: {file_type}")
        
        # CRITICAL: Leer contenido ANTES de guardar en modelo (Django consume el archivo)
        try:
            file_content = file_obj.read()
            logger.info(f"Contenido leído: {len(file_content)} bytes")
        except Exception as e:
            logger.exception(f"Error leyendo archivo: {str(e)}")
            return Response({'error': f'Error leyendo archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        file_obj.seek(0)  # Reset para que Django pueda guardarlo
        
        import_record = BulkLeadImport.objects.create(
            file=file_obj,
            file_type=file_type,
            status='procesando'
        )
        
        try:
            logger.info(f"Iniciando parse de archivo {file_type}")
            rows = self._parse_file_content(file_content, file_type)
            logger.info(f"Parse exitoso, {len(rows)} filas encontradas")
            
            import_record.total_rows = len(rows)
            import_record.save()
            
            error_list = []
            for idx, row in enumerate(rows):
                try:
                    # Clean whitespace from all values
                    row = {k: (str(v).strip() if v else '') for k, v in row.items()}
                    
                    # Map both Spanish and English column names
                    company_name = row.get('Empresa') or row.get('company_name') or 'N/A'
                    contact_name = row.get('Nombre Contacto') or row.get('contact_name') or ''
                    email = row.get('Correo') or row.get('email') or ''
                    phone = row.get('Teléfono') or row.get('phone') or ''
                    whatsapp = row.get('WhatsApp') or row.get('whatsapp') or ''
                    country = row.get('País') or row.get('country') or 'Ecuador'
                    city = row.get('Ciudad') or row.get('city') or ''
                    notes = row.get('Notas') or row.get('notes') or ''
                    is_active_importer_val = row.get('¿Es Importador Activo?') or row.get('is_active_importer') or 'False'
                    ruc_val = row.get('RUC') or row.get('ruc') or ''
                    
                    # Dividir contact_name en first_name y last_name
                    name_parts = contact_name.strip().split(' ', 1) if contact_name else ['', '']
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                    
                    logger.debug(f"Fila {idx+1}: Empresa='{company_name}', Nombre='{first_name}', Apellido='{last_name}', Email='{email}'")
                    
                    lead = Lead.objects.create(
                        company_name=company_name[:255],
                        first_name=first_name[:255],
                        last_name=last_name[:255],
                        email=email[:255],
                        phone=phone[:50],
                        whatsapp=whatsapp[:50],
                        country=country[:100],
                        city=city[:100],
                        source='bulk_import',
                        notes=notes,
                        is_active_importer=str(is_active_importer_val).lower() in ['true', 'sí', 'si', '1', 'verdadero'],
                        ruc=ruc_val[:13]
                    )
                    import_record.imported_rows += 1
                    logger.info(f"✅ Fila {idx+1} creada exitosamente")
                except Exception as e:
                    import_record.error_rows += 1
                    error_msg = f"Fila {idx+1}: {str(e)}"
                    error_list.append(error_msg)
                    logger.error(error_msg)
            
            import_record.error_details = '\n'.join(error_list)
            import_record.status = 'completado'
            import_record.save()
            
            return Response(BulkLeadImportSerializer(import_record).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            error_tb = traceback.format_exc()
            logger.error(f"Error en upload_file: {str(e)}\n{error_tb}")
            import_record.status = 'error'
            import_record.error_details = f"{str(e)}\n\nTraceback:\n{error_tb}"
            import_record.save()
            return Response({
                'error': str(e),
                'detail': error_tb,
                'import_id': import_record.id
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _parse_file_content(self, content, file_type):
        """Parse file content that was already read into memory"""
        rows = []
        try:
            if file_type == 'csv':
                try:
                    text = content.decode('utf-8')
                except UnicodeDecodeError:
                    text = content.decode('latin-1')
                
                # Remover BOM si existe
                if text.startswith('\ufeff'):
                    text = text[1:]
                    logger.info("BOM UTF-8 removido del contenido")
                
                # Auto-detect delimiter (comma or semicolon)
                first_line = text.split('\n')[0] if text else ''
                delimiter = ';' if ';' in first_line else ','
                logger.info(f"CSV delimiter detectado: '{delimiter}'")
                
                reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                rows = [row for row in reader if any(row.values())]
                logger.info(f"CSV leído con {len(rows)} filas")
            
            elif file_type == 'xlsx':
                import openpyxl
                logger.info(f"Leyendo XLSX con openpyxl...")
                
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                logger.info(f"Sheet activa: {ws.title}, max_row: {ws.max_row}")
                
                if not ws or ws.max_row < 2:
                    logger.warning("Archivo Excel está vacío")
                    return []
                
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                
                logger.info(f"Headers encontrados: {headers}")
                
                if not headers:
                    logger.warning("No se encontraron headers en el Excel")
                    return []
                
                for row_idx in range(2, ws.max_row + 1):
                    values = []
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        val = str(cell.value).strip() if cell.value else ''
                        values.append(val)
                    
                    if any(values):
                        row_dict = dict(zip(headers, values))
                        rows.append(row_dict)
                        logger.debug(f"Fila {row_idx}: {row_dict}")
                
                wb.close()
                logger.info(f"Total filas leídas: {len(rows)}")
            
            elif file_type == 'txt':
                try:
                    text = content.decode('utf-8')
                except UnicodeDecodeError:
                    text = content.decode('latin-1')
                
                for line in text.split('\n'):
                    if line.strip():
                        parts = [p.strip() for p in line.split('\t')]
                        rows.append({
                            'Empresa': parts[0] if len(parts) > 0 else '',
                            'Nombre Contacto': parts[1] if len(parts) > 1 else ''
                        })
        
        except Exception as e:
            error_tb = traceback.format_exc()
            logger.error(f"Error en _parse_file {file_type}: {str(e)}\n{error_tb}")
            raise ValueError(f"Error parseando archivo {file_type}: {str(e)}")
        
        return rows


class LeadCotizacionViewSet(viewsets.ModelViewSet):
    """ViewSet for LEAD user quotation requests - restricted to LEAD role users only"""
    serializer_class = LeadCotizacionSerializer
    permission_classes = [IsAuthenticated, IsLeadUser]
    filterset_fields = ['estado', 'tipo_carga']
    search_fields = ['numero_cotizacion', 'descripcion_mercancia', 'origen_pais']
    ordering_fields = ['fecha_creacion', 'total_usd', 'estado']
    
    def get_queryset(self):
        queryset = LeadCotizacion.objects.filter(lead_user=self.request.user)
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return LeadCotizacionDetailSerializer
        return LeadCotizacionSerializer
    
    def perform_create(self, serializer):
        serializer.save(lead_user=self.request.user)
    
    @action(detail=False, methods=['get'], url_path='por-estado')
    def por_estado(self, request):
        """Get quotations grouped by status"""
        queryset = self.get_queryset()
        estados = {}
        for estado_code, estado_name in LeadCotizacion.ESTADO_CHOICES:
            estados[estado_code] = {
                'nombre': str(estado_name),
                'count': queryset.filter(estado=estado_code).count()
            }
        return Response(estados)
    
    @action(detail=True, methods=['post'], url_path='generar-escenarios')
    def generar_escenarios(self, request, pk=None):
        """Generate quote scenarios with strict route-aligned rate matching"""
        cotizacion = self.get_object()
        from decimal import Decimal
        
        cotizacion.escenarios.all().delete()
        scenarios_created = []
        
        tipo_carga_map = {'aerea': 'aereo', 'maritima': 'maritimo', 'terrestre': 'terrestre'}
        primary_transport = tipo_carga_map.get(cotizacion.tipo_carga, 'aereo')
        
        def find_freight_rate_strict(transport_prefix):
            """Find freight rate with strict route alignment - requires origin+destination match"""
            base_qs = FreightRate.objects.filter(transport_type__startswith=transport_prefix, is_active=True)
            return base_qs.filter(
                origin_country__iexact=cotizacion.origen_pais,
                destination_country__iexact='Ecuador'
            ).first()
        
        air_freight = find_freight_rate_strict('aereo')
        sea_freight = find_freight_rate_strict('maritimo')
        
        if not air_freight and not sea_freight:
            return Response({
                'error': 'No hay tarifas de flete configuradas para esta ruta',
                'detalles': [
                    f'Origen: {cotizacion.origen_pais}',
                    f'Destino: Ecuador ({cotizacion.destino_ciudad})',
                    'No se encontraron tarifas aéreas ni marítimas para esta ruta'
                ],
                'accion_requerida': 'Configure tarifas de flete para la ruta solicitada antes de generar escenarios',
                'configuracion_necesaria': {
                    'modelo': 'FreightRate',
                    'campos_requeridos': {
                        'origin_country': cotizacion.origen_pais,
                        'destination_country': 'Ecuador',
                        'transport_type': 'aereo o maritimo_*'
                    }
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        insurance = InsuranceRate.objects.filter(is_active=True).first()
        if not insurance:
            return Response({
                'error': 'No hay tarifa de seguro configurada',
                'accion_requerida': 'Configure al menos una tarifa de seguro activa'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        brokerage = CustomsBrokerageRate.objects.filter(is_active=True).first()
        if not brokerage:
            return Response({
                'error': 'No hay tarifa de agenciamiento aduanero configurada',
                'accion_requerida': 'Configure al menos una tarifa de agenciamiento activa'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        inland = None
        if cotizacion.requiere_transporte_interno:
            inland = InlandTransportQuoteRate.objects.filter(
                destination_city__iexact=cotizacion.destino_ciudad, is_active=True
            ).first()
            if not inland:
                return Response({
                    'error': f'No hay tarifa de transporte interno a {cotizacion.destino_ciudad}',
                    'accion_requerida': f'Configure tarifa de transporte interno para {cotizacion.destino_ciudad}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        customs_rate = CustomsDutyRate.objects.filter(is_active=True).first()
        
        air_rate = air_freight.rate_usd if air_freight else Decimal('4.50')
        sea_rate = sea_freight.rate_usd if sea_freight else Decimal('0.85')
        insurance_pct = (insurance.rate_percentage / 100) if insurance else Decimal('0.005')
        brokerage_fee = brokerage.fixed_rate_usd if brokerage else Decimal('150')
        inland_cost = inland.rate_usd if inland else Decimal('80')
        air_transit = air_freight.transit_time_days if air_freight else 5
        sea_transit = sea_freight.transit_time_days if sea_freight else 25
        
        def calculate_duties(flete, seguro):
            """Calculate customs duties based on CIF value"""
            from .gemini_service import SENAE_TRIBUTOS_2025
            cif_value = cotizacion.valor_mercancia_usd + flete + seguro
            if customs_rate:
                duties = customs_rate.calculate_duties(cif_value)
            else:
                ad_valorem = cif_value * SENAE_TRIBUTOS_2025['ad_valorem_default']
                fodinfa = cif_value * SENAE_TRIBUTOS_2025['fodinfa_rate']
                base_iva = cif_value + ad_valorem + fodinfa
                iva = base_iva * SENAE_TRIBUTOS_2025['iva_rate']
                duties = {
                    'ad_valorem': ad_valorem,
                    'fodinfa': fodinfa,
                    'ice': Decimal('0'),
                    'salvaguardia': Decimal('0'),
                    'iva': iva,
                    'total': ad_valorem + fodinfa + iva
                }
            return duties
        
        def create_scenario_with_items(nombre, tipo, flete, transit_days, freight_rate_obj, notas):
            seguro = cotizacion.valor_mercancia_usd * insurance_pct if cotizacion.requiere_seguro else Decimal('0')
            transporte = inland_cost if cotizacion.requiere_transporte_interno else Decimal('0')
            duties = calculate_duties(flete, seguro)
            
            total = flete + seguro + brokerage_fee + transporte + duties['total'] + Decimal('25')
            
            scenario = QuoteScenario.objects.create(
                cotizacion=cotizacion,
                nombre=nombre,
                tipo=tipo,
                freight_rate=freight_rate_obj,
                insurance_rate=insurance if cotizacion.requiere_seguro else None,
                brokerage_rate=brokerage,
                inland_transport_rate=inland if cotizacion.requiere_transporte_interno else None,
                flete_usd=flete,
                seguro_usd=seguro,
                agenciamiento_usd=brokerage_fee,
                transporte_interno_usd=transporte,
                otros_usd=duties['total'] + Decimal('25'),
                total_usd=total,
                tiempo_transito_dias=transit_days,
                notas=notas
            )
            
            line_items_data = [
                ('flete', f'Flete Internacional ({cotizacion.peso_kg} kg)', flete),
                ('seguro', f'Seguro de Carga ({insurance_pct*100:.2f}%)', seguro),
                ('arancel', f'Ad Valorem ({customs_rate.ad_valorem_percentage if customs_rate else 10}%)', duties['ad_valorem']),
                ('fodinfa', 'FODINFA (0.5%)', duties['fodinfa']),
                ('iva', 'IVA (15%)', duties['iva']),
                ('ice', 'ICE', duties.get('ice', Decimal('0'))),
                ('salvaguardia', 'Salvaguardia', duties.get('salvaguardia', Decimal('0'))),
                ('agenciamiento', 'Agenciamiento Aduanero', brokerage_fee),
                ('transporte_interno', f'Transporte a {cotizacion.destino_ciudad}', transporte),
                ('otros', 'Gastos Administrativos', Decimal('25')),
            ]
            
            for cat, desc, amount in line_items_data:
                if amount > 0:
                    QuoteLineItem.objects.create(
                        escenario=scenario,
                        categoria=cat,
                        descripcion=desc,
                        cantidad=Decimal('1'),
                        precio_unitario_usd=amount,
                        subtotal_usd=amount,
                        es_estimado=(cat in ['arancel', 'iva', 'fodinfa', 'ice', 'salvaguardia'])
                    )
            
            return scenario
        
        air_flete = cotizacion.peso_kg * air_rate
        air_scenario = create_scenario_with_items(
            'Opción Aérea Express', 'express', air_flete, air_transit, air_freight,
            'Transporte aéreo con tiempo de tránsito reducido'
        )
        scenarios_created.append(air_scenario)
        
        sea_flete = cotizacion.peso_kg * sea_rate
        sea_scenario = create_scenario_with_items(
            'Opción Marítima Económica', 'economico', sea_flete, sea_transit, sea_freight,
            'Transporte marítimo con mejor precio'
        )
        scenarios_created.append(sea_scenario)
        
        std_rate = (air_rate + sea_rate) / 2
        std_flete = cotizacion.peso_kg * std_rate
        std_scenario = create_scenario_with_items(
            'Opción Estándar', 'estandar', std_flete, int((air_transit + sea_transit) / 2), None,
            'Balance entre precio y tiempo de tránsito'
        )
        scenarios_created.append(std_scenario)
        
        cotizacion.estado = 'cotizado'
        cotizacion.save()
        
        return Response({
            'message': f'Se generaron {len(scenarios_created)} escenarios de cotización',
            'ruta': f'{cotizacion.origen_pais} → Ecuador ({cotizacion.destino_ciudad})',
            'tarifas_verificadas': True,
            'escenarios': QuoteScenarioSerializer(scenarios_created, many=True).data,
            'cotizacion': LeadCotizacionDetailSerializer(cotizacion).data
        })
    
    @action(detail=True, methods=['post'], url_path='seleccionar-escenario')
    def seleccionar_escenario(self, request, pk=None):
        """Select a quote scenario"""
        cotizacion = self.get_object()
        serializer = QuoteScenarioSelectSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        scenario_id = serializer.validated_data['scenario_id']
        try:
            scenario = cotizacion.escenarios.get(id=scenario_id)
        except QuoteScenario.DoesNotExist:
            return Response({'error': 'Escenario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        cotizacion.escenarios.update(is_selected=False)
        scenario.is_selected = True
        scenario.save()
        
        cotizacion.flete_usd = scenario.flete_usd
        cotizacion.seguro_usd = scenario.seguro_usd
        cotizacion.aduana_usd = scenario.agenciamiento_usd
        cotizacion.transporte_interno_usd = scenario.transporte_interno_usd
        cotizacion.otros_usd = scenario.otros_usd
        cotizacion.total_usd = scenario.total_usd
        cotizacion.save()
        
        return Response({
            'message': 'Escenario seleccionado exitosamente',
            'escenario': QuoteScenarioSerializer(scenario).data,
            'cotizacion': LeadCotizacionDetailSerializer(cotizacion).data
        })
    
    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Approve a quotation"""
        cotizacion = self.get_object()
        
        if cotizacion.estado != 'cotizado':
            return Response(
                {'error': 'Solo se pueden aprobar cotizaciones con estado "cotizado"'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        cotizacion.aprobar()
        return Response({
            'message': 'Cotización aprobada exitosamente',
            'cotizacion': LeadCotizacionDetailSerializer(cotizacion).data
        })
    
    @action(detail=True, methods=['post'], url_path='instruccion-embarque')
    def instruccion_embarque(self, request, pk=None):
        """Send shipping instructions and generate RO number"""
        cotizacion = self.get_object()
        
        if cotizacion.estado != 'aprobada':
            return Response(
                {'error': 'Solo se pueden enviar instrucciones para cotizaciones aprobadas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = LeadCotizacionInstruccionSerializer(data=request.data)
        if serializer.is_valid():
            cotizacion.shipper_name = serializer.validated_data['shipper_name']
            cotizacion.shipper_address = serializer.validated_data['shipper_address']
            cotizacion.consignee_name = serializer.validated_data['consignee_name']
            cotizacion.consignee_address = serializer.validated_data['consignee_address']
            cotizacion.notify_party = serializer.validated_data.get('notify_party', '')
            cotizacion.fecha_embarque_estimada = serializer.validated_data['fecha_embarque_estimada']
            cotizacion.save()
            
            ro_number = cotizacion.generar_ro()
            
            return Response({
                'success': True,
                'message': f'RO generado exitosamente: {ro_number}',
                'ro_number': ro_number,
                'cotizacion': LeadCotizacionDetailSerializer(cotizacion).data
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class FreightRateViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    """CRUD for freight rates"""
    queryset = FreightRate.objects.all()
    serializer_class = FreightRateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['transport_mode', 'origin_port', 'destination_port', 'is_active']
    search_fields = ['origin_port', 'destination_port', 'carrier_name']
    ordering_fields = ['rate_usd', 'transit_time_days', 'created_at']


class InsuranceRateViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    """CRUD for insurance rates"""
    queryset = InsuranceRate.objects.all()
    serializer_class = InsuranceRateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['coverage_type', 'is_active']
    search_fields = ['name']
    ordering_fields = ['rate_percentage', 'created_at']
    
    @action(detail=True, methods=['post'], url_path='calculate-premium')
    def calculate_premium(self, request, pk=None):
        """Calculate insurance premium for a cargo value"""
        rate = self.get_object()
        cargo_value = request.data.get('cargo_value')
        if cargo_value is None:
            return Response({'error': 'cargo_value is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from decimal import Decimal
            premium = rate.calculate_premium(Decimal(str(cargo_value)))
            return Response({
                'cargo_value': str(cargo_value),
                'premium': str(premium),
                'rate_percentage': str(rate.rate_percentage)
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CustomsDutyRateViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only access to SENAE customs duty rates (government tariffs)"""
    queryset = CustomsDutyRate.objects.filter(is_active=True)
    serializer_class = CustomsDutyRateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_active', 'requires_import_license', 'requires_phytosanitary', 'requires_inen_certification']
    search_fields = ['hs_code', 'description']
    ordering_fields = ['hs_code', 'ad_valorem_percentage', 'created_at']
    
    @action(detail=True, methods=['post'], url_path='calculate-duties')
    def calculate_duties(self, request, pk=None):
        """Calculate all customs duties for a CIF value"""
        rate = self.get_object()
        serializer = CustomsDutyCalculationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        cif_value = serializer.validated_data['cif_value_usd']
        quantity = serializer.validated_data.get('quantity', 1)
        
        duties = rate.calculate_duties(cif_value, quantity)
        return Response({
            'cif_value_usd': str(cif_value),
            'quantity': quantity,
            'hs_code': rate.hs_code,
            'duties': {k: str(v) for k, v in duties.items()}
        })
    
    @action(detail=False, methods=['get'], url_path='search-hs')
    def search_hs(self, request):
        """Search customs duty rates by HS code prefix"""
        hs_code = request.query_params.get('code', '')
        if len(hs_code) < 2:
            return Response({'error': 'Provide at least 2 digits of HS code'}, status=status.HTTP_400_BAD_REQUEST)
        
        rates = CustomsDutyRate.objects.filter(hs_code__startswith=hs_code, is_active=True)[:20]
        serializer = self.get_serializer(rates, many=True)
        return Response(serializer.data)


class InlandTransportQuoteRateViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    """CRUD for inland transport rates"""
    queryset = InlandTransportQuoteRate.objects.all()
    serializer_class = InlandTransportQuoteRateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['origin_city', 'destination_city', 'vehicle_type', 'is_active']
    search_fields = ['origin_city', 'destination_city', 'carrier_name']
    ordering_fields = ['rate_usd', 'distance_km', 'created_at']


class CustomsBrokerageRateViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    """CRUD for customs brokerage fees"""
    queryset = CustomsBrokerageRate.objects.all()
    serializer_class = CustomsBrokerageRateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['service_type', 'is_active']
    search_fields = ['name']
    ordering_fields = ['fixed_rate_usd', 'created_at']
    
    @action(detail=True, methods=['post'], url_path='calculate-fee')
    def calculate_fee(self, request, pk=None):
        """Calculate brokerage fee for a CIF value"""
        rate = self.get_object()
        serializer = BrokerageFeeCalculationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        cif_value = serializer.validated_data['cif_value_usd']
        fee = rate.calculate_fee(cif_value)
        
        return Response({
            'cif_value_usd': str(cif_value),
            'fee_usd': str(fee),
            'service_type': rate.get_service_type_display()
        })


class ShipmentViewSet(viewsets.ModelViewSet):
    """ViewSet for shipment tracking - LEAD users only"""
    permission_classes = [IsAuthenticated, IsLeadUser]
    filterset_fields = ['current_status', 'transport_type']
    search_fields = ['tracking_number', 'bl_awb_number', 'container_number', 'description']
    ordering_fields = ['created_at', 'estimated_arrival', 'current_status']
    
    def get_queryset(self):
        return Shipment.objects.filter(lead_user=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ShipmentCreateSerializer
        if self.action == 'retrieve':
            return ShipmentDetailSerializer
        return ShipmentSerializer
    
    def perform_create(self, serializer):
        serializer.save(lead_user=self.request.user)
    
    @action(detail=False, methods=['get'], url_path='por-estado')
    def por_estado(self, request):
        """Get shipment counts by status for dashboard"""
        queryset = self.get_queryset()
        estados = {}
        for status_code, status_name in Shipment.STATUS_CHOICES:
            estados[status_code] = {
                'nombre': str(status_name),
                'count': queryset.filter(current_status=status_code).count()
            }
        return Response(estados)
    
    @action(detail=True, methods=['post'], url_path='agregar-evento')
    def agregar_evento(self, request, pk=None):
        """Add a tracking event to a shipment"""
        shipment = self.get_object()
        serializer = AddTrackingEventSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        event = ShipmentTracking.objects.create(
            shipment=shipment,
            status=serializer.validated_data['status'],
            location=serializer.validated_data['location'],
            description=serializer.validated_data['description'],
            event_datetime=serializer.validated_data['event_datetime']
        )
        
        return Response({
            'message': 'Evento de tracking agregado',
            'evento': ShipmentTrackingSerializer(event).data,
            'shipment': ShipmentDetailSerializer(shipment).data
        })
    
    @action(detail=True, methods=['get'], url_path='historial')
    def historial(self, request, pk=None):
        """Get full tracking history for a shipment"""
        shipment = self.get_object()
        events = shipment.tracking_events.all()
        return Response({
            'tracking_number': shipment.tracking_number,
            'current_status': shipment.get_current_status_display(),
            'eventos': ShipmentTrackingSerializer(events, many=True).data
        })
    
    @action(detail=False, methods=['get'], url_path='buscar/(?P<tracking_number>[^/.]+)')
    def buscar(self, request, tracking_number=None):
        """Search shipment by tracking number"""
        try:
            shipment = self.get_queryset().get(tracking_number=tracking_number)
            return Response(ShipmentDetailSerializer(shipment).data)
        except Shipment.DoesNotExist:
            return Response({'error': 'Embarque no encontrado'}, status=status.HTTP_404_NOT_FOUND)


class PreLiquidationViewSet(viewsets.ModelViewSet):
    """ViewSet for pre-liquidation with AI HS code classification"""
    serializer_class = PreLiquidationSerializer
    permission_classes = [IsAuthenticated, IsLeadUser]
    
    def get_queryset(self):
        user = self.request.user
        from django.db.models import Q
        return PreLiquidation.objects.filter(
            Q(cotizacion__lead_user=user) | Q(quote_submission__owner=user)
        )
    
    def create(self, request, *args, **kwargs):
        """Create pre-liquidation supporting both cotizacion (LeadCotizacion) and quote_submission_id (QuoteSubmission)"""
        from decimal import Decimal
        
        cotizacion_id = request.data.get('cotizacion')
        quote_submission_id = request.data.get('quote_submission_id')
        
        if not cotizacion_id and not quote_submission_id:
            return Response(
                {'error': 'Se requiere cotizacion o quote_submission_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            quote_submission = None
            lead_cotizacion = None
            
            product_description = request.data.get('product_description', '')
            fob_value = Decimal(str(request.data.get('fob_value_usd', 0) or 0))
            freight_usd = Decimal(str(request.data.get('freight_usd', 0) or 0))
            insurance_usd = Decimal(str(request.data.get('insurance_usd', 0) or 0))
            
            if quote_submission_id:
                try:
                    quote_submission = QuoteSubmission.objects.get(
                        id=quote_submission_id,
                        owner=request.user
                    )
                except QuoteSubmission.DoesNotExist:
                    return Response(
                        {'error': 'QuoteSubmission no encontrado'},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                if not product_description:
                    product_description = quote_submission.cargo_description or quote_submission.product_description or ''
                if fob_value == 0 and quote_submission.fob_value_usd:
                    fob_value = quote_submission.fob_value_usd
                    
                pre_liq = PreLiquidation.objects.create(
                    quote_submission=quote_submission,
                    product_description=product_description,
                    fob_value_usd=fob_value,
                    freight_usd=freight_usd,
                    insurance_usd=insurance_usd,
                )
            elif cotizacion_id:
                try:
                    lead_cotizacion = LeadCotizacion.objects.get(
                        id=cotizacion_id,
                        lead_user=request.user
                    )
                except LeadCotizacion.DoesNotExist:
                    return Response(
                        {'error': 'LeadCotizacion no encontrada'},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                pre_liq = PreLiquidation.objects.create(
                    cotizacion=lead_cotizacion,
                    product_description=product_description or lead_cotizacion.descripcion_mercancia,
                    fob_value_usd=fob_value or lead_cotizacion.valor_mercancia_usd,
                    freight_usd=freight_usd or (lead_cotizacion.flete_usd or Decimal('0')),
                    insurance_usd=insurance_usd or (lead_cotizacion.seguro_usd or Decimal('0')),
                )
            
            pre_liq.calculate_cif()
            self._suggest_hs_code(pre_liq)
            self._calculate_estimated_duties(pre_liq)
            pre_liq.save()
            
            return Response(PreLiquidationSerializer(pre_liq).data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error creating pre-liquidation: {e}")
            return Response(
                {'error': f'Error al calcular pre-liquidación: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def perform_create(self, serializer):
        pre_liq = serializer.save()
        pre_liq.calculate_cif()
        self._suggest_hs_code(pre_liq)
        self._calculate_estimated_duties(pre_liq)
        pre_liq.save()
    
    def _calculate_estimated_duties(self, pre_liq):
        """Calculate estimated duties using suggested HS code or default rates"""
        from decimal import Decimal
        cif = pre_liq.cif_value_usd
        
        try:
            if pre_liq.suggested_hs_code and pre_liq.suggested_hs_code != '9999.00.00':
                duty_rate = CustomsDutyRate.objects.filter(
                    hs_code=pre_liq.suggested_hs_code, 
                    is_active=True
                ).first()
                if duty_rate:
                    duties = duty_rate.calculate_duties(cif)
                    pre_liq.ad_valorem_usd = duties['ad_valorem']
                    pre_liq.fodinfa_usd = duties['fodinfa']
                    pre_liq.ice_usd = duties['ice']
                    pre_liq.salvaguardia_usd = duties['salvaguardia']
                    pre_liq.iva_usd = duties['iva']
                    pre_liq.total_tributos_usd = duties['total']
                    return
        except Exception:
            pass
        
        from .gemini_service import SENAE_TRIBUTOS_2025
        pre_liq.ad_valorem_usd = cif * SENAE_TRIBUTOS_2025['ad_valorem_default']
        pre_liq.fodinfa_usd = cif * SENAE_TRIBUTOS_2025['fodinfa_rate']
        base_iva = cif + pre_liq.ad_valorem_usd + pre_liq.fodinfa_usd
        pre_liq.iva_usd = base_iva * SENAE_TRIBUTOS_2025['iva_rate']
        pre_liq.ice_usd = Decimal('0')
        pre_liq.salvaguardia_usd = Decimal('0')
        pre_liq.total_tributos_usd = pre_liq.ad_valorem_usd + pre_liq.fodinfa_usd + pre_liq.iva_usd
    
    def _suggest_hs_code(self, pre_liq):
        """AI-powered HS code suggestion using Gemini AI"""
        from .gemini_service import suggest_hs_code
        
        origin_country = ''
        if pre_liq.cotizacion:
            origin_country = getattr(pre_liq.cotizacion, 'origin', '') or getattr(pre_liq.cotizacion, 'origen_pais', '') or ''
        
        fob_value = float(pre_liq.fob_value_usd) if pre_liq.fob_value_usd else 0
        
        result = suggest_hs_code(
            product_description=pre_liq.product_description,
            origin_country=origin_country,
            fob_value=fob_value
        )
        
        pre_liq.suggested_hs_code = result.get('suggested_hs_code', '9999.00.00')
        pre_liq.hs_code_confidence = result.get('confidence', 30)
        
        reasoning = result.get('reasoning', 'Clasificacion pendiente')
        notes = result.get('notes', '')
        ai_status = result.get('ai_status', 'unknown')
        
        if ai_status.startswith('fallback'):
            reasoning = f"[Sistema de respaldo] {reasoning}"
        
        if notes and notes not in reasoning:
            reasoning = f"{reasoning} | {notes}"
        
        pre_liq.ai_reasoning = reasoning
        pre_liq.ai_status = ai_status
        
        pre_liq.requires_permit = result.get('requires_permit', False)
        permit_info = result.get('permit_info', None)
        if permit_info and isinstance(permit_info, dict):
            pre_liq.permit_institucion = permit_info.get('institucion', '')
            pre_liq.permit_tipo = permit_info.get('permiso', '')
            pre_liq.permit_descripcion = permit_info.get('descripcion', '')
            pre_liq.permit_tramite_previo = permit_info.get('tramite_previo', False)
            pre_liq.permit_tiempo_estimado = permit_info.get('tiempo_estimado', '')
        
        special_taxes = result.get('special_taxes', [])
        if special_taxes and isinstance(special_taxes, list):
            pre_liq.special_taxes = special_taxes
    
    @action(detail=True, methods=['post'], url_path='confirmar-hs')
    def confirmar_hs(self, request, pk=None):
        """Confirm HS code and calculate duties with CIF recalculation"""
        pre_liq = self.get_object()
        hs_code = request.data.get('hs_code', pre_liq.suggested_hs_code)
        
        if not hs_code:
            return Response({'error': 'Se requiere código HS'}, status=status.HTTP_400_BAD_REQUEST)
        
        pre_liq.confirmed_hs_code = hs_code
        pre_liq.calculate_cif()
        
        try:
            duty_rate = CustomsDutyRate.objects.get(hs_code=hs_code, is_active=True)
            duties = duty_rate.calculate_duties(pre_liq.cif_value_usd)
            
            pre_liq.ad_valorem_usd = duties['ad_valorem']
            pre_liq.fodinfa_usd = duties['fodinfa']
            pre_liq.ice_usd = duties['ice']
            pre_liq.salvaguardia_usd = duties['salvaguardia']
            pre_liq.iva_usd = duties['iva']
            pre_liq.total_tributos_usd = duties['total']
            pre_liq.is_confirmed = True
            pre_liq.confirmed_by = request.user
            pre_liq.confirmed_at = timezone.now()
            pre_liq.save()
            
            return Response({
                'message': 'Código HS confirmado y tributos calculados',
                'hs_code': hs_code,
                'cif_value_usd': str(pre_liq.cif_value_usd),
                'tributos': {
                    'ad_valorem': str(pre_liq.ad_valorem_usd),
                    'fodinfa': str(pre_liq.fodinfa_usd),
                    'ice': str(pre_liq.ice_usd),
                    'salvaguardia': str(pre_liq.salvaguardia_usd),
                    'iva': str(pre_liq.iva_usd),
                    'total': str(pre_liq.total_tributos_usd)
                },
                'pre_liquidacion': PreLiquidationSerializer(pre_liq).data
            })
        except CustomsDutyRate.DoesNotExist:
            from decimal import Decimal
            from .gemini_service import SENAE_TRIBUTOS_2025
            cif = pre_liq.cif_value_usd
            pre_liq.ad_valorem_usd = cif * SENAE_TRIBUTOS_2025['ad_valorem_default']
            pre_liq.fodinfa_usd = cif * SENAE_TRIBUTOS_2025['fodinfa_rate']
            base_iva = cif + pre_liq.ad_valorem_usd + pre_liq.fodinfa_usd
            pre_liq.iva_usd = base_iva * SENAE_TRIBUTOS_2025['iva_rate']
            pre_liq.ice_usd = Decimal('0')
            pre_liq.salvaguardia_usd = Decimal('0')
            pre_liq.total_tributos_usd = pre_liq.ad_valorem_usd + pre_liq.fodinfa_usd + pre_liq.iva_usd
            pre_liq.is_confirmed = True
            pre_liq.confirmed_by = request.user
            pre_liq.confirmed_at = timezone.now()
            pre_liq.save()
            
            return Response({
                'message': 'Código HS confirmado con tarifas estimadas (no existe tarifa específica)',
                'hs_code': hs_code,
                'nota': 'Se usaron tarifas estimadas: Ad Valorem 10%, FODINFA 0.5%, IVA 15% (2025)',
                'pre_liquidacion': PreLiquidationSerializer(pre_liq).data
            })
    
    @action(detail=False, methods=['post'], url_path='sugerir-hs')
    def sugerir_hs(self, request):
        """AI endpoint for HS code suggestion using Gemini AI"""
        from .gemini_service import suggest_hs_code
        
        serializer = HSCodeSuggestionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        description = serializer.validated_data['product_description']
        origin_country = request.data.get('origin_country', '')
        try:
            fob_value = float(request.data.get('fob_value', 0) or 0)
        except (ValueError, TypeError):
            fob_value = 0
        
        result = suggest_hs_code(
            product_description=description,
            origin_country=origin_country,
            fob_value=fob_value
        )
        
        ai_status = result.get('ai_status', 'unknown')
        
        return Response({
            'suggested_hs_code': result.get('suggested_hs_code', '9999.00.00'),
            'confidence': result.get('confidence', 30),
            'category': result.get('category', ''),
            'reasoning': result.get('reasoning', 'Clasificacion pendiente'),
            'notes': result.get('notes', ''),
            'ai_powered': ai_status == 'success',
            'ai_status': ai_status,
            'ad_valorem_rate': result.get('ad_valorem_rate', 0.10),
            'requires_permit': result.get('requires_permit', False),
            'permit_info': result.get('permit_info', None),
            'special_taxes': result.get('special_taxes', []),
            'tributos_2025': result.get('tributos_2025', {
                'iva_rate': 0.15,
                'fodinfa_rate': 0.005,
                'ad_valorem_rate': 0.10
            })
        })
    
    @action(detail=True, methods=['post'], url_path='editar-hs')
    def editar_hs(self, request, pk=None):
        """Update the suggested HS code before confirmation"""
        pre_liq = self.get_object()
        
        if pre_liq.is_confirmed:
            return Response({
                'error': 'No se puede editar el código HS después de confirmar la pre-liquidación'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = PreLiquidationUpdateHSCodeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        hs_code = serializer.validated_data['hs_code']
        notes = serializer.validated_data.get('notes', '')
        
        pre_liq.suggested_hs_code = hs_code
        if notes:
            pre_liq.ai_reasoning = f"{pre_liq.ai_reasoning} | Editado por usuario: {notes}"
        pre_liq.ai_status = 'user_modified'
        
        self._calculate_estimated_duties(pre_liq)
        pre_liq.save()
        
        return Response({
            'success': True,
            'message': f'Partida arancelaria actualizada a {hs_code}',
            'pre_liquidacion': PreLiquidationSerializer(pre_liq).data
        })
    
    @action(detail=True, methods=['post'], url_path='solicitar-asistencia')
    def solicitar_asistencia(self, request, pk=None):
        """Request customs assistance - sends email to Freight Forwarder"""
        pre_liq = self.get_object()
        
        if pre_liq.assistance_requested:
            return Response({
                'error': 'Ya se ha solicitado asistencia para esta pre-liquidación',
                'assistance_status': pre_liq.assistance_status
            }, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = RequestAssistanceSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        notes = serializer.validated_data.get('notes', '')
        
        pre_liq.assistance_requested = True
        pre_liq.assistance_requested_at = timezone.now()
        pre_liq.assistance_notes = notes
        pre_liq.assistance_status = 'pending'
        pre_liq.save()
        
        try:
            user = request.user
            cotizacion = pre_liq.cotizacion
            
            subject = f"Solicitud de Asistencia Aduanera - Pre-Liquidación {cotizacion.numero_cotizacion}"
            message = f"""
Se ha recibido una solicitud de asistencia aduanera:

Cliente: {user.get_full_name()} ({user.email})
Empresa: {user.company_name or 'N/A'}
Cotización: {cotizacion.numero_cotizacion}

Descripción del Producto: {pre_liq.product_description}
Código HS Sugerido: {pre_liq.suggested_hs_code}
Valor CIF: ${pre_liq.cif_value_usd}
Total Tributos Estimados: ${pre_liq.total_tributos_usd}

Notas del Cliente:
{notes or 'Sin notas adicionales'}

Por favor, contacte al cliente para brindar asistencia con la clasificación arancelaria.

---
Sistema ImportaYa.ia
"""
            
            ff_email = getattr(settings, 'FREIGHT_FORWARDER_EMAIL', 'operaciones@integralcargosolutions.ec')
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [ff_email],
                fail_silently=True,
            )
        except Exception as e:
            logger.error(f"Error sending assistance request email: {str(e)}")
        
        return Response({
            'success': True,
            'message': 'Solicitud de asistencia enviada exitosamente. Un asesor se comunicará contigo pronto.',
            'pre_liquidacion': PreLiquidationSerializer(pre_liq).data
        })
    
    @action(detail=True, methods=['get', 'post'], url_path='documentos', parser_classes=[MultiPartParser, FormParser])
    def documentos(self, request, pk=None):
        """Manage pre-liquidation documents with file upload support"""
        pre_liq = self.get_object()
        
        if request.method == 'GET':
            documents = pre_liq.documents.all()
            return Response({
                'success': True,
                'documents': PreLiquidationDocumentSerializer(documents, many=True).data
            })
        
        uploaded_file = request.FILES.get('file')
        document_type = request.data.get('document_type', 'otro')
        notes = request.data.get('notes', '')
        
        if uploaded_file:
            import os
            from django.core.files.storage import default_storage
            
            max_size = 10 * 1024 * 1024
            if uploaded_file.size > max_size:
                return Response({
                    'error': 'El archivo es demasiado grande. Máximo permitido: 10MB'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            allowed_types = [
                'application/pdf',
                'image/jpeg', 'image/png', 'image/gif',
                'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/vnd.ms-excel',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ]
            if uploaded_file.content_type not in allowed_types:
                return Response({
                    'error': f'Tipo de archivo no permitido: {uploaded_file.content_type}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            upload_dir = f'pre_liquidation_docs/{pre_liq.id}/'
            file_path = default_storage.save(
                os.path.join(upload_dir, uploaded_file.name),
                uploaded_file
            )
            
            document = PreLiquidationDocument.objects.create(
                pre_liquidation=pre_liq,
                document_type=document_type,
                file_name=uploaded_file.name,
                file_path=file_path,
                file_size=uploaded_file.size,
                mime_type=uploaded_file.content_type,
                notes=notes,
                uploaded_by=request.user
            )
        else:
            file_name = request.data.get('file_name', '')
            file_path = request.data.get('file_path', '')
            file_size = request.data.get('file_size', 0)
            mime_type = request.data.get('mime_type', '')
            
            if not file_name or not file_path:
                return Response({
                    'error': 'Se requiere un archivo o los campos file_name y file_path'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            document = PreLiquidationDocument.objects.create(
                pre_liquidation=pre_liq,
                document_type=document_type,
                file_name=file_name,
                file_path=file_path,
                file_size=file_size,
                mime_type=mime_type,
                notes=notes,
                uploaded_by=request.user
            )
        
        return Response({
            'success': True,
            'message': 'Documento registrado exitosamente',
            'document': PreLiquidationDocumentSerializer(document).data
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['delete'], url_path='documentos/(?P<doc_id>[^/.]+)')
    def eliminar_documento(self, request, pk=None, doc_id=None):
        """Delete a pre-liquidation document"""
        pre_liq = self.get_object()
        
        try:
            document = pre_liq.documents.get(pk=doc_id)
            document.delete()
            return Response({
                'success': True,
                'message': 'Documento eliminado exitosamente'
            })
        except PreLiquidationDocument.DoesNotExist:
            return Response({
                'error': 'Documento no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)


class AIAssistantAPIView(APIView):
    """
    AI Assistant endpoint for ImportaYa.ia
    Handles text queries (Mode A) and document analysis (Mode B)
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        from .gemini_service import ai_assistant_chat
        
        message = request.data.get('message', '').strip()
        image_data = request.data.get('image_data')
        image_mime_type = request.data.get('image_mime_type')
        
        if not message and not image_data:
            return Response({
                'error': 'Se requiere un mensaje o una imagen para procesar'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not message:
            message = "Analiza este documento y extrae la informacion relevante para importacion a Ecuador."
        
        result = ai_assistant_chat(
            message=message,
            image_data=image_data,
            image_mime_type=image_mime_type
        )
        
        return Response(result)


class LogisticsProviderViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing logistics providers (navieras, consolidadores, aerolíneas)
    """
    from .serializers import LogisticsProviderSerializer
    from .models import LogisticsProvider
    
    queryset = LogisticsProvider.objects.all()
    serializer_class = LogisticsProviderSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = LogisticsProvider.objects.all()
        
        transport_type = self.request.query_params.get('transport_type')
        if transport_type:
            queryset = queryset.filter(transport_type=transport_type.upper())
        
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('transport_type', 'priority', 'name')
    
    @action(detail=False, methods=['get'], url_path='by-type')
    def by_type(self, request):
        """Get providers grouped by transport type"""
        from .models import LogisticsProvider
        
        result = {}
        for transport_type in ['FCL', 'LCL', 'AEREO']:
            providers = LogisticsProvider.objects.filter(
                transport_type=transport_type,
                is_active=True
            ).values('id', 'name', 'code', 'priority')
            result[transport_type] = list(providers)
        
        return Response(result)
    
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Get summary counts by transport type"""
        from .models import LogisticsProvider
        from django.db.models import Count
        
        summary = LogisticsProvider.objects.filter(is_active=True).values(
            'transport_type'
        ).annotate(count=Count('id')).order_by('transport_type')
        
        return Response({
            'total': LogisticsProvider.objects.filter(is_active=True).count(),
            'by_type': list(summary)
        })


class ProviderRateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing provider rates/tarifas
    """
    from .serializers import ProviderRateSerializer
    from .models import ProviderRate
    
    queryset = ProviderRate.objects.all()
    serializer_class = ProviderRateSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        from django.utils import timezone
        
        queryset = ProviderRate.objects.select_related('provider').all()
        
        provider_id = self.request.query_params.get('provider')
        if provider_id:
            queryset = queryset.filter(provider_id=provider_id)
        
        transport_type = self.request.query_params.get('transport_type')
        if transport_type:
            queryset = queryset.filter(provider__transport_type=transport_type.upper())
        
        origin = self.request.query_params.get('origin')
        if origin:
            queryset = queryset.filter(origin_port__icontains=origin)
        
        destination = self.request.query_params.get('destination')
        if destination:
            queryset = queryset.filter(destination=destination.upper())
        
        container_type = self.request.query_params.get('container_type')
        if container_type:
            queryset = queryset.filter(container_type=container_type.upper())
        
        valid_only = self.request.query_params.get('valid_only')
        if valid_only and valid_only.lower() == 'true':
            today = timezone.now().date()
            queryset = queryset.filter(
                valid_from__lte=today,
                valid_to__gte=today,
                is_active=True
            )
        
        return queryset.order_by('provider__name', 'origin_port', 'rate_usd')
    
    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        """
        Search for best rates based on route and transport type
        Used by Gemini AI for automatic quote generation
        """
        from django.utils import timezone
        from .models import ProviderRate
        
        transport_type = request.query_params.get('transport_type', '').upper()
        origin = request.query_params.get('origin', '')
        destination = request.query_params.get('destination', 'GYE')
        container_type = request.query_params.get('container_type', '')
        limit = int(request.query_params.get('limit', 5))
        
        if not transport_type:
            return Response({
                'error': 'Se requiere transport_type (FCL, LCL, AEREO)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        today = timezone.now().date()
        
        queryset = ProviderRate.objects.select_related('provider').filter(
            provider__transport_type=transport_type,
            provider__is_active=True,
            is_active=True,
            valid_from__lte=today,
            valid_to__gte=today
        )
        
        if origin:
            queryset = queryset.filter(origin_port__icontains=origin)
        
        if destination:
            queryset = queryset.filter(destination=destination.upper())
        
        if container_type and transport_type == 'FCL':
            queryset = queryset.filter(container_type=container_type.upper())
        
        rates = queryset.order_by('rate_usd')[:limit]
        
        result = []
        for rate in rates:
            result.append({
                'provider_id': rate.provider.id,
                'provider_name': rate.provider.name,
                'provider_code': rate.provider.code,
                'origin_port': rate.origin_port,
                'origin_country': rate.origin_country,
                'destination': rate.destination,
                'container_type': rate.container_type,
                'rate_usd': float(rate.rate_usd),
                'unit': rate.unit,
                'transit_days_min': rate.transit_days_min,
                'transit_days_max': rate.transit_days_max,
                'free_days': rate.free_days,
                'thc_origin_usd': float(rate.thc_origin_usd),
                'thc_destination_usd': float(rate.thc_destination_usd),
            })
        
        return Response({
            'transport_type': transport_type,
            'origin': origin,
            'destination': destination,
            'container_type': container_type,
            'rates_found': len(result),
            'rates': result
        })


class AirportRegionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for airport regions (read-only)
    """
    from .serializers import AirportRegionSerializer
    from .models import AirportRegion
    
    queryset = AirportRegion.objects.filter(is_active=True)
    serializer_class = AirportRegionSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        return AirportRegion.objects.filter(is_active=True).order_by('display_order')


class AirportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for world airports management.
    
    Business Logic:
    - Users search by ciudad_exacta (user-friendly city names)
    - System uses iata_code internally for rate lookups
    """
    from .serializers import AirportSerializer, AirportSearchSerializer
    from .models import Airport
    
    queryset = Airport.objects.all()
    serializer_class = AirportSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = Airport.objects.select_related('region').all()
        
        region = self.request.query_params.get('region')
        if region:
            queryset = queryset.filter(region_name__icontains=region)
        
        country = self.request.query_params.get('country')
        if country:
            queryset = queryset.filter(country__icontains=country)
        
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        cargo_only = self.request.query_params.get('cargo_only')
        if cargo_only and cargo_only.lower() == 'true':
            queryset = queryset.filter(is_cargo_capable=True)
        
        return queryset.order_by('region_name', 'country', 'ciudad_exacta')
    
    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        """
        Search airports by ciudad_exacta (primary search field for users).
        Also searches by airport name, IATA code, and country.
        
        Query params:
            q: Search query (required, min 2 chars)
            limit: Max results (default 10, max 50)
        
        Returns airports sorted by relevance.
        """
        from .serializers import AirportSearchSerializer
        from .models import Airport
        
        query = request.query_params.get('q', '').strip()
        limit = min(int(request.query_params.get('limit', 10)), 50)
        
        if len(query) < 2:
            return Response({
                'error': 'La búsqueda requiere al menos 2 caracteres',
                'results': []
            })
        
        airports = Airport.search_by_city(query, limit=limit)
        serializer = AirportSearchSerializer(airports, many=True)
        
        return Response({
            'query': query,
            'count': len(serializer.data),
            'results': serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='by-iata/(?P<iata_code>[A-Z]{3})')
    def by_iata(self, request, iata_code=None):
        """
        Get airport by IATA code (internal system lookup).
        Used for freight rate calculations.
        """
        from .serializers import AirportSerializer
        from .models import Airport
        
        airport = Airport.get_by_iata(iata_code)
        
        if not airport:
            return Response({
                'error': f'Aeropuerto con código IATA {iata_code} no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        
        serializer = AirportSerializer(airport)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='by-country')
    def by_country(self, request):
        """
        Get airports grouped by country.
        """
        from .models import Airport
        from django.db.models import Count
        
        region = request.query_params.get('region')
        
        queryset = Airport.objects.filter(is_active=True, is_cargo_capable=True)
        if region:
            queryset = queryset.filter(region_name__icontains=region)
        
        countries = queryset.values('country', 'region_name').annotate(
            airport_count=Count('id')
        ).order_by('region_name', 'country')
        
        result = {}
        for item in countries:
            region_name = item['region_name']
            if region_name not in result:
                result[region_name] = []
            result[region_name].append({
                'country': item['country'],
                'airport_count': item['airport_count']
            })
        
        return Response(result)
    
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Get summary statistics of airports in database.
        """
        from .models import Airport
        from django.db.models import Count
        
        total = Airport.objects.filter(is_active=True).count()
        cargo_capable = Airport.objects.filter(is_active=True, is_cargo_capable=True).count()
        
        by_region = Airport.objects.filter(is_active=True).values(
            'region_name'
        ).annotate(count=Count('id')).order_by('region_name')
        
        countries = Airport.objects.filter(is_active=True).values('country').distinct().count()
        
        return Response({
            'total_airports': total,
            'cargo_capable': cargo_capable,
            'countries': countries,
            'by_region': list(by_region)
        })


class ContainerViewSet(viewsets.ViewSet):
    """
    API endpoints for container selection and multi-container FCL quotes.
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'], url_path='available')
    def available_containers(self, request):
        """
        Get all available container types for FCL selection.
        Returns automatic (instant quote) and manual (special equipment) containers.
        """
        from .container_logic import get_contenedores_disponibles
        return Response(get_contenedores_disponibles())
    
    @action(detail=False, methods=['post'], url_path='validate-selection')
    def validate_selection(self, request):
        """
        Validate a multi-container selection.
        
        Request body:
        {
            "selecciones": [
                {"tipo": "20GP", "cantidad": 2},
                {"tipo": "40HC", "cantidad": 1}
            ]
        }
        """
        from .container_logic import validar_seleccion_multicontenedor_json
        
        selecciones = request.data.get('selecciones', [])
        resultado = validar_seleccion_multicontenedor_json(selecciones)
        
        return Response(resultado)
    
    @action(detail=False, methods=['post'], url_path='calculate-capacity')
    def calculate_capacity(self, request):
        """
        Calculate total capacity for a container selection.
        """
        from .container_logic import calcular_capacidad_seleccion
        
        selecciones = request.data.get('selecciones', [])
        resultado = calcular_capacidad_seleccion(selecciones)
        
        return Response(resultado)
    
    @action(detail=False, methods=['post'], url_path='optimize')
    def optimize(self, request):
        """
        Get optimized container recommendation based on cargo dimensions.
        """
        from .container_logic import optimizar_contenedor_json
        
        volumen = request.data.get('volumen_cbm', 0)
        peso = request.data.get('peso_kg', 0)
        
        try:
            resultado = optimizar_contenedor_json(volumen, peso)
            return Response(resultado)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ManualQuoteRequestViewSet(OwnerFilterMixin, viewsets.ModelViewSet):
    """
    API endpoints for manual quote requests (special containers).
    """
    from .models import ManualQuoteRequest
    from .serializers import ManualQuoteRequestSerializer
    
    queryset = ManualQuoteRequest.objects.all()
    serializer_class = ManualQuoteRequestSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'lead']
    search_fields = ['company_name', 'request_number']
    
    @action(detail=False, methods=['get'], url_path='pending')
    def pending_requests(self, request):
        """Get all pending manual quote requests for admin."""
        pending = self.get_queryset().filter(status='pendiente').order_by('-created_at')
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='urgent')
    def urgent_requests(self, request):
        """Get urgent requests (pending > 2 days)."""
        from django.utils import timezone
        from datetime import timedelta
        
        cutoff = timezone.now() - timedelta(days=2)
        urgent = self.get_queryset().filter(
            status='pendiente',
            created_at__lt=cutoff
        ).order_by('created_at')
        
        serializer = self.get_serializer(urgent, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='submit-quote')
    def submit_quote(self, request, pk=None):
        """
        Admin submits the manual quote with costs.
        """
        obj = self.get_object()
        
        obj.cost_freight_usd = request.data.get('cost_freight_usd')
        obj.cost_local_charges_usd = request.data.get('cost_local_charges_usd')
        obj.cost_special_equipment_usd = request.data.get('cost_special_equipment_usd', 0)
        obj.profit_margin_usd = request.data.get('profit_margin_usd', 150)
        obj.valid_until = request.data.get('valid_until')
        obj.transit_time_days = request.data.get('transit_time_days')
        obj.provider_name = request.data.get('provider_name', '')
        obj.customer_message = request.data.get('customer_message', '')
        obj.status = 'cotizacion_lista'
        obj.quoted_at = timezone.now()
        obj.save()
        
        return Response({'status': 'Cotización lista', 'final_price_usd': float(obj.final_price_usd or 0)})


class PortViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoints for world ports (read-only).
    
    Endpoints:
    - GET /api/sales/ports/ - Lista todos los puertos
    - GET /api/sales/ports/{id}/ - Detalle de puerto
    - GET /api/sales/ports/search/?q=shanghai - Búsqueda por nombre/país/código
    - GET /api/sales/ports/by-region/?region=Asia - Puertos por región
    - GET /api/sales/ports/by-locode/{LOCODE}/ - Búsqueda por UN/LOCODE
    - GET /api/sales/ports/summary/ - Resumen estadístico
    """
    queryset = Port.objects.filter(is_active=True)
    serializer_class = PortSerializer
    permission_classes = [AllowAny]
    filterset_fields = ['region', 'country']
    search_fields = ['name', 'country', 'un_locode']
    
    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        """
        Búsqueda de puertos por nombre, país o código.
        
        Query params:
        - q: texto de búsqueda (mínimo 2 caracteres)
        - limit: número máximo de resultados (default 20)
        """
        query = request.query_params.get('q', '').strip()
        limit = int(request.query_params.get('limit', 20))
        
        if len(query) < 2:
            return Response({'error': 'Búsqueda requiere mínimo 2 caracteres'}, status=status.HTTP_400_BAD_REQUEST)
        
        ports = Port.search_by_name(query, limit=limit)
        serializer = PortSearchSerializer(ports, many=True)
        
        return Response({
            'count': len(serializer.data),
            'query': query,
            'results': serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='by-region')
    def by_region(self, request):
        """
        Obtiene puertos filtrados por región.
        
        Query params:
        - region: nombre de la región (Norteamérica, Latinoamérica, Europa, África, Asia, Oceanía)
        """
        region = request.query_params.get('region', '').strip()
        
        valid_regions = ['Norteamérica', 'Latinoamérica', 'Europa', 'África', 'Asia', 'Oceanía']
        if region not in valid_regions:
            return Response({
                'error': f'Región inválida. Opciones: {", ".join(valid_regions)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        ports = Port.get_by_region(region)
        serializer = PortSerializer(ports, many=True)
        
        return Response({
            'region': region,
            'count': len(serializer.data),
            'ports': serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='by-locode/(?P<locode>[A-Z]{5})')
    def by_locode(self, request, locode=None):
        """
        Obtiene un puerto por su código UN/LOCODE.
        """
        try:
            port = Port.objects.get(un_locode=locode.upper(), is_active=True)
            serializer = PortSerializer(port)
            return Response(serializer.data)
        except Port.DoesNotExist:
            return Response({'error': f'Puerto con código {locode} no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Resumen estadístico de la base de datos de puertos.
        """
        total = Port.objects.filter(is_active=True).count()
        
        by_region = {}
        for region in ['Norteamérica', 'Latinoamérica', 'Europa', 'África', 'Asia', 'Oceanía']:
            by_region[region] = Port.objects.filter(region=region, is_active=True).count()
        
        countries = Port.objects.filter(is_active=True).values_list('country', flat=True).distinct().count()
        
        return Response({
            'total_ports': total,
            'total_countries': countries,
            'by_region': by_region
        })
    
    @action(detail=False, methods=['get'], url_path='by-country')
    def by_country(self, request):
        """
        Obtiene puertos agrupados por país.
        """
        from django.db.models import Count
        
        country = request.query_params.get('country', '').strip()
        
        if country:
            ports = Port.objects.filter(country__icontains=country, is_active=True)
            serializer = PortSerializer(ports, many=True)
            return Response({
                'country': country,
                'count': len(serializer.data),
                'ports': serializer.data
            })
        
        countries = Port.objects.filter(is_active=True).values('country', 'region').annotate(
            count=Count('id')
        ).order_by('region', 'country')
        
        return Response({
            'countries': list(countries)
        })


class ShippingInstructionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Shipping Instructions (Instrucciones de Embarque).
    Flujo: Cotización aprobada → Upload documentos → AI extraction → Formulario → RO → Notificación
    """
    queryset = ShippingInstruction.objects.all()
    serializer_class = ShippingInstructionSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        """Filtra por usuario autenticado (owner de la cotización)"""
        user = self.request.user
        return ShippingInstruction.objects.filter(
            quote_submission__owner=user
        ).select_related('quote_submission').prefetch_related('documents')
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ShippingInstructionDetailSerializer
        elif self.action == 'create' or self.action == 'init_from_quote':
            return ShippingInstructionCreateSerializer
        elif self.action in ['form', 'partial_update']:
            return ShippingInstructionFormSerializer
        return ShippingInstructionSerializer
    
    @action(detail=False, methods=['post'], url_path='init')
    def init_from_quote(self, request):
        """
        POST /api/sales/shipping-instructions/init/
        Inicializa un Shipping Instruction desde una cotización aprobada.
        Acepta quote_submission_id o lead_cotizacion_id.
        """
        quote_submission_id = request.data.get('quote_submission_id')
        lead_cotizacion_id = request.data.get('lead_cotizacion_id')
        
        if not quote_submission_id and not lead_cotizacion_id:
            return Response(
                {'error': 'Se requiere quote_submission_id o lead_cotizacion_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if lead_cotizacion_id:
                from .models import LeadCotizacion
                lead_cotizacion = LeadCotizacion.objects.get(
                    id=lead_cotizacion_id,
                    lead_user=request.user
                )
                quote_submission = lead_cotizacion.quote_submission
                if not quote_submission:
                    return Response(
                        {'error': 'La cotización no tiene un QuoteSubmission asociado'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                quote_submission = QuoteSubmission.objects.get(
                    id=quote_submission_id,
                    owner=request.user
                )
        except (QuoteSubmission.DoesNotExist, LeadCotizacion.DoesNotExist):
            return Response(
                {'error': 'Cotización no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Verify the quote submission is in an approved status before allowing shipping instructions
        if quote_submission.status not in ['aprobada', 'ro_generado', 'en_transito']:
            return Response(
                {'error': 'La cotización debe estar aprobada para enviar instrucciones de embarque'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        existing = ShippingInstruction.objects.filter(quote_submission=quote_submission).first()
        if existing:
            return Response(
                ShippingInstructionDetailSerializer(existing).data,
                status=status.HTTP_200_OK
            )
        
        shipping_instruction = ShippingInstruction.objects.create(
            quote_submission=quote_submission,
            shipper_name=quote_submission.company_name or '',
            shipper_address=quote_submission.city or '',
            shipper_contact=quote_submission.contact_name or '',
            shipper_email=quote_submission.contact_email or '',
            shipper_phone=quote_submission.contact_phone or '',
            consignee_name=quote_submission.company_name or '',
            consignee_address=quote_submission.city or '',
            origin_port=quote_submission.origin or '',
            destination_port=quote_submission.destination or '',
            cargo_description=quote_submission.cargo_description or quote_submission.product_description or '',
            gross_weight_kg=quote_submission.cargo_weight_kg,
            volume_cbm=quote_submission.cargo_volume_cbm,
            incoterm=quote_submission.incoterm or 'FOB',
            status='draft'
        )
        
        return Response(
            ShippingInstructionDetailSerializer(shipping_instruction).data,
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=True, methods=['get', 'post'], url_path='documents')
    def documents(self, request, pk=None):
        """
        GET/POST /api/sales/shipping-instructions/{id}/documents/
        Lista o sube documentos para el Shipping Instruction.
        """
        shipping_instruction = self.get_object()
        
        if request.method == 'GET':
            documents = shipping_instruction.documents.all()
            serializer = ShippingInstructionDocumentSerializer(documents, many=True)
            return Response(serializer.data)
        
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'Se requiere un archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        document_type = request.data.get('document_type', 'other')
        
        document = ShippingInstructionDocument.objects.create(
            shipping_instruction=shipping_instruction,
            document_type=document_type,
            file=file,
            original_filename=file.name,
            uploaded_by=request.user
        )
        
        shipping_instruction.status = 'documents_uploaded'
        shipping_instruction.save(update_fields=['status'])
        
        return Response(
            ShippingInstructionDocumentSerializer(document).data,
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=True, methods=['post'], url_path='process-ai')
    def process_ai(self, request, pk=None):
        """
        POST /api/sales/shipping-instructions/{id}/process-ai/
        Procesa documentos con Gemini AI para extraer datos.
        """
        shipping_instruction = self.get_object()
        
        documents = shipping_instruction.documents.filter(ai_processed=False)
        if not documents.exists():
            return Response(
                {'message': 'No hay documentos pendientes de procesar'},
                status=status.HTTP_200_OK
            )
        
        try:
            from .gemini_service import extract_shipping_data_from_documents
            
            extracted_data = extract_shipping_data_from_documents(
                shipping_instruction_id=shipping_instruction.id
            )
            
            if extracted_data:
                for field, value in extracted_data.items():
                    if hasattr(shipping_instruction, field) and value:
                        current_value = getattr(shipping_instruction, field)
                        if not current_value:
                            setattr(shipping_instruction, field, value)
                
                shipping_instruction.ai_extracted_data = extracted_data
                shipping_instruction.status = 'ai_processed'
                shipping_instruction.save()
                
                documents.update(ai_processed=True)
            
            return Response({
                'message': 'Documentos procesados exitosamente',
                'extracted_data': extracted_data,
                'shipping_instruction': ShippingInstructionDetailSerializer(shipping_instruction).data
            })
            
        except Exception as e:
            logger.error(f"Error processing AI for SI {pk}: {str(e)}")
            return Response(
                {'error': f'Error procesando con AI: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get', 'patch'], url_path='form')
    def form(self, request, pk=None):
        """
        GET/PATCH /api/sales/shipping-instructions/{id}/form/
        Obtiene o actualiza el formulario híbrido del SI.
        """
        shipping_instruction = self.get_object()
        
        if request.method == 'GET':
            serializer = ShippingInstructionFormSerializer(shipping_instruction)
            return Response({
                'form_data': serializer.data,
                'ai_suggestions': shipping_instruction.ai_extracted_data or {},
                'status': shipping_instruction.status
            })
        
        serializer = ShippingInstructionFormSerializer(
            shipping_instruction,
            data=request.data,
            partial=True
        )
        
        if serializer.is_valid():
            serializer.save()
            
            if shipping_instruction.status in ['draft', 'documents_uploaded', 'ai_processed']:
                shipping_instruction.status = 'form_in_progress'
                shipping_instruction.save(update_fields=['status'])
            
            return Response({
                'message': 'Formulario actualizado',
                'form_data': serializer.data
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'], url_path='finalize')
    def finalize(self, request, pk=None):
        """
        POST /api/sales/shipping-instructions/{id}/finalize/
        Finaliza el Shipping Instruction tras validar campos requeridos.
        """
        shipping_instruction = self.get_object()
        
        required_fields = [
            ('shipper_name', 'Nombre del Shipper'),
            ('consignee_name', 'Nombre del Consignee'),
            ('origin_port', 'Puerto de Origen'),
            ('destination_port', 'Puerto de Destino'),
            ('cargo_description', 'Descripción de la Carga'),
        ]
        
        missing_fields = []
        for field, label in required_fields:
            if not getattr(shipping_instruction, field):
                missing_fields.append(label)
        
        if missing_fields:
            return Response({
                'error': 'Faltan campos requeridos',
                'missing_fields': missing_fields
            }, status=status.HTTP_400_BAD_REQUEST)
        
        shipping_instruction.status = 'finalized'
        shipping_instruction.finalized_at = timezone.now()
        shipping_instruction.save(update_fields=['status', 'finalized_at'])
        
        return Response({
            'message': 'Shipping Instruction finalizado',
            'shipping_instruction': ShippingInstructionDetailSerializer(shipping_instruction).data
        })
    
    @action(detail=True, methods=['post'], url_path='generate-ro')
    def generate_ro(self, request, pk=None):
        """
        POST /api/sales/shipping-instructions/{id}/generate-ro/
        Genera el número de Routing Order (RO).
        """
        shipping_instruction = self.get_object()
        
        if shipping_instruction.status != 'finalized':
            return Response({
                'error': 'El SI debe estar finalizado para generar RO'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if shipping_instruction.ro_number:
            return Response({
                'message': 'RO ya generado',
                'ro_number': shipping_instruction.ro_number
            })
        
        year = timezone.now().year
        month = timezone.now().month
        
        count = ShippingInstruction.objects.filter(
            ro_number__isnull=False,
            created_at__year=year,
            created_at__month=month
        ).count() + 1
        
        ro_number = f"RO-{year}{month:02d}-{count:04d}"
        
        shipping_instruction.ro_number = ro_number
        shipping_instruction.ro_generated_at = timezone.now()
        shipping_instruction.status = 'ro_generated'
        shipping_instruction.save(update_fields=['ro_number', 'ro_generated_at', 'status'])
        
        return Response({
            'message': 'Routing Order generado exitosamente',
            'ro_number': ro_number,
            'shipping_instruction': ShippingInstructionDetailSerializer(shipping_instruction).data
        })
    
    @action(detail=True, methods=['post'], url_path='notify-forwarder')
    def notify_forwarder(self, request, pk=None):
        """
        POST /api/sales/shipping-instructions/{id}/notify-forwarder/
        Envía notificación email al Freight Forwarder con el SI.
        """
        shipping_instruction = self.get_object()
        
        if not shipping_instruction.ro_number:
            return Response({
                'error': 'Se requiere generar RO antes de notificar'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            quote_submission = shipping_instruction.quote_submission
            
            subject = f"Nueva Instrucción de Embarque - RO: {shipping_instruction.ro_number}"
            message = f"""
Estimado Freight Forwarder,

Se ha recibido una nueva Instrucción de Embarque lista para procesamiento.

ROUTING ORDER: {shipping_instruction.ro_number}
SOLICITUD: {quote_submission.submission_number}

SHIPPER:
- Nombre: {shipping_instruction.shipper_name}
- Dirección: {shipping_instruction.shipper_address}
- Contacto: {shipping_instruction.shipper_contact}
- Email: {shipping_instruction.shipper_email}
- Teléfono: {shipping_instruction.shipper_phone}

CONSIGNEE:
- Nombre: {shipping_instruction.consignee_name}
- Dirección: {shipping_instruction.consignee_address}

NOTIFICANTE:
- Nombre: {shipping_instruction.notify_party_name or 'Mismo que consignee'}

DETALLES DEL EMBARQUE:
- Origen: {shipping_instruction.origin_port}
- Destino: {shipping_instruction.destination_port}
- Incoterm: {shipping_instruction.incoterm}
- Descripción: {shipping_instruction.cargo_description}
- Peso Bruto: {shipping_instruction.gross_weight_kg} kg
- Volumen: {shipping_instruction.volume_cbm} CBM
- Número de Bultos: {shipping_instruction.package_count or 'N/A'}
- Tipo de Embalaje: {shipping_instruction.package_type or 'N/A'}

INSTRUCCIONES ESPECIALES:
{shipping_instruction.special_instructions or 'Ninguna'}

Por favor confirme recepción y proporcione la referencia de booking.

Sistema ImportaYa.ia
La logística de carga integral, ahora es Inteligente!
"""
            
            forwarder_email = getattr(settings, 'FREIGHT_FORWARDER_EMAIL', 'forwarder@importaya.ia')
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [forwarder_email],
                fail_silently=False,
            )
            
            shipping_instruction.forwarder_notified_at = timezone.now()
            shipping_instruction.status = 'sent_to_forwarder'
            shipping_instruction.save(update_fields=['forwarder_notified_at', 'status'])
            
            logger.info(f"Forwarder notification sent for SI {pk} - RO: {shipping_instruction.ro_number}")
            
            return Response({
                'message': 'Notificación enviada al Freight Forwarder',
                'ro_number': shipping_instruction.ro_number,
                'notified_at': shipping_instruction.forwarder_notified_at
            })
            
        except Exception as e:
            logger.error(f"Error sending forwarder notification: {str(e)}")
            return Response({
                'error': f'Error enviando notificación: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['put'], url_path='forwarder-reference')
    def forwarder_reference(self, request, pk=None):
        """
        PUT /api/sales/shipping-instructions/{id}/forwarder-reference/
        Guarda la referencia externa del Freight Forwarder.
        """
        shipping_instruction = self.get_object()
        
        reference = request.data.get('forwarder_reference')
        if not reference:
            return Response({
                'error': 'Se requiere forwarder_reference'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        shipping_instruction.forwarder_reference = reference
        shipping_instruction.status = 'confirmed'
        shipping_instruction.save(update_fields=['forwarder_reference', 'status'])
        
        return Response({
            'message': 'Referencia del forwarder guardada',
            'forwarder_reference': reference,
            'status': shipping_instruction.status
        })


class CargoTrackingViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para el módulo unificado de Cargo Tracking y sus Instrucciones de Embarque.
    Muestra ROs con SI confirmada y su timeline de 14 hitos.
    """
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        confirmed_statuses = ['ro_generated', 'sent_to_forwarder', 'confirmed', 'in_transit', 'delivered']
        
        if hasattr(user, 'lead_profile'):
            return ShippingInstruction.objects.filter(
                quote_submission__lead__email=user.email,
                status__in=confirmed_statuses
            ).select_related('quote_submission', 'quote_submission__lead').prefetch_related('milestones')
        
        return ShippingInstruction.objects.filter(
            status__in=confirmed_statuses
        ).select_related('quote_submission', 'quote_submission__lead').prefetch_related('milestones')
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return CargoTrackingDetailSerializer
        return CargoTrackingListSerializer
    
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        
        if not instance.milestones.exists():
            ShipmentMilestone.create_initial_milestones(instance)
        
        ShipmentMilestone.get_current_milestone(instance)
        
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='milestones')
    def milestones(self, request, pk=None):
        """
        GET /api/sales/cargo-tracking/{id}/milestones/
        Retorna todos los hitos del embarque ordenados.
        """
        shipping_instruction = self.get_object()
        
        if not shipping_instruction.milestones.exists():
            ShipmentMilestone.create_initial_milestones(shipping_instruction)
        
        milestones = shipping_instruction.milestones.all().order_by('milestone_order')
        serializer = ShipmentMilestoneSerializer(milestones, many=True)
        
        current = ShipmentMilestone.get_current_milestone(shipping_instruction)
        
        return Response({
            'ro_number': shipping_instruction.ro_number,
            'current_milestone': current.milestone_key if current else None,
            'current_milestone_order': current.milestone_order if current else 0,
            'milestones': serializer.data
        })
    
    @action(detail=True, methods=['put'], url_path='update-milestone')
    def update_milestone(self, request, pk=None):
        """
        PUT /api/sales/cargo-tracking/{id}/update-milestone/
        Actualiza un hito específico.
        Body: { milestone_key, actual_date, planned_date, meta_data, notes }
        """
        shipping_instruction = self.get_object()
        milestone_key = request.data.get('milestone_key')
        
        if not milestone_key:
            return Response({'error': 'Se requiere milestone_key'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            milestone = ShipmentMilestone.objects.get(
                shipping_instruction=shipping_instruction,
                milestone_key=milestone_key
            )
        except ShipmentMilestone.DoesNotExist:
            return Response({'error': 'Hito no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        actual_date = request.data.get('actual_date')
        if actual_date:
            if isinstance(actual_date, str):
                try:
                    milestone.actual_date = datetime.fromisoformat(actual_date.replace('Z', '+00:00'))
                except:
                    milestone.actual_date = timezone.now()
            else:
                milestone.actual_date = actual_date
        
        planned_date = request.data.get('planned_date')
        if planned_date:
            if isinstance(planned_date, str):
                try:
                    milestone.planned_date = datetime.fromisoformat(planned_date.replace('Z', '+00:00'))
                except:
                    pass
            else:
                milestone.planned_date = planned_date
        
        if 'meta_data' in request.data:
            milestone.meta_data = request.data['meta_data']
        
        if 'notes' in request.data:
            milestone.notes = request.data['notes']
        
        milestone.save()
        
        ShipmentMilestone.get_current_milestone(shipping_instruction)
        
        return Response({
            'message': 'Hito actualizado exitosamente',
            'milestone': ShipmentMilestoneSerializer(milestone).data
        })


class ShipmentMilestoneViewSet(viewsets.ModelViewSet):
    """
    ViewSet para CRUD de hitos de embarque.
    """
    serializer_class = ShipmentMilestoneSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = ShipmentMilestone.objects.all()
        
        ro_id = self.request.query_params.get('ro_id')
        if ro_id:
            queryset = queryset.filter(shipping_instruction_id=ro_id)
        
        ro_number = self.request.query_params.get('ro_number')
        if ro_number:
            queryset = queryset.filter(shipping_instruction__ro_number=ro_number)
        
        return queryset.select_related('shipping_instruction').order_by('milestone_order')
