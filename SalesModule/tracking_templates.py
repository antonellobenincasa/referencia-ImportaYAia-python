"""
Cargo Tracking Template Service for ImportaYa.ia
Handles Excel template generation and parsing for freight forwarder milestone updates.
"""
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils import timezone
from .models import ShippingInstruction, ShipmentMilestone
from .notification_service import NotificationService
import logging

logger = logging.getLogger(__name__)


class TrackingTemplateService:
    """Service for generating and parsing cargo tracking templates."""
    
    MILESTONE_COLUMNS = [
        ('A', 'RO Number', 'ro_number'),
        ('B', 'Consignatario', 'consignee'),
        ('C', 'POL', 'pol'),
        ('D', 'POD', 'pod'),
        ('E', 'Hito', 'milestone_key'),
        ('F', 'Nombre del Hito', 'milestone_label'),
        ('G', 'Estado', 'status'),
        ('H', 'Fecha Planificada', 'planned_date'),
        ('I', 'Fecha Real', 'actual_date'),
        ('J', 'Notas', 'notes'),
        ('K', 'Booking Number', 'booking_number'),
        ('L', 'Container Number', 'container_number'),
        ('M', 'BL Number', 'bl_number'),
        ('N', 'Naviera', 'carrier'),
    ]
    
    HEADER_FILL = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="00C9B7", end_color="00C9B7", fill_type="solid")
    SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    STATUS_MAP = {
        'PENDING': 'Pendiente',
        'IN_PROGRESS': 'En Progreso',
        'COMPLETED': 'Completado',
    }
    
    REVERSE_STATUS_MAP = {
        'pendiente': 'PENDING',
        'en progreso': 'IN_PROGRESS',
        'completado': 'COMPLETED',
    }
    
    @classmethod
    def generate_template(cls, shipping_instructions=None, include_data=False):
        """
        Generate an Excel template for cargo tracking.
        
        Args:
            shipping_instructions: QuerySet of ShippingInstruction objects
            include_data: If True, populate with existing milestone data
        
        Returns:
            BytesIO object containing the Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tracking Template"
        
        ws.merge_cells('A1:N1')
        ws['A1'] = "ImportaYa.ia - Plantilla de Actualización de Tracking"
        ws['A1'].font = Font(size=16, bold=True, color="0A2540")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells('A2:N2')
        ws['A2'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Zona Horaria: America/Guayaquil"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        for col_idx, (col_letter, header, _) in enumerate(cls.MILESTONE_COLUMNS, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cls.BORDER
        
        ws.row_dimensions[4].height = 25
        
        col_widths = [15, 25, 15, 15, 25, 35, 15, 18, 18, 40, 20, 20, 20, 20]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        row = 5
        
        if include_data and shipping_instructions:
            for si in shipping_instructions:
                milestones = si.milestones.all().order_by('milestone_order')
                
                for milestone in milestones:
                    quote = si.quote
                    
                    ws.cell(row=row, column=1, value=si.ro_number or '')
                    ws.cell(row=row, column=2, value=si.consignee_name or (quote.company_name if quote else ''))
                    ws.cell(row=row, column=3, value=getattr(quote, 'origin_port', '') if quote else '')
                    ws.cell(row=row, column=4, value=getattr(quote, 'destination_port', '') if quote else '')
                    ws.cell(row=row, column=5, value=milestone.milestone_key)
                    ws.cell(row=row, column=6, value=milestone.get_milestone_key_display())
                    ws.cell(row=row, column=7, value=cls.STATUS_MAP.get(milestone.status, milestone.status))
                    
                    if milestone.planned_date:
                        ws.cell(row=row, column=8, value=milestone.planned_date.strftime('%d/%m/%Y %H:%M'))
                    
                    if milestone.actual_date:
                        ws.cell(row=row, column=9, value=milestone.actual_date.strftime('%d/%m/%Y %H:%M'))
                    
                    ws.cell(row=row, column=10, value=milestone.notes or '')
                    
                    meta = milestone.meta_data or {}
                    ws.cell(row=row, column=11, value=meta.get('booking_number', si.booking_number or ''))
                    ws.cell(row=row, column=12, value=meta.get('container_number', ''))
                    ws.cell(row=row, column=13, value=meta.get('bl_number', si.bl_number or ''))
                    ws.cell(row=row, column=14, value=si.carrier or '')
                    
                    for col_idx in range(1, 15):
                        ws.cell(row=row, column=col_idx).border = cls.BORDER
                    
                    row += 1
        else:
            for i in range(10):
                for col_idx in range(1, 15):
                    cell = ws.cell(row=row + i, column=col_idx, value='')
                    cell.border = cls.BORDER
        
        instructions_ws = wb.create_sheet("Instrucciones")
        instructions = [
            ("INSTRUCCIONES PARA LLENAR LA PLANTILLA", ""),
            ("", ""),
            ("Columna", "Descripción"),
            ("RO Number", "Número de Routing Order (obligatorio) - No modificar"),
            ("Consignatario", "Nombre del consignatario"),
            ("POL", "Puerto de Origen"),
            ("POD", "Puerto de Destino"),
            ("Hito", "Código del hito (no modificar) - Ej: SI_ENVIADA, BOOKING_CONFIRMADO"),
            ("Nombre del Hito", "Descripción del hito (informativo)"),
            ("Estado", "Pendiente, En Progreso, o Completado"),
            ("Fecha Planificada", "Formato: DD/MM/AAAA HH:MM"),
            ("Fecha Real", "Formato: DD/MM/AAAA HH:MM - Al llenar, el hito se marca como completado"),
            ("Notas", "Observaciones adicionales"),
            ("Booking Number", "Número de booking de la naviera"),
            ("Container Number", "Número(s) de contenedor"),
            ("BL Number", "Número de Bill of Lading"),
            ("Naviera", "Nombre de la naviera"),
            ("", ""),
            ("CÓDIGOS DE HITOS VÁLIDOS:", ""),
            ("SI_ENVIADA", "SI Enviada"),
            ("SI_RECIBIDA_FORWARDER", "SI Recibida por Forwarder"),
            ("SHIPPER_CONTACTADO", "Shipper Contactado"),
            ("BOOKING_ORDER_RECIBIDA", "Booking Order Recibida en Origen"),
            ("ETD_SOLICITADO", "ETD Solicitado / Gestionando Reserva"),
            ("BOOKING_CONFIRMADO", "Booking Confirmado"),
            ("SO_LIBERADO", "S/O Liberado al Shipper"),
            ("RETIRO_CARGUE_VACIOS", "Fecha Retiro y Cargue de Vacíos"),
            ("CONTENEDORES_CARGADOS", "Contenedores Cargados"),
            ("GATE_IN", "Gate In (Entrada Terminal Origen)"),
            ("ETD", "ETD (Estimated Time of Departure)"),
            ("ATD", "ATD (Actual Time of Departure)"),
            ("TRANSHIPMENT", "Transhipment Date"),
            ("ETA_DESTINO", "ETA Destino Final"),
        ]
        
        for row_idx, (col1, col2) in enumerate(instructions, 1):
            instructions_ws.cell(row=row_idx, column=1, value=col1)
            instructions_ws.cell(row=row_idx, column=2, value=col2)
            if row_idx == 1:
                instructions_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
            elif row_idx == 3 or row_idx == 19:
                instructions_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                instructions_ws.cell(row=row_idx, column=2).font = Font(bold=True)
        
        instructions_ws.column_dimensions['A'].width = 30
        instructions_ws.column_dimensions['B'].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @classmethod
    def parse_template(cls, file_content, send_notifications=True):
        """
        Parse an uploaded Excel template and update milestones.
        
        Args:
            file_content: BytesIO or file-like object with Excel content
            send_notifications: If True, send notifications for updated milestones
        
        Returns:
            dict with results: {
                'success': bool,
                'updated': int,
                'errors': list,
                'warnings': list,
                'details': list
            }
        """
        results = {
            'success': True,
            'updated': 0,
            'errors': [],
            'warnings': [],
            'details': []
        }
        
        try:
            wb = openpyxl.load_workbook(file_content)
            ws = wb.active
            
            updated_milestones_by_si = {}
            
            for row_idx in range(5, ws.max_row + 1):
                ro_number = str(ws.cell(row=row_idx, column=1).value or '').strip()
                if not ro_number:
                    continue
                
                milestone_key = str(ws.cell(row=row_idx, column=5).value or '').strip().upper()
                if not milestone_key:
                    results['warnings'].append(f"Fila {row_idx}: Sin código de hito, omitida")
                    continue
                
                try:
                    si = ShippingInstruction.objects.get(ro_number=ro_number)
                except ShippingInstruction.DoesNotExist:
                    results['errors'].append(f"Fila {row_idx}: RO {ro_number} no encontrado")
                    continue
                
                try:
                    milestone = ShipmentMilestone.objects.get(
                        shipping_instruction=si,
                        milestone_key=milestone_key
                    )
                except ShipmentMilestone.DoesNotExist:
                    results['errors'].append(f"Fila {row_idx}: Hito {milestone_key} no encontrado para RO {ro_number}")
                    continue
                
                updated = False
                old_status = milestone.status
                
                status_raw = str(ws.cell(row=row_idx, column=7).value or '').strip().lower()
                if status_raw and status_raw in cls.REVERSE_STATUS_MAP:
                    new_status = cls.REVERSE_STATUS_MAP[status_raw]
                    if milestone.status != new_status:
                        milestone.status = new_status
                        updated = True
                
                planned_date_raw = ws.cell(row=row_idx, column=8).value
                if planned_date_raw:
                    try:
                        if isinstance(planned_date_raw, datetime):
                            milestone.planned_date = timezone.make_aware(planned_date_raw) if timezone.is_naive(planned_date_raw) else planned_date_raw
                            updated = True
                        elif isinstance(planned_date_raw, str):
                            parsed = datetime.strptime(planned_date_raw.strip(), '%d/%m/%Y %H:%M')
                            milestone.planned_date = timezone.make_aware(parsed)
                            updated = True
                    except Exception as e:
                        results['warnings'].append(f"Fila {row_idx}: Error parseando fecha planificada: {e}")
                
                actual_date_raw = ws.cell(row=row_idx, column=9).value
                if actual_date_raw:
                    try:
                        if isinstance(actual_date_raw, datetime):
                            milestone.actual_date = timezone.make_aware(actual_date_raw) if timezone.is_naive(actual_date_raw) else actual_date_raw
                            updated = True
                        elif isinstance(actual_date_raw, str):
                            parsed = datetime.strptime(actual_date_raw.strip(), '%d/%m/%Y %H:%M')
                            milestone.actual_date = timezone.make_aware(parsed)
                            updated = True
                    except Exception as e:
                        results['warnings'].append(f"Fila {row_idx}: Error parseando fecha real: {e}")
                
                notes = ws.cell(row=row_idx, column=10).value
                if notes and str(notes).strip():
                    milestone.notes = str(notes).strip()
                    updated = True
                
                meta = milestone.meta_data or {}
                meta_updated = False
                
                booking_number = ws.cell(row=row_idx, column=11).value
                if booking_number and str(booking_number).strip():
                    meta['booking_number'] = str(booking_number).strip()
                    meta_updated = True
                
                container_number = ws.cell(row=row_idx, column=12).value
                if container_number and str(container_number).strip():
                    meta['container_number'] = str(container_number).strip()
                    meta_updated = True
                
                bl_number = ws.cell(row=row_idx, column=13).value
                if bl_number and str(bl_number).strip():
                    meta['bl_number'] = str(bl_number).strip()
                    meta_updated = True
                
                if meta_updated:
                    milestone.meta_data = meta
                    updated = True
                
                if updated:
                    milestone.save()
                    results['updated'] += 1
                    results['details'].append(f"RO {ro_number} - {milestone.get_milestone_key_display()} actualizado")
                    
                    if ro_number not in updated_milestones_by_si:
                        updated_milestones_by_si[ro_number] = {
                            'si': si,
                            'milestones': [],
                            'old_statuses': {}
                        }
                    updated_milestones_by_si[ro_number]['milestones'].append(milestone)
                    updated_milestones_by_si[ro_number]['old_statuses'][milestone.id] = old_status
            
            if send_notifications:
                for ro_number, data in updated_milestones_by_si.items():
                    si = data['si']
                    milestones = data['milestones']
                    
                    if len(milestones) > 1:
                        NotificationService.send_bulk_milestone_update_notification(si, milestones)
                    else:
                        for milestone in milestones:
                            old_status = data['old_statuses'].get(milestone.id)
                            NotificationService.send_milestone_notification(milestone, old_status)
            
            if results['errors']:
                results['success'] = False
                
        except Exception as e:
            logger.error(f"Error parsing template: {e}")
            results['success'] = False
            results['errors'].append(f"Error al procesar el archivo: {str(e)}")
        
        return results
    
    @classmethod
    def generate_empty_template(cls):
        """Generate an empty template for new data entry."""
        return cls.generate_template(include_data=False)
    
    @classmethod
    def generate_template_for_ro(cls, ro_number):
        """Generate a template for a specific RO number."""
        try:
            si = ShippingInstruction.objects.get(ro_number=ro_number)
            return cls.generate_template(shipping_instructions=[si], include_data=True)
        except ShippingInstruction.DoesNotExist:
            return None
    
    @classmethod
    def generate_template_for_active_shipments(cls):
        """Generate a template with all active shipments (RO generated, not delivered)."""
        active_sis = ShippingInstruction.objects.filter(
            ro_number__isnull=False,
            status__in=['ro_generated', 'pending', 'in_progress']
        ).select_related('quote_submission').prefetch_related('milestones')
        
        return cls.generate_template(shipping_instructions=active_sis, include_data=True)
