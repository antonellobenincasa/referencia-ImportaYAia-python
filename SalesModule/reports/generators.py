from django.db.models import Count, Sum, Avg, Q, F, Min, Max
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from SalesModule.models import (
    Lead, Opportunity, Quote, TaskReminder, QuoteSubmission, CostRate,
    LeadCotizacion, QuoteScenario, QuoteLineItem, Shipment, PreLiquidation,
    FreightRate, InsuranceRate, CustomsDutyRate
)
from CommsModule.models import InboxMessage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io


def generate_sales_metrics_report(start_date=None, end_date=None):
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    total_leads = Lead.objects.filter(created_at__range=[start_date, end_date]).count()
    converted_leads = Lead.objects.filter(
        created_at__range=[start_date, end_date],
        status='convertido'
    ).count()
    
    total_quotes = Quote.objects.filter(created_at__range=[start_date, end_date]).count()
    quotes_sent = Quote.objects.filter(
        created_at__range=[start_date, end_date],
        status='enviado'
    ).count()
    quotes_accepted = Quote.objects.filter(
        created_at__range=[start_date, end_date],
        status='aceptado'
    ).count()
    
    total_quote_value = Quote.objects.filter(
        created_at__range=[start_date, end_date]
    ).aggregate(total=Sum('final_price'))['total'] or 0
    
    avg_quote_value = Quote.objects.filter(
        created_at__range=[start_date, end_date]
    ).aggregate(avg=Avg('final_price'))['avg'] or 0
    
    opportunities = Opportunity.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('stage').annotate(count=Count('id'))
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        },
        'leads': {
            'total': total_leads,
            'converted': converted_leads,
            'conversion_rate': round((converted_leads / total_leads * 100) if total_leads > 0 else 0, 2)
        },
        'quotes': {
            'total': total_quotes,
            'sent': quotes_sent,
            'accepted': quotes_accepted,
            'acceptance_rate': round((quotes_accepted / quotes_sent * 100) if quotes_sent > 0 else 0, 2),
            'total_value': float(total_quote_value),
            'average_value': float(avg_quote_value)
        },
        'opportunities_by_stage': list(opportunities)
    }


def generate_lead_conversion_report(start_date=None, end_date=None):
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    leads_by_status = Lead.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('status').annotate(count=Count('id'))
    
    leads_by_source = Lead.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('source').annotate(count=Count('id'))
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        },
        'leads_by_status': list(leads_by_status),
        'leads_by_source': list(leads_by_source)
    }


def generate_communication_report(start_date=None, end_date=None):
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    messages_by_source = InboxMessage.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('source').annotate(count=Count('id'))
    
    messages_by_status = InboxMessage.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('status').annotate(count=Count('id'))
    
    total_messages = InboxMessage.objects.filter(
        created_at__range=[start_date, end_date]
    ).count()
    
    responded_messages = InboxMessage.objects.filter(
        created_at__range=[start_date, end_date],
        status='respondido'
    ).count()
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        },
        'total_messages': total_messages,
        'responded_messages': responded_messages,
        'response_rate': round((responded_messages / total_messages * 100) if total_messages > 0 else 0, 2),
        'messages_by_source': list(messages_by_source),
        'messages_by_status': list(messages_by_status)
    }


def generate_quote_analytics_report(start_date=None, end_date=None):
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    quotes_by_status = Quote.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('status').annotate(count=Count('id'))
    
    quotes_by_incoterm = Quote.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('incoterm').annotate(count=Count('id'))
    
    quotes_by_cargo_type = Quote.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('cargo_type').annotate(count=Count('id'))
    
    avg_profit_margin = Quote.objects.filter(
        created_at__range=[start_date, end_date]
    ).aggregate(avg=Avg('profit_margin'))['avg'] or 0
    
    submissions_by_status = QuoteSubmission.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('status').annotate(count=Count('id'))
    
    submissions_by_transport = QuoteSubmission.objects.filter(
        created_at__range=[start_date, end_date]
    ).values('transport_type').annotate(count=Count('id'))
    
    total_submission_value = QuoteSubmission.objects.filter(
        created_at__range=[start_date, end_date],
        final_price__isnull=False
    ).aggregate(total=Sum('final_price'))['total'] or 0
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        },
        'quotes_by_status': list(quotes_by_status),
        'quotes_by_incoterm': list(quotes_by_incoterm),
        'quotes_by_cargo_type': list(quotes_by_cargo_type),
        'average_profit_margin': float(avg_profit_margin),
        'quote_submissions': {
            'total': QuoteSubmission.objects.filter(created_at__range=[start_date, end_date]).count(),
            'by_status': list(submissions_by_status),
            'by_transport_type': list(submissions_by_transport),
            'total_value': float(total_submission_value)
        }
    }


def generate_quote_submissions_report(start_date=None, end_date=None):
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    submissions = QuoteSubmission.objects.filter(
        created_at__range=[start_date, end_date]
    ).values_list(
        'id', 'company_name', 'contact_name', 'contact_email',
        'origin', 'destination', 'transport_type', 'status',
        'cost_rate', 'profit_markup', 'final_price', 'created_at'
    )
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        },
        'total_submissions': QuoteSubmission.objects.filter(created_at__range=[start_date, end_date]).count(),
        'submissions': [
            {
                'id': s[0],
                'empresa': s[1],
                'contacto': s[2],
                'email': s[3],
                'origen': s[4],
                'destino': s[5],
                'transporte': s[6],
                'estado': s[7],
                'costo': float(s[8]) if s[8] else 0,
                'margen': float(s[9]) if s[9] else 0,
                'precio_final': float(s[10]) if s[10] else 0,
                'fecha': s[11].strftime('%Y-%m-%d %H:%M')
            } for s in submissions
        ]
    }


def export_report_to_excel(report_data, report_type):
    wb = Workbook()
    ws = wb.active
    ws.title = report_type
    
    ws['A1'] = f'Reporte: {report_type}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    row = 3
    for key, value in report_data.items():
        ws[f'A{row}'] = str(key)
        ws[f'B{row}'] = str(value)
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_report_to_pdf(report_data, report_type):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph(f'<b>Reporte: {report_type}</b>', styles['Title'])
    elements.append(title)
    
    data = [['Campo', 'Valor']]
    for key, value in report_data.items():
        data.append([str(key), str(value)])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output


def generate_dashboard_summary(user=None):
    """Generate consolidated dashboard summary for LEAD portal widgets"""
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    cotizaciones_qs = LeadCotizacion.objects.all()
    shipments_qs = Shipment.objects.all()
    pre_liquidations_qs = PreLiquidation.objects.all()
    
    if user:
        cotizaciones_qs = cotizaciones_qs.filter(lead_user=user)
        shipments_qs = shipments_qs.filter(cotizacion__lead_user=user)
        pre_liquidations_qs = pre_liquidations_qs.filter(cotizacion__lead_user=user)
    
    cotizaciones_por_estado = cotizaciones_qs.values('estado').annotate(count=Count('id'))
    cotizaciones_dict = {item['estado']: item['count'] for item in cotizaciones_por_estado}
    
    shipments_por_estado = shipments_qs.values('current_status').annotate(count=Count('id'))
    shipments_dict = {item['current_status']: item['count'] for item in shipments_por_estado}
    
    total_tributos = pre_liquidations_qs.filter(
        is_confirmed=True
    ).aggregate(
        total=Sum('total_tributos_usd')
    )['total'] or Decimal('0')
    
    valor_total_cif = pre_liquidations_qs.filter(
        is_confirmed=True
    ).aggregate(
        total=Sum('cif_value_usd')
    )['total'] or Decimal('0')
    
    cotizaciones_mes = cotizaciones_qs.filter(
        fecha_creacion__date__gte=month_start
    ).count()
    
    valor_cotizado_mes = cotizaciones_qs.filter(
        fecha_creacion__date__gte=month_start,
        total_usd__isnull=False
    ).aggregate(total=Sum('total_usd'))['total'] or Decimal('0')
    
    embarques_activos = shipments_qs.exclude(
        current_status__in=['entregado', 'cancelado']
    ).count()
    
    return {
        'resumen': {
            'cotizaciones_pendientes': cotizaciones_dict.get('pendiente', 0),
            'cotizaciones_aprobadas': cotizaciones_dict.get('aprobada', 0),
            'embarques_en_transito': shipments_dict.get('en_transito', 0),
            'embarques_en_aduana': shipments_dict.get('en_aduana', 0),
            'embarques_activos': embarques_activos,
        },
        'metricas_mes': {
            'cotizaciones_nuevas': cotizaciones_mes,
            'valor_cotizado_usd': float(valor_cotizado_mes),
        },
        'financiero': {
            'total_tributos_pagados_usd': float(total_tributos),
            'valor_total_cif_usd': float(valor_total_cif),
        },
        'cotizaciones_por_estado': cotizaciones_dict,
        'embarques_por_estado': shipments_dict,
    }


def generate_cost_analytics_report(start_date=None, end_date=None, user=None):
    """Detailed cost analytics by route and period"""
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    cotizaciones_qs = LeadCotizacion.objects.filter(
        fecha_creacion__range=[start_date, end_date]
    )
    if user:
        cotizaciones_qs = cotizaciones_qs.filter(lead_user=user)
    
    costos_por_ruta = cotizaciones_qs.values(
        'origen_pais', 'destino_ciudad'
    ).annotate(
        total_cotizaciones=Count('id'),
        valor_promedio=Avg('total_usd'),
        valor_total=Sum('total_usd'),
        peso_promedio=Avg('peso_kg'),
    ).order_by('-total_cotizaciones')[:10]
    
    costos_por_transporte = cotizaciones_qs.values(
        'tipo_carga'
    ).annotate(
        total=Count('id'),
        valor_promedio=Avg('total_usd'),
        valor_total=Sum('total_usd'),
    )
    
    line_items = QuoteLineItem.objects.filter(
        scenario__cotizacion__fecha_creacion__range=[start_date, end_date]
    )
    if user:
        line_items = line_items.filter(scenario__cotizacion__lead_user=user)
    
    desglose_costos = line_items.values('category').annotate(
        total=Sum('amount_usd'),
        promedio=Avg('amount_usd'),
        count=Count('id')
    ).order_by('-total')
    
    pre_liquidations = PreLiquidation.objects.filter(
        created_at__range=[start_date, end_date],
        is_confirmed=True
    )
    if user:
        pre_liquidations = pre_liquidations.filter(cotizacion__lead_user=user)
    
    tributos_desglose = {
        'ad_valorem': float(pre_liquidations.aggregate(t=Sum('ad_valorem_usd'))['t'] or 0),
        'fodinfa': float(pre_liquidations.aggregate(t=Sum('fodinfa_usd'))['t'] or 0),
        'ice': float(pre_liquidations.aggregate(t=Sum('ice_usd'))['t'] or 0),
        'salvaguardia': float(pre_liquidations.aggregate(t=Sum('salvaguardia_usd'))['t'] or 0),
        'iva': float(pre_liquidations.aggregate(t=Sum('iva_usd'))['t'] or 0),
        'total': float(pre_liquidations.aggregate(t=Sum('total_tributos_usd'))['t'] or 0),
    }
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date),
            'end_date': end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date),
        },
        'rutas_principales': [
            {
                'origen': r['origen_pais'],
                'destino': r['destino_ciudad'],
                'cotizaciones': r['total_cotizaciones'],
                'valor_promedio_usd': float(r['valor_promedio'] or 0),
                'valor_total_usd': float(r['valor_total'] or 0),
                'peso_promedio_kg': float(r['peso_promedio'] or 0),
            } for r in costos_por_ruta
        ],
        'por_tipo_transporte': [
            {
                'tipo': c['tipo_carga'],
                'total': c['total'],
                'valor_promedio_usd': float(c['valor_promedio'] or 0),
                'valor_total_usd': float(c['valor_total'] or 0),
            } for c in costos_por_transporte
        ],
        'desglose_costos': [
            {
                'categoria': d['category'],
                'total_usd': float(d['total'] or 0),
                'promedio_usd': float(d['promedio'] or 0),
                'transacciones': d['count'],
            } for d in desglose_costos
        ],
        'tributos_aduaneros': tributos_desglose,
    }


def generate_operational_kpis(start_date=None, end_date=None, user=None):
    """Operational KPIs: transit times, customs processing, delivery performance"""
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=90)
    
    shipments_qs = Shipment.objects.filter(
        created_at__range=[start_date, end_date]
    )
    if user:
        shipments_qs = shipments_qs.filter(cotizacion__lead_user=user)
    
    completed_shipments = shipments_qs.filter(
        current_status='entregado',
        actual_delivery__isnull=False,
        actual_departure__isnull=False
    )
    
    transit_times = []
    for ship in completed_shipments:
        if ship.actual_delivery and ship.actual_departure:
            days = (ship.actual_delivery - ship.actual_departure).days
            transit_times.append({
                'transport_type': ship.transport_type,
                'days': days
            })
    
    avg_transit_by_type = {}
    for t in ['aereo', 'maritimo', 'terrestre']:
        times = [x['days'] for x in transit_times if x['transport_type'] == t]
        if times:
            avg_transit_by_type[t] = sum(times) / len(times)
    
    on_time_deliveries = completed_shipments.filter(
        actual_delivery__lte=F('estimated_delivery')
    ).count()
    
    total_completed = completed_shipments.count()
    on_time_rate = (on_time_deliveries / total_completed * 100) if total_completed > 0 else 0
    
    status_counts = shipments_qs.values('current_status').annotate(count=Count('id'))
    
    by_transport = shipments_qs.values('transport_type').annotate(
        total=Count('id'),
        entregados=Count('id', filter=Q(current_status='entregado')),
        en_transito=Count('id', filter=Q(current_status='en_transito')),
    )
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date),
            'end_date': end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date),
        },
        'kpis': {
            'total_embarques': shipments_qs.count(),
            'embarques_completados': total_completed,
            'tasa_entrega_a_tiempo': round(on_time_rate, 1),
            'tiempo_transito_promedio_dias': avg_transit_by_type,
        },
        'por_estado': list(status_counts),
        'por_transporte': list(by_transport),
    }


def generate_import_trends_report(start_date=None, end_date=None, user=None):
    """Import trends by category, origin, and time period"""
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=180)
    
    pre_liq_qs = PreLiquidation.objects.filter(
        created_at__range=[start_date, end_date],
        is_confirmed=True
    )
    if user:
        pre_liq_qs = pre_liq_qs.filter(cotizacion__lead_user=user)
    
    by_hs_chapter = {}
    for pl in pre_liq_qs:
        if pl.confirmed_hs_code:
            chapter = pl.confirmed_hs_code[:2]
            if chapter not in by_hs_chapter:
                by_hs_chapter[chapter] = {
                    'count': 0,
                    'fob_total': Decimal('0'),
                    'cif_total': Decimal('0'),
                    'tributos_total': Decimal('0')
                }
            by_hs_chapter[chapter]['count'] += 1
            by_hs_chapter[chapter]['fob_total'] += pl.fob_value_usd
            by_hs_chapter[chapter]['cif_total'] += pl.cif_value_usd
            by_hs_chapter[chapter]['tributos_total'] += pl.total_tributos_usd
    
    hs_chapter_names = {
        '84': 'Maquinaria y equipos',
        '85': 'Electrónica y equipos eléctricos',
        '87': 'Vehículos y partes',
        '61': 'Prendas de vestir (tejido)',
        '62': 'Prendas de vestir (no tejido)',
        '39': 'Plásticos',
        '73': 'Manufactura de hierro/acero',
        '94': 'Muebles',
        '90': 'Instrumentos ópticos/médicos',
        '38': 'Productos químicos diversos',
    }
    
    trends_by_chapter = [
        {
            'capitulo_hs': ch,
            'descripcion': hs_chapter_names.get(ch, f'Capítulo {ch}'),
            'importaciones': data['count'],
            'fob_total_usd': float(data['fob_total']),
            'cif_total_usd': float(data['cif_total']),
            'tributos_total_usd': float(data['tributos_total']),
        }
        for ch, data in sorted(by_hs_chapter.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
    ]
    
    cotizaciones = LeadCotizacion.objects.filter(
        fecha_creacion__range=[start_date, end_date]
    )
    if user:
        cotizaciones = cotizaciones.filter(lead_user=user)
    
    by_origin = cotizaciones.values('origen_pais').annotate(
        count=Count('id'),
        valor_total=Sum('total_usd')
    ).order_by('-count')[:10]
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date),
            'end_date': end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date),
        },
        'por_categoria_hs': trends_by_chapter,
        'por_pais_origen': [
            {
                'pais': o['origen_pais'],
                'importaciones': o['count'],
                'valor_total_usd': float(o['valor_total'] or 0),
            } for o in by_origin
        ],
        'total_importaciones': pre_liq_qs.count(),
        'valor_total_fob_usd': float(pre_liq_qs.aggregate(t=Sum('fob_value_usd'))['t'] or 0),
        'valor_total_cif_usd': float(pre_liq_qs.aggregate(t=Sum('cif_value_usd'))['t'] or 0),
    }


def generate_financial_summary(start_date=None, end_date=None, user=None):
    """Financial summary with duties paid and cost breakdown"""
    if not end_date:
        end_date = timezone.now()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    pre_liq_qs = PreLiquidation.objects.filter(
        created_at__range=[start_date, end_date],
        is_confirmed=True
    )
    if user:
        pre_liq_qs = pre_liq_qs.filter(cotizacion__lead_user=user)
    
    tributos_agg = pre_liq_qs.aggregate(
        ad_valorem=Sum('ad_valorem_usd'),
        fodinfa=Sum('fodinfa_usd'),
        ice=Sum('ice_usd'),
        salvaguardia=Sum('salvaguardia_usd'),
        iva=Sum('iva_usd'),
        total_tributos=Sum('total_tributos_usd'),
        total_fob=Sum('fob_value_usd'),
        total_cif=Sum('cif_value_usd'),
    )
    
    cotizaciones = LeadCotizacion.objects.filter(
        fecha_creacion__range=[start_date, end_date]
    )
    if user:
        cotizaciones = cotizaciones.filter(lead_user=user)
    
    cotizaciones_agg = cotizaciones.aggregate(
        total_cotizado=Sum('total_usd'),
        cotizaciones_count=Count('id'),
        aprobadas=Count('id', filter=Q(estado='aprobada')),
    )
    
    scenarios = QuoteScenario.objects.filter(
        cotizacion__fecha_creacion__range=[start_date, end_date],
        is_selected=True
    )
    if user:
        scenarios = scenarios.filter(cotizacion__lead_user=user)
    
    line_items_agg = QuoteLineItem.objects.filter(
        scenario__in=scenarios
    ).values('category').annotate(
        total=Sum('amount_usd')
    ).order_by('-total')
    
    return {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date),
            'end_date': end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date),
        },
        'resumen_tributos': {
            'ad_valorem_usd': float(tributos_agg['ad_valorem'] or 0),
            'fodinfa_usd': float(tributos_agg['fodinfa'] or 0),
            'ice_usd': float(tributos_agg['ice'] or 0),
            'salvaguardia_usd': float(tributos_agg['salvaguardia'] or 0),
            'iva_usd': float(tributos_agg['iva'] or 0),
            'total_tributos_usd': float(tributos_agg['total_tributos'] or 0),
        },
        'resumen_valores': {
            'total_fob_usd': float(tributos_agg['total_fob'] or 0),
            'total_cif_usd': float(tributos_agg['total_cif'] or 0),
        },
        'cotizaciones': {
            'total': cotizaciones_agg['cotizaciones_count'] or 0,
            'aprobadas': cotizaciones_agg['aprobadas'] or 0,
            'valor_total_usd': float(cotizaciones_agg['total_cotizado'] or 0),
        },
        'desglose_costos': [
            {
                'categoria': item['category'],
                'total_usd': float(item['total'] or 0),
            } for item in line_items_agg
        ],
        'pre_liquidaciones_confirmadas': pre_liq_qs.count(),
    }
